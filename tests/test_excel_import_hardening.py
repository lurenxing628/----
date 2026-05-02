from __future__ import annotations

import importlib
import io
import json
import os
import re
import sys
from html import unescape
from pathlib import Path

import openpyxl
import pytest
from flask import g

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from core.infrastructure.database import ensure_schema, get_connection
from core.infrastructure.errors import ErrorCode, ValidationError
from core.services.common.excel_service import ImportMode, ImportPreviewRow, RowStatus
from core.services.common.excel_templates import build_xlsx_bytes
from core.services.common.excel_validators import (
    get_batch_row_validate_and_normalize,
    get_operator_calendar_row_validate_and_normalize,
)
from core.services.common.normalize import is_blank_value
from core.services.process.op_type_excel_import_service import OpTypeExcelImportService
from core.services.scheduler.batch_service import BatchService
from web.routes.excel_utils import ensure_unique_ids


def _new_conn(tmp_path) -> tuple:
    db_path = tmp_path / "aps_test.db"
    ensure_schema(str(db_path), logger=None, schema_path=os.path.join(str(REPO_ROOT), "schema.sql"), backup_dir=None)
    return get_connection(str(db_path)), str(db_path)


def _build_app(tmp_path, monkeypatch):
    test_db = tmp_path / "aps_test.db"
    test_logs = tmp_path / "logs"
    test_backups = tmp_path / "backups"
    test_templates = tmp_path / "templates_excel"
    test_logs.mkdir(exist_ok=True)
    test_backups.mkdir(exist_ok=True)
    test_templates.mkdir(exist_ok=True)

    monkeypatch.setenv("APS_ENV", "development")
    monkeypatch.setenv("APS_DB_PATH", str(test_db))
    monkeypatch.setenv("APS_LOG_DIR", str(test_logs))
    monkeypatch.setenv("APS_BACKUP_DIR", str(test_backups))
    monkeypatch.setenv("APS_EXCEL_TEMPLATE_DIR", str(test_templates))

    ensure_schema(str(test_db), logger=None, schema_path=os.path.join(str(REPO_ROOT), "schema.sql"), backup_dir=None)
    app_mod = importlib.import_module("app")
    return app_mod.create_app(), str(test_db)


def _make_xlsx(headers, rows) -> bytes:
    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.append(list(headers))
        for row in rows:
            ws.append(list(row))
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    finally:
        wb.close()


def _extract_raw_rows_json(html: str) -> str:
    m = re.search(r'<textarea name="raw_rows_json"[^>]*>(.*?)</textarea>', html, re.S)
    if not m:
        raise RuntimeError("未能从页面提取 raw_rows_json")
    return unescape(m.group(1)).strip()


def _extract_hidden_input(html: str, name: str) -> str:
    for m in re.finditer(r"<input[^>]+>", html, re.I):
        tag = m.group(0)
        if re.search(rf'name="{re.escape(name)}"', tag):
            vm = re.search(r'value="([^"]*)"', tag)
            return unescape(vm.group(1)).strip() if vm else ""
    raise RuntimeError(f"未能从页面提取隐藏字段：{name}")


def test_batch_quantity_float_is_rejected_without_truncation(tmp_path) -> None:
    conn, _db_path = _new_conn(tmp_path)
    try:
        validator = get_batch_row_validate_and_normalize(conn, parts_cache={"P001": object()}, inplace=True)
        row = {"批次号": "B001", "图号": "P001", "数量": 100.5}
        assert validator(row) == "“数量”必须是整数"

        with pytest.raises(ValidationError, match="必须是整数"):
            BatchService._normalize_int(100.5, field="数量", allow_none=False)
    finally:
        conn.close()


def test_batch_validator_accepts_parts_cache_without_conn() -> None:
    validator = get_batch_row_validate_and_normalize(
        parts_cache={"P001": object()},
        inplace=True,
    )
    row = {
        "批次号": "B001",
        "图号": "P001",
        "数量": "2",
        "优先级": "normal",
        "齐套": "yes",
    }

    assert validator(row) is None
    assert row["数量"] == 2


def test_batch_validator_requires_conn_when_parts_cache_missing() -> None:
    with pytest.raises(ValueError, match="缺少 conn 或 parts_cache"):
        get_batch_row_validate_and_normalize(inplace=True)


def test_blank_helper_does_not_treat_zero_as_blank() -> None:
    assert is_blank_value(0) is False
    assert is_blank_value("0") is False
    assert is_blank_value(None) is True
    assert is_blank_value("   ") is True


def test_import_service_does_not_treat_zero_id_as_blank(tmp_path) -> None:
    conn, _db_path = _new_conn(tmp_path)
    try:
        service = OpTypeExcelImportService(conn)
        stats = service.apply_preview_rows(
            [
                ImportPreviewRow(
                    row_num=2,
                    status=RowStatus.NEW,
                    data={"工种ID": 0, "工种名称": "零号工种", "归属": "internal"},
                    message="新增",
                )
            ],
            mode=ImportMode.OVERWRITE,
            existing_ids=set(),
        )
        row = conn.execute("SELECT op_type_id, name, category FROM OpTypes WHERE op_type_id = ?", ("0",)).fetchone()
        assert row is not None
        assert row["name"] == "零号工种"
        assert row["category"] == "internal"
        assert stats["new_count"] == 1
    finally:
        conn.close()


def test_ensure_unique_ids_detects_integer_like_float_duplicates() -> None:
    with pytest.raises(ValidationError, match="存在重复的“id”"):
        ensure_unique_ids([{"id": 1}, {"id": 1.0}], id_column="id")


def test_operator_calendar_preview_fallback_trims_time_suffix(tmp_path, monkeypatch) -> None:
    from core.services.common.excel_backend_factory import get_excel_backend
    from web.bootstrap.request_services import RequestServices

    app, db_path = _build_app(tmp_path, monkeypatch)
    conn = get_connection(db_path)
    conn.execute("INSERT INTO Operators (operator_id, name, status) VALUES (?, ?, ?)", ("OP001", "张三", "active"))
    conn.commit()
    conn.close()

    route_mod = importlib.import_module("web.routes.personnel_excel_operator_calendar")
    captured = {}

    def _fake_render_template(_template_name, **kwargs):
        captured.update(kwargs)
        return kwargs

    monkeypatch.setattr(route_mod, "render_template", _fake_render_template)

    file_bytes = _make_xlsx(
        ["工号", "日期", "类型", "班次开始", "班次结束", "可用工时", "效率", "允许普通件", "允许急件", "说明"],
        [["OP001", "2026-01-25 00:00:00", "holiday", "08:00", "", 0, 0.8, "no", "no", "示例"]],
    )

    with app.test_request_context(
        "/personnel/excel/operator_calendar/preview",
        method="POST",
        data={"mode": ImportMode.OVERWRITE.value, "file": (io.BytesIO(file_bytes), "calendar.xlsx")},
        content_type="multipart/form-data",
    ):
        req_conn = get_connection(db_path)
        try:
            g.app_logger = app.logger
            g.db = req_conn
            g.op_logger = None
            g.services = RequestServices(
                db=req_conn,
                app_logger=app.logger,
                op_logger=None,
                get_excel_backend=get_excel_backend,
            )
            route_mod.excel_operator_calendar_preview()
        finally:
            req_conn.close()

    rows = json.loads(captured["raw_rows_json"])
    assert rows[0]["日期"] == "2026-01-25"
    assert rows[0]["__id"] == "OP001|2026-01-25"


def test_operator_calendar_validator_rejects_bool_and_nonfinite_numeric_inputs(tmp_path) -> None:
    conn, _db_path = _new_conn(tmp_path)
    try:
        conn.execute("INSERT INTO Operators (operator_id, name, status) VALUES (?, ?, ?)", ("OP001", "张三", "active"))
        conn.commit()

        validator = get_operator_calendar_row_validate_and_normalize(
            conn,
            holiday_default_efficiency=0.8,
            inplace=True,
        )
        base_row = {
            "工号": "OP001",
            "日期": "2026-04-03",
            "类型": "workday",
            "班次开始": "08:00",
            "班次结束": "",
            "可用工时": 8,
            "效率": 1.0,
            "允许普通件": "yes",
            "允许急件": "yes",
            "说明": "strict-number",
        }
        cases = [
            ("可用工时", True, "“可用工时”必须是数字"),
            ("效率", False, "“效率”必须是数字"),
            ("可用工时", float("nan"), "“可用工时”必须是有限数字"),
            ("效率", float("nan"), "“效率”必须是有限数字"),
            ("可用工时", float("inf"), "“可用工时”必须是有限数字"),
            ("效率", float("inf"), "“效率”必须是有限数字"),
            ("可用工时", float("-inf"), "“可用工时”必须是有限数字"),
            ("效率", float("-inf"), "“效率”必须是有限数字"),
        ]

        for field, value, expected in cases:
            row = dict(base_row)
            row[field] = value
            assert validator(row) == expected
    finally:
        conn.close()


def test_operator_calendar_preview_and_confirm_reject_bool_numeric_cells(tmp_path, monkeypatch) -> None:
    app, db_path = _build_app(tmp_path, monkeypatch)
    conn = get_connection(db_path)
    conn.execute("INSERT INTO Operators (operator_id, name, status) VALUES (?, ?, ?)", ("OP001", "张三", "active"))
    conn.commit()
    conn.close()

    client = app.test_client()
    file_bytes = _make_xlsx(
        ["工号", "日期", "类型", "班次开始", "班次结束", "可用工时", "效率", "允许普通件", "允许急件", "说明"],
        [["OP001", "2026-04-04", "workday", "08:00", "", True, 1.0, "yes", "yes", "bool-cell"]],
    )

    preview_resp = client.post(
        "/personnel/excel/operator_calendar/preview",
        data={"mode": ImportMode.OVERWRITE.value, "file": (io.BytesIO(file_bytes), "operator_calendar.xlsx")},
        content_type="multipart/form-data",
    )
    preview_html = preview_resp.get_data(as_text=True)
    assert preview_resp.status_code == 200
    assert "“可用工时”必须是数字" in preview_html

    raw_rows_json = _extract_raw_rows_json(preview_html)
    preview_baseline = _extract_hidden_input(preview_html, "preview_baseline")
    assert preview_baseline

    confirm_resp = client.post(
        "/personnel/excel/operator_calendar/confirm",
        data={
            "mode": ImportMode.OVERWRITE.value,
            "filename": "operator_calendar.xlsx",
            "raw_rows_json": raw_rows_json,
            "preview_baseline": preview_baseline,
        },
        follow_redirects=True,
    )
    confirm_html = confirm_resp.get_data(as_text=True)
    assert confirm_resp.status_code == 200
    assert "导入被拒绝：Excel 存在 1 行错误。" in confirm_html
    assert "“可用工时”必须是数字" in confirm_html

    verify_conn = get_connection(db_path)
    try:
        count = verify_conn.execute("SELECT COUNT(1) FROM OperatorCalendar WHERE operator_id=? AND date=?", ("OP001", "2026-04-04")).fetchone()[0]
        assert int(count) == 0
    finally:
        verify_conn.close()


def test_upload_over_limit_returns_413(tmp_path, monkeypatch) -> None:
    app, _db_path = _build_app(tmp_path, monkeypatch)
    client = app.test_client()

    big_file = io.BytesIO(b"x" * (17 * 1024 * 1024))
    resp = client.post(
        "/personnel/excel/operators/preview",
        data={"mode": ImportMode.OVERWRITE.value, "file": (big_file, "too-large.xlsx")},
        content_type="multipart/form-data",
    )

    assert resp.status_code == 413
    assert "16MB" in resp.get_data(as_text=True)


def test_exact_file_limit_is_not_rejected_by_multipart_overhead(tmp_path, monkeypatch) -> None:
    app, _db_path = _build_app(tmp_path, monkeypatch)
    file_bytes = _make_xlsx(
        ["工号", "姓名", "状态", "班组", "备注"],
        [["OP001", "张三", "active", None, "ok"]],
    )
    app.config["EXCEL_MAX_UPLOAD_BYTES"] = len(file_bytes)
    app.config["MAX_CONTENT_LENGTH"] = len(file_bytes) + 1024
    client = app.test_client()

    resp = client.post(
        "/personnel/excel/operators/preview",
        data={"mode": ImportMode.OVERWRITE.value, "file": (io.BytesIO(file_bytes), "operators.xlsx")},
        content_type="multipart/form-data",
    )

    assert resp.status_code == 200
    assert "上传文件超过" not in resp.get_data(as_text=True)


def test_op_type_preview_and_confirm_reject_duplicate_name_conflict(tmp_path, monkeypatch) -> None:
    app, db_path = _build_app(tmp_path, monkeypatch)
    conn = get_connection(db_path)
    conn.execute(
        "INSERT INTO OpTypes (op_type_id, name, category) VALUES (?, ?, ?)",
        ("OT001", "数车", "internal"),
    )
    conn.commit()
    conn.close()

    client = app.test_client()
    file_bytes = _make_xlsx(
        ["工种ID", "工种名称", "归属"],
        [["OT002", "数车", "internal"]],
    )

    preview_resp = client.post(
        "/process/excel/op-types/preview",
        data={"mode": ImportMode.OVERWRITE.value, "file": (io.BytesIO(file_bytes), "op_types.xlsx")},
        content_type="multipart/form-data",
    )

    preview_html = preview_resp.get_data(as_text=True)
    assert preview_resp.status_code == 200
    assert "工种名称“数车”已被工种编号“OT001”使用，名称不能重复。" in preview_html

    raw_rows_json = _extract_raw_rows_json(preview_html)
    preview_baseline = _extract_hidden_input(preview_html, "preview_baseline")
    assert preview_baseline

    confirm_resp = client.post(
        "/process/excel/op-types/confirm",
        data={
            "mode": ImportMode.OVERWRITE.value,
            "filename": "op_types.xlsx",
            "raw_rows_json": raw_rows_json,
            "preview_baseline": preview_baseline,
        },
        follow_redirects=True,
    )

    confirm_html = confirm_resp.get_data(as_text=True)
    assert confirm_resp.status_code == 200
    assert "导入被拒绝：Excel 存在 1 行错误。" in confirm_html
    assert 'data-flash="success"' not in confirm_html

    verify_conn = get_connection(db_path)
    try:
        rows = verify_conn.execute("SELECT op_type_id, name FROM OpTypes ORDER BY op_type_id").fetchall()
        assert [(row["op_type_id"], row["name"]) for row in rows] == [("OT001", "数车")]
    finally:
        verify_conn.close()


def test_file_body_over_limit_returns_file_too_large_error(tmp_path, monkeypatch) -> None:
    app, _db_path = _build_app(tmp_path, monkeypatch)
    app.config["EXCEL_MAX_UPLOAD_BYTES"] = 1 * 1024 * 1024
    app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024
    client = app.test_client()

    resp = client.post(
        "/personnel/excel/operators/preview",
        data={
            "mode": ImportMode.OVERWRITE.value,
            "file": (io.BytesIO(b"x" * (app.config["EXCEL_MAX_UPLOAD_BYTES"] + 1)), "too-large.xlsx"),
        },
        content_type="multipart/form-data",
    )

    body = resp.get_data(as_text=True)
    assert resp.status_code == 413
    assert ErrorCode.FILE_TOO_LARGE.value in body
    assert "上传文件超过 1MB" in body


def test_scheduler_calendar_preview_rejects_duplicate_dates_after_canonicalization(tmp_path, monkeypatch) -> None:
    app, _db_path = _build_app(tmp_path, monkeypatch)
    client = app.test_client()
    file_bytes = _make_xlsx(
        ["日期", "类型", "可用工时", "效率", "允许普通件", "允许急件", "说明"],
        [
            ["2026-1-21", "workday", 8, 1.0, "yes", "yes", "row1"],
            ["2026/01/21", "holiday", 0, 0.8, "no", "no", "row2"],
        ],
    )

    resp = client.post(
        "/scheduler/excel/calendar/preview",
        data={"mode": ImportMode.OVERWRITE.value, "file": (io.BytesIO(file_bytes), "calendar.xlsx")},
        content_type="multipart/form-data",
    )

    body = resp.get_data(as_text=True)
    assert resp.status_code == 400
    assert "存在重复的“日期”" in body
    assert "2026-01-21" in body


def test_build_xlsx_bytes_sanitizes_formula_like_strings() -> None:
    output = build_xlsx_bytes(
        ["备注"],
        [["=cmd|' /C calc'!A0"], ["+SUM(1,1)"], ["-1"], ["@A1"]],
        sanitize_formula=True,
    )
    wb = openpyxl.load_workbook(output)
    try:
        ws = wb.active
        assert ws is not None
        assert ws["A2"].value == "'=cmd|' /C calc'!A0"
        assert ws["A3"].value == "'+SUM(1,1)"
        assert ws["A4"].value == "'-1"
        assert ws["A5"].value == "'@A1"
        assert ws.freeze_panes == "A2"
    finally:
        wb.close()
