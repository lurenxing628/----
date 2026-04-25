from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]


def _read(rel_path: str) -> str:
    return (REPO_ROOT / rel_path).read_text(encoding="utf-8")


def test_scheduler_config_and_batch_hints_are_user_facing_chinese() -> None:
    expected_holiday_hint = "假期也安排生产且未单独填写效率时，系统会使用这里的效率值；请输入大于 0 的数字。"
    expected_batch_manage_hint = (
        "只在零件还没有工序模板、需要根据工艺路线自动生成模板时起作用："
        "如果遇到系统不认识的工种、找不到供应商、或供应商默认周期无效，将直接报错并拒绝创建。"
    )
    expected_batch_schedule_hint = (
        "仅对“派工方式”“智能派工策略”“自动分配设备人员”这几项设置生效："
        "开启后，如果填写了无效的值，将直接报错并拒绝排产，不会自动使用默认值替代。"
    )

    for rel_path in ("templates/scheduler/config.html", "web_new_test/templates/scheduler/config.html"):
        source = _read(rel_path)
        assert expected_holiday_hint in source
        assert "假期安排生产但未单独设置效率时" not in source
        assert "&gt;0" not in source

    for rel_path in (
        "templates/scheduler/batches_manage.html",
        "web_new_test/templates/scheduler/batches_manage.html",
    ):
        source = _read(rel_path)
        assert expected_batch_manage_hint in source
        assert "解析器不支持 strict_mode" not in source

    for rel_path in ("templates/scheduler/batches.html", "web_new_test/templates/scheduler/batches.html"):
        source = _read(rel_path)
        assert expected_batch_schedule_hint in source
        assert "dispatch_mode / dispatch_rule / auto_assign_enabled" not in source
        assert "设了截止日期的话，排不完会提示失败。" not in source


def test_scheduler_analysis_gantt_and_logs_do_not_surface_internal_terms() -> None:
    analysis = _read("templates/scheduler/analysis.html")
    assert "dispatch_mode_zh" in analysis
    assert "dispatch_rule_zh" in analysis
    assert "attempts / 优化曲线 / 超期明细" not in analysis
    assert "r.dispatch_mode }}/{{ r.dispatch_rule" not in analysis

    for rel_path in ("templates/scheduler/gantt.html", "web_new_test/templates/scheduler/gantt.html"):
        source = _read(rel_path)
        assert "排程数据" in source
        assert "Schedule 数据" not in source

    logs = _read("templates/system/logs.html")
    assert "按英文值筛选" not in logs
    assert "可输入模块名称或代码" in logs
    assert "可输入动作名称或代码" in logs
    assert "<code>{{ r.module }}</code>" not in logs
    assert "<code>{{ r.action }}</code>" not in logs
    assert "{{ r.module_label }}" in logs
    assert "{{ r.action_label }}" in logs
    assert 'title="{{ r.module }}"' not in logs
    assert 'title="{{ r.action }}"' not in logs
    assert 'title="{{ r.target_type }}"' not in logs


def test_debug_details_do_not_expose_flask_endpoint_names_to_users() -> None:
    rel_paths = (
        "templates/scheduler/config.html",
        "web_new_test/templates/scheduler/config.html",
        "templates/scheduler/batches.html",
        "web_new_test/templates/scheduler/batches.html",
        "templates/personnel/list.html",
        "templates/personnel/detail.html",
        "templates/personnel/calendar.html",
    )
    for rel_path in rel_paths:
        source = _read(rel_path)
        assert "后端接口未注册（endpoint" not in source
        assert "endpoint）" not in source
        assert "missing_preset_endpoints|join" not in source


def test_process_excel_current_tables_render_chinese_display_fields() -> None:
    op_types = _read("templates/process/excel_import_op_types.html")
    assert 'r["归属显示"]' in op_types
    assert '<td>{{ r["归属"] }}</td>' not in op_types

    part_operation_hours = _read("templates/process/excel_import_part_operation_hours.html")
    assert 'r["归属显示"]' in part_operation_hours
    assert '<td>{{ r["归属"] }}</td>' not in part_operation_hours

    suppliers = _read("templates/process/excel_import_suppliers.html")
    assert 'r["状态显示"]' in suppliers
    assert '<td>{{ r["状态"] }}</td>' not in suppliers

    op_type_service = _read("core/services/process/op_type_service.py")
    assert '"归属显示": source_type_label(ot.category)' in op_type_service

    supplier_service = _read("core/services/process/supplier_service.py")
    assert '"状态显示": supplier_status_label(s.status)' in supplier_service

    part_operation_hours_route = _read("web/routes/process_excel_part_operation_hours.py")
    assert '"归属显示": source_type_label(source)' in part_operation_hours_route


def test_manuals_keep_backend_supported_english_aliases_but_mark_them_as_compatible() -> None:
    from core.models.enums import (
        BatchPriority,
        CalendarDayType,
        MachineStatus,
        OperatorStatus,
        ReadyStatus,
        SourceType,
        SupplierStatus,
        YesNo,
    )
    from core.services.common.enum_normalizers import (
        normalize_machine_status,
        normalize_op_type_category,
        normalize_operator_status,
        normalize_skill_level,
        normalize_supplier_status,
        normalize_yesno_narrow,
    )
    from core.services.common.normalization_matrix import (
        normalize_batch_priority_value,
        normalize_calendar_day_type_value,
        normalize_ready_status_value,
    )

    manual_sources = "\n".join(
        _read(rel_path)
        for rel_path in (
            "web/viewmodels/page_manuals_process.py",
            "web/viewmodels/page_manuals_scheduler.py",
            "web/viewmodels/page_manuals_personnel.py",
            "web/viewmodels/page_manuals_equipment.py",
            "static/docs/scheduler_manual.md",
        )
    )

    expected_alias_phrases = (
        "兼容英文标准值 `active`/`inactive`",
        "兼容英文标准值 `active`/`inactive`/`maintain`",
        "兼容英文标准值 `internal`/`external`",
        "兼容英文标准值 `beginner`/`normal`/`expert`",
        "兼容英文标准值 `yes`/`no`",
        "兼容英文标准值 `yes`/`no`、`true`/`false`、`on`/`off`",
        "兼容英文标准值 `normal`/`urgent`/`critical`",
        "兼容英文标准值 `yes`/`no`/`partial`",
        "兼容英文标准值 `workday`/`holiday`",
        "兼容英文标准值 `true`/`false`",
    )
    for phrase in expected_alias_phrases:
        assert phrase in manual_sources

    assert normalize_operator_status("active") == OperatorStatus.ACTIVE.value
    assert normalize_operator_status("inactive") == OperatorStatus.INACTIVE.value
    assert normalize_machine_status("maintain") == MachineStatus.MAINTAIN.value
    assert normalize_supplier_status("active") == SupplierStatus.ACTIVE.value
    assert normalize_supplier_status("inactive") == SupplierStatus.INACTIVE.value
    assert normalize_op_type_category("internal") == SourceType.INTERNAL.value
    assert normalize_op_type_category("external") == SourceType.EXTERNAL.value
    assert normalize_skill_level("beginner") == "beginner"
    assert normalize_skill_level("normal") == "normal"
    assert normalize_skill_level("expert") == "expert"
    assert normalize_yesno_narrow("yes") == YesNo.YES.value
    assert normalize_yesno_narrow("no") == YesNo.NO.value
    from core.services.personnel.operator_machine_normalizers import normalize_yes_no_optional

    assert normalize_yes_no_optional("true", field="主操设备") == YesNo.YES.value
    assert normalize_yes_no_optional("false", field="主操设备") == YesNo.NO.value
    assert normalize_yes_no_optional("on", field="主操设备") == YesNo.YES.value
    assert normalize_yes_no_optional("off", field="主操设备") == YesNo.NO.value
    assert normalize_batch_priority_value("urgent") == BatchPriority.URGENT.value
    assert normalize_batch_priority_value("critical") == BatchPriority.CRITICAL.value
    assert normalize_ready_status_value("partial") == ReadyStatus.PARTIAL.value
    assert normalize_calendar_day_type_value("workday") == CalendarDayType.WORKDAY.value
    assert normalize_calendar_day_type_value("holiday") == CalendarDayType.HOLIDAY.value
    assert normalize_calendar_day_type_value("weekend") == CalendarDayType.HOLIDAY.value

    static_manual = _read("static/docs/scheduler_manual.md")
    assert "根据工艺路线字符串自动补建模板" in static_manual
    assert "route_raw 自动补建模板" not in static_manual
    assert "排产方案、智能派工策略、自动分配设备人员" in static_manual
    assert "dispatch_mode / dispatch_rule / auto_assign_enabled" not in static_manual


def test_supplier_manual_matches_required_default_days_and_template_columns() -> None:
    manual_sources = "\n".join(
        _read(rel_path)
        for rel_path in (
            "web/viewmodels/page_manuals_process.py",
            "static/docs/scheduler_manual.md",
        )
    )
    assert "默认周期填正数(天)，必须填写" in manual_sources
    assert "默认周期` 填大于 0 的天数；不能留空" in manual_sources
    assert "下载的模板包含“状态”和“备注”两列" in manual_sources
    assert "模板只有4列" not in manual_sources
    assert "不填默认1天" not in manual_sources
    assert "不填默认 1 天" not in manual_sources
    assert "留空默认 1 天" not in manual_sources
    assert "周期默认1天" not in manual_sources


def test_excel_templates_default_to_chinese_enum_values_accepted_by_backend() -> None:
    from core.models.enums import (
        BatchPriority,
        CalendarDayType,
        MachineStatus,
        OperatorStatus,
        ReadyStatus,
        SourceType,
        SupplierStatus,
        YesNo,
    )
    from core.services.common.enum_normalizers import (
        normalize_machine_status,
        normalize_op_type_category,
        normalize_operator_status,
        normalize_skill_level,
        normalize_supplier_status,
        normalize_yesno_narrow,
    )
    from core.services.common.excel_templates import get_default_templates
    from core.services.common.normalization_matrix import (
        normalize_batch_priority_value,
        normalize_calendar_day_type_value,
        normalize_ready_status_value,
    )

    raw_enum_values = {
        "active",
        "inactive",
        "maintain",
        "internal",
        "external",
        "beginner",
        "normal",
        "expert",
        "urgent",
        "critical",
        "partial",
        "workday",
        "holiday",
        "yes",
        "no",
    }
    normalizers = {
        ("人员基本信息.xlsx", "状态"): (normalize_operator_status, {OperatorStatus.ACTIVE.value, OperatorStatus.INACTIVE.value}),
        ("设备信息.xlsx", "状态"): (normalize_machine_status, set(MachineStatus._value2member_map_)),
        ("工种配置.xlsx", "归属"): (normalize_op_type_category, set(SourceType._value2member_map_)),
        ("供应商配置.xlsx", "状态"): (normalize_supplier_status, set(SupplierStatus._value2member_map_)),
        ("人员设备关联.xlsx", "技能等级"): (lambda value: normalize_skill_level(value, default="normal"), {"beginner", "normal", "expert"}),
        ("设备人员关联.xlsx", "技能等级"): (lambda value: normalize_skill_level(value, default="normal"), {"beginner", "normal", "expert"}),
        ("人员设备关联.xlsx", "主操设备"): (normalize_yesno_narrow, set(YesNo._value2member_map_)),
        ("设备人员关联.xlsx", "主操设备"): (normalize_yesno_narrow, set(YesNo._value2member_map_)),
        ("批次信息.xlsx", "优先级"): (normalize_batch_priority_value, set(BatchPriority._value2member_map_)),
        ("批次信息.xlsx", "齐套"): (normalize_ready_status_value, set(ReadyStatus._value2member_map_)),
        ("工作日历.xlsx", "类型"): (normalize_calendar_day_type_value, {CalendarDayType.WORKDAY.value, CalendarDayType.HOLIDAY.value}),
        ("人员专属工作日历.xlsx", "类型"): (
            normalize_calendar_day_type_value,
            {CalendarDayType.WORKDAY.value, CalendarDayType.HOLIDAY.value},
        ),
        ("工作日历.xlsx", "允许普通件"): (normalize_yesno_narrow, set(YesNo._value2member_map_)),
        ("工作日历.xlsx", "允许急件"): (normalize_yesno_narrow, set(YesNo._value2member_map_)),
        ("人员专属工作日历.xlsx", "允许普通件"): (normalize_yesno_narrow, set(YesNo._value2member_map_)),
        ("人员专属工作日历.xlsx", "允许急件"): (normalize_yesno_narrow, set(YesNo._value2member_map_)),
    }

    for template_def in get_default_templates():
        filename = str(template_def["filename"])
        headers = list(template_def["headers"])
        enum_cols = template_def.get("format_spec", {}).get("enum_cols", {}) or {}
        sample_rows = list(template_def.get("sample_rows") or [])
        enum_cells = []
        for row in sample_rows:
            for col_idx, cell in enumerate(row):
                if col_idx in enum_cols and isinstance(cell, str):
                    enum_cells.append(cell)
        enum_values = enum_cells + [str(value) for values in enum_cols.values() for value in values]
        assert not (set(enum_values) & raw_enum_values), f"{filename} 模板默认值不应继续使用英文枚举"

        for col_idx, values in enum_cols.items():
            header = str(headers[col_idx])
            key = (filename, header)
            assert key in normalizers, f"{filename} 的枚举列 {header} 缺少后端归一化测试"
            normalize, allowed = normalizers[key]
            for value in values:
                assert normalize(value) in allowed, f"{filename}.{header} 的模板值 {value!r} 后端不接受"


def test_ensure_excel_templates_refreshes_known_stale_generated_template(tmp_path) -> None:
    from core.services.common.excel_templates import build_xlsx_bytes, ensure_excel_templates

    stale_path = tmp_path / "人员基本信息.xlsx"
    stale_path.write_bytes(
        build_xlsx_bytes(
            ["工号", "姓名", "状态", "班组", "备注"],
            [["OP001", "张三", "active", None, "旧模板"]],
            format_spec={"enum_cols": {2: ["active", "inactive"]}},
        ).getvalue()
    )

    stats = ensure_excel_templates(str(tmp_path))
    assert "人员基本信息.xlsx" in stats["created"]

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=stale_path, data_only=True)
        ws = workbook.active
        assert ws["C2"].value == "在岗"
    finally:
        if workbook is not None:
            workbook.close()


def test_template_download_preserves_existing_disk_file_when_headers_match(tmp_path) -> None:
    from flask import Flask

    from core.services.common.excel_templates import build_xlsx_bytes
    from web.routes.excel_utils import send_excel_template_file

    custom_path = tmp_path / "人员基本信息.xlsx"
    custom_path.write_bytes(
        build_xlsx_bytes(
            ["工号", "姓名", "状态", "班组", "备注"],
            [["OP999", "自定义示例", "active", "A组", "保留用户模板"]],
            format_spec={"enum_cols": {2: ["active", "inactive", "在岗", "停用"]}},
        ).getvalue()
    )

    app = Flask(__name__)

    @app.get("/template")
    def _template():
        return send_excel_template_file(str(custom_path), download_name="人员基本信息.xlsx")

    response = app.test_client().get("/template")
    assert response.status_code == 200

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=io.BytesIO(response.data), data_only=True)
        ws = workbook.active
        assert ws["A2"].value == "OP999"
        assert ws["C2"].value == "active"
        assert ws["E2"].value == "保留用户模板"
    finally:
        if workbook is not None:
            workbook.close()


def test_ensure_excel_templates_preserves_user_custom_template_with_extra_rows(tmp_path) -> None:
    from core.services.common.excel_templates import build_xlsx_bytes, ensure_excel_templates

    custom_path = tmp_path / "人员基本信息.xlsx"
    custom_path.write_bytes(
        build_xlsx_bytes(
            ["工号", "姓名", "状态", "班组", "备注"],
            [
                ["OP001", "张三", "active", "A组", "用户自定义示例"],
                ["OP002", "李四", "inactive", "B组", "用户额外示例"],
            ],
            format_spec={"enum_cols": {2: ["active", "inactive", "在岗", "停用"]}},
        ).getvalue()
    )

    stats = ensure_excel_templates(str(tmp_path))
    assert "人员基本信息.xlsx" in stats["skipped"]

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=custom_path, data_only=True)
        ws = workbook.active
        assert ws["C2"].value == "active"
        assert ws["E3"].value == "用户额外示例"
    finally:
        if workbook is not None:
            workbook.close()


def test_import_errors_present_chinese_first_and_english_as_compatible_aliases() -> None:
    sources = "\n".join(
        _read(rel_path)
        for rel_path in (
            "core/services/common/excel_validators.py",
            "web/routes/process_excel_op_types.py",
            "web/routes/process_excel_suppliers.py",
            "web/routes/domains/scheduler/scheduler_excel_calendar.py",
            "core/services/personnel/operator_machine_normalizers.py",
        )
    )
    assert "允许：normal/urgent/critical；或中文" not in sources
    assert "允许：yes/no/partial；或中文" not in sources
    assert "允许：workday/holiday；或中文" not in sources
    assert "允许：yes/no/true/false/1/0；或中文" not in sources
    assert "允许：internal / external；或中文" not in sources
    assert "允许：active / inactive；或中文" not in sources
    assert "可填写：普通 / 急件 / 特急；也兼容英文标准值 normal/urgent/critical" in sources
    assert "可填写：齐套 / 未齐套 / 部分齐套 / 是 / 否；也兼容英文标准值 yes/no/partial" in sources
    assert "可填写：工作日 / 假期 / 周末 / 节假日；也兼容英文标准值 workday/holiday" in sources
    assert "可填写：是 / 否；也兼容英文标准值 yes/no/true/false/1/0" in sources
    assert "可填写：内部 / 外部 / 内 / 外；也兼容英文标准值 internal/external" in sources
    assert "可填写：启用 / 停用 / 在用 / 正常 / 禁用；也兼容英文标准值 active/inactive" in sources
    assert "也兼容英文标准值 yes/no、true/false、1/0、on/off" in sources


def test_excel_exports_use_chinese_labels_for_enum_columns() -> None:
    route_expectations = {
        "web/routes/personnel_excel_operators.py": ("operator_status_label",),
        "web/routes/equipment_excel_machines.py": ("machine_status_label",),
        "web/routes/process_excel_op_types.py": ("source_type_label",),
        "web/routes/process_excel_suppliers.py": ("supplier_status_label",),
        "web/routes/personnel_excel_links.py": ("skill_level_label", "yes_no_label"),
        "web/routes/equipment_excel_links.py": ("skill_level_label", "yes_no_label"),
        "web/routes/domains/scheduler/scheduler_excel_batches.py": ("batch_priority_label", "ready_status_label"),
        "web/routes/domains/scheduler/scheduler_excel_calendar.py": ("calendar_day_type_label", "yes_no_label"),
        "web/routes/personnel_excel_operator_calendar.py": ("calendar_day_type_label", "yes_no_label"),
        "web/routes/process_excel_part_operations.py": ("source_type_label",),
    }
    for rel_path, helpers in route_expectations.items():
        source = _read(rel_path)
        for helper in helpers:
            assert helper in source, f"{rel_path} 导出枚举字段应使用 {helper}"


def test_frontend_scripts_keep_internal_details_out_of_user_messages() -> None:
    gantt_boot = _read("static/js/gantt_boot.js")
    assert "页面脚本加载不完整" in gantt_boot
    assert "reportClientError" in gantt_boot
    assert "缺失：" not in gantt_boot
    assert "未找到数据接口 URL（data-url；兼容 data-data-url）" not in gantt_boot
    assert "dataUrl=" not in gantt_boot
    assert "甘特图数据请求超过" in gantt_boot
    assert ">${fetchTimeoutMs}ms" not in gantt_boot

    gantt_render = _read("static/js/gantt_render.js")
    assert "甘特图装饰刷新失败" in gantt_render
    assert "Gantt decorate failed" not in gantt_render
    assert "间隔（分钟）" in gantt_render
    assert "间隔(分)" not in gantt_render

    resource_dispatch = _read("static/js/resource_dispatch.js")
    assert "未知退化提示" in resource_dispatch
    assert "parts.push(escapeHtml(code));" not in resource_dispatch


def test_process_and_scheduler_errors_use_chinese_terms() -> None:
    route_parser = _read("core/services/process/route_parser.py")
    assert "严格模式已拒绝" in route_parser
    assert "strict_mode 已拒绝" not in route_parser
    assert "默认周期无法解析（{raw_default_days!r}）" not in route_parser
    assert "默认周期无效（{raw_default_days!r}）" not in route_parser

    external_group_service = _read("core/services/process/external_group_service.py")
    assert "兼容模式已按 1.0 天回退" in external_group_service
    assert "append_unique_text_messages(user_warnings, user_warning_text)" in external_group_service
    assert "safe_warning(self.logger, log_warning_text)" in external_group_service

    batch_template_ops = _read("core/services/scheduler/batch_template_ops.py")
    assert "不支持严格模式" in batch_template_ops
    assert "不支持 strict_mode" not in batch_template_ops


def test_scheduler_analysis_hides_internal_schema_and_attempt_tags() -> None:
    analysis_vm = _read("web/viewmodels/scheduler_analysis_vm.py")
    assert "当前版本摘要缺少部分分析字段" in analysis_vm
    assert "新 schema 字段" not in analysis_vm
    assert '"comparison_metric": "优化对比指标"' in analysis_vm
    assert '"best_score_schema": "评分顺序"' in analysis_vm

    analysis_template = _read("templates/scheduler/analysis.html")
    assert "compat_fallback.missing_field_labels" in analysis_template
    assert "compat_fallback.missing_fields | join" not in analysis_template
    assert "<th>方案</th>" in analysis_template
    assert "{{ r.tag }}" not in analysis_template
    assert "方案 {{ loop.index }}" in analysis_template


@pytest.mark.parametrize(
    ("rel_path", "bad_tokens", "good_tokens"),
    (
        (
            "core/services/equipment/machine_service.py",
            ("璁惧",),
            ("设备编号不能为空", "设备名称不能为空"),
        ),
        (
            "core/services/process/op_type_service.py",
            ("鈥", "宸ョ"),
            ("“工种ID”不能为空", "“工种名称”不能为空"),
        ),
        (
            "core/services/equipment/machine_downtime_service.py",
            ("鍋滄満",),
            ("停机记录 ID 缺失，无法执行取消。",),
        ),
    ),
)
def test_known_garbled_error_messages_are_repaired(
    rel_path: str,
    bad_tokens: tuple[str, ...],
    good_tokens: tuple[str, ...],
) -> None:
    source = _read(rel_path)
    for token in bad_tokens:
        assert token not in source
    for token in good_tokens:
        assert token in source
