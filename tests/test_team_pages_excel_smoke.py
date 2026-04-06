from __future__ import annotations

import importlib
import io
import os
import sys
from pathlib import Path
from typing import List

REPO_ROOT = Path(__file__).resolve().parents[1]


def _assert_status(resp, name: str, expect: int = 200) -> None:
    if resp.status_code != expect:
        body = resp.data.decode("utf-8", errors="ignore") if getattr(resp, "data", None) else ""
        raise AssertionError(f"{name} 返回 {resp.status_code}，期望 {expect}，body={body[:500]}")


def _xlsx_headers(content: bytes) -> List[str]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content))
    try:
        ws = wb.active
        if ws is None:
            return []
        return [str(ws.cell(1, idx).value or "").strip() for idx in range(1, ws.max_column + 1)]
    finally:
        wb.close()


def test_team_pages_and_excel_routes_show_team_columns_and_headers(tmp_path, monkeypatch) -> None:
    repo_root = str(REPO_ROOT)
    if repo_root not in sys.path:
        sys.path.insert(0, repo_root)

    test_db = tmp_path / "aps_test.db"
    test_logs = tmp_path / "logs"
    test_backups = tmp_path / "backups"
    test_templates = tmp_path / "templates_excel"
    test_logs.mkdir(parents=True, exist_ok=True)
    test_backups.mkdir(parents=True, exist_ok=True)
    test_templates.mkdir(parents=True, exist_ok=True)

    monkeypatch.setenv("APS_ENV", "development")
    monkeypatch.setenv("APS_DB_PATH", str(test_db))
    monkeypatch.setenv("APS_LOG_DIR", str(test_logs))
    monkeypatch.setenv("APS_BACKUP_DIR", str(test_backups))
    monkeypatch.setenv("APS_EXCEL_TEMPLATE_DIR", str(test_templates))

    from core.infrastructure.database import ensure_schema, get_connection

    ensure_schema(str(test_db), logger=None, schema_path=os.path.join(repo_root, "schema.sql"))

    conn = get_connection(str(test_db))
    try:
        conn.execute(
            "INSERT INTO ResourceTeams (team_id, name, status) VALUES (?, ?, ?)",
            ("TEAM-01", "车工一组", "active"),
        )
        conn.execute(
            "INSERT INTO Operators (operator_id, name, status, team_id, remark) VALUES (?, ?, ?, ?, ?)",
            ("OP001", "张三", "active", "TEAM-01", "人员备注"),
        )
        conn.execute(
            "INSERT INTO Machines (machine_id, name, status, team_id, remark) VALUES (?, ?, ?, ?, ?)",
            ("MC001", "数控车床1", "active", "TEAM-01", "设备备注"),
        )
        conn.commit()
    finally:
        conn.close()

    sys.modules.pop("app", None)
    app_mod = importlib.import_module("app")
    app = app_mod.create_app()
    client = app.test_client()

    assert _xlsx_headers((test_templates / "人员基本信息.xlsx").read_bytes()) == ["工号", "姓名", "状态", "班组", "备注"]
    assert _xlsx_headers((test_templates / "设备信息.xlsx").read_bytes()) == ["设备编号", "设备名称", "工种", "班组", "状态"]

    resp_team_page = client.get("/personnel/teams")
    _assert_status(resp_team_page, "GET /personnel/teams")
    html_team_page = resp_team_page.data.decode("utf-8", errors="ignore")
    assert "班组管理" in html_team_page
    assert "车工一组" in html_team_page

    resp_personnel = client.get("/personnel/?team_id=TEAM-01")
    _assert_status(resp_personnel, "GET /personnel/?team_id=TEAM-01")
    html_personnel = resp_personnel.data.decode("utf-8", errors="ignore")
    assert "车工一组" in html_personnel
    assert "OP001" in html_personnel
    assert "班组筛选" in html_personnel

    resp_equipment = client.get("/equipment/?team_id=TEAM-01")
    _assert_status(resp_equipment, "GET /equipment/?team_id=TEAM-01")
    html_equipment = resp_equipment.data.decode("utf-8", errors="ignore")
    assert "车工一组" in html_equipment
    assert "MC001" in html_equipment
    assert "班组筛选" in html_equipment

    resp_personnel_invalid = client.get("/personnel/?team_id=TEAM-404")
    assert resp_personnel_invalid.status_code == 302
    assert resp_personnel_invalid.headers["Location"].endswith("/personnel/")

    resp_equipment_invalid = client.get("/equipment/?team_id=TEAM-404")
    assert resp_equipment_invalid.status_code == 302
    assert resp_equipment_invalid.headers["Location"].endswith("/equipment/")

    resp_operator_excel = client.get("/personnel/excel/operators")
    _assert_status(resp_operator_excel, "GET /personnel/excel/operators")
    html_operator_excel = resp_operator_excel.data.decode("utf-8", errors="ignore")
    assert "班组" in html_operator_excel
    assert "车工一组" in html_operator_excel

    resp_machine_excel = client.get("/equipment/excel/machines")
    _assert_status(resp_machine_excel, "GET /equipment/excel/machines")
    html_machine_excel = resp_machine_excel.data.decode("utf-8", errors="ignore")
    assert "班组" in html_machine_excel
    assert "车工一组" in html_machine_excel

    resp_operator_template = client.get("/personnel/excel/operators/template")
    _assert_status(resp_operator_template, "GET /personnel/excel/operators/template")
    assert _xlsx_headers(resp_operator_template.data) == ["工号", "姓名", "状态", "班组", "备注"]

    resp_machine_template = client.get("/equipment/excel/machines/template")
    _assert_status(resp_machine_template, "GET /equipment/excel/machines/template")
    assert _xlsx_headers(resp_machine_template.data) == ["设备编号", "设备名称", "工种", "班组", "状态"]

    resp_operator_export = client.get("/personnel/excel/operators/export")
    _assert_status(resp_operator_export, "GET /personnel/excel/operators/export")
    assert _xlsx_headers(resp_operator_export.data) == ["工号", "姓名", "状态", "班组", "备注"]

    resp_machine_export = client.get("/equipment/excel/machines/export")
    _assert_status(resp_machine_export, "GET /equipment/excel/machines/export")
    assert _xlsx_headers(resp_machine_export.data) == ["设备编号", "设备名称", "工种", "班组", "状态"]

    (test_templates / "人员基本信息.xlsx").unlink()
    resp_excel_demo_template = client.get("/excel-demo/template")
    _assert_status(resp_excel_demo_template, "GET /excel-demo/template")
    assert _xlsx_headers(resp_excel_demo_template.data) == ["工号", "姓名", "状态", "班组", "备注"]
