from __future__ import annotations

import importlib
import io
import json
import os
import sys
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[1]


def _assert_status(resp, name: str, expect: int = 200) -> None:
    if resp.status_code != expect:
        body = resp.data.decode("utf-8", errors="ignore") if getattr(resp, "data", None) else ""
        raise AssertionError(f"{name} 返回 {resp.status_code}，期望 {expect}，body={body[:800]}")


def _set_ui_mode_cookie(client, mode: str) -> None:
    try:
        client.set_cookie("aps_ui_mode", mode, domain="localhost")
    except TypeError:
        client.set_cookie("localhost", "aps_ui_mode", mode)


def test_scheduler_resource_dispatch_page_data_export_and_dashboard_entry(tmp_path, monkeypatch) -> None:
    repo_root = str(REPO_ROOT)
    if repo_root not in sys.path:
        sys.path.insert(0, repo_root)

    test_db = tmp_path / "aps_test.db"
    test_logs = tmp_path / "logs"
    test_backups = tmp_path / "backups"
    test_templates = tmp_path / "templates_excel"
    test_logs.mkdir(parents=True, exist_ok=True)
    test_backups.mkdir(parents=True, exist_ok=True)
    test_templates.mkdir(parents=True, exist_ok=True)

    monkeypatch.setenv("APS_ENV", "development")
    monkeypatch.setenv("APS_DB_PATH", str(test_db))
    monkeypatch.setenv("APS_LOG_DIR", str(test_logs))
    monkeypatch.setenv("APS_BACKUP_DIR", str(test_backups))
    monkeypatch.setenv("APS_EXCEL_TEMPLATE_DIR", str(test_templates))

    from core.infrastructure.database import ensure_schema, get_connection

    ensure_schema(str(test_db), logger=None, schema_path=os.path.join(repo_root, "schema.sql"))

    conn = get_connection(str(test_db))
    try:
        conn.execute("INSERT INTO ResourceTeams (team_id, name, status) VALUES (?, ?, ?)", ("TEAM-OP", "装配一组", "active"))
        conn.execute("INSERT INTO ResourceTeams (team_id, name, status) VALUES (?, ?, ?)", ("TEAM-MC", "设备保障组", "active"))
        conn.execute(
            "INSERT INTO Operators (operator_id, name, status, team_id, remark) VALUES (?, ?, ?, ?, ?)",
            ("OP001", "张三", "active", "TEAM-OP", ""),
        )
        conn.execute(
            "INSERT INTO Machines (machine_id, name, status, team_id, remark) VALUES (?, ?, ?, ?, ?)",
            ("MC001", "数控车床1", "active", "TEAM-MC", ""),
        )
        conn.execute(
            "INSERT INTO Parts (part_no, part_name, route_raw) VALUES (?, ?, ?)",
            ("PART-001", "回转壳体", "[]"),
        )
        conn.execute(
            "INSERT INTO Batches (batch_id, part_no, part_name, quantity, due_date, priority, ready_status, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("B001", "PART-001", "回转壳体", 5, "2026-03-08", "urgent", "yes", "scheduled"),
        )
        cur = conn.execute(
            "INSERT INTO BatchOperations (op_code, batch_id, piece_id, seq, op_type_name, source, machine_id, operator_id, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("OP-B001-10", "B001", "P1", 10, "车削", "internal", "MC001", "OP001", "scheduled"),
        )
        assert cur.lastrowid is not None
        op_id = int(cur.lastrowid)
        conn.execute(
            "INSERT INTO Schedule (op_id, machine_id, operator_id, start_time, end_time, lock_status, version) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (op_id, "MC001", "OP001", "2026-03-02 08:00:00", "2026-03-02 12:00:00", "locked", 1),
        )
        conn.execute(
            "INSERT INTO ScheduleHistory (version, strategy, batch_count, op_count, result_status, result_summary, created_by) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                1,
                "greedy",
                1,
                1,
                "success",
                json.dumps(
                    {"overdue_batches": {"count": 1, "items": [{"batch_id": "B001", "hours": 4}]}},
                    ensure_ascii=False,
                ),
                "test",
            ),
        )
        conn.commit()
    finally:
        conn.close()

    sys.modules.pop("app", None)
    app_mod = importlib.import_module("app")
    app = app_mod.create_app()
    client = app.test_client()

    _set_ui_mode_cookie(client, "v1")
    resp_dashboard_v1 = client.get("/")
    _assert_status(resp_dashboard_v1, "GET / (v1)")
    html_dashboard_v1 = resp_dashboard_v1.data.decode("utf-8", errors="ignore")
    assert "资源排班" in html_dashboard_v1

    _set_ui_mode_cookie(client, "v2")
    resp_dashboard_v2 = client.get("/")
    _assert_status(resp_dashboard_v2, "GET / (v2)")
    html_dashboard_v2 = resp_dashboard_v2.data.decode("utf-8", errors="ignore")
    assert "资源排班中心" in html_dashboard_v2

    query = "scope_type=operator&operator_id=OP001&period_preset=week&query_date=2026-03-02&version=1"

    resp_page = client.get(f"/scheduler/resource-dispatch?{query}")
    _assert_status(resp_page, "GET /scheduler/resource-dispatch")
    html_page = resp_page.data.decode("utf-8", errors="ignore")
    assert "资源排班中心" in html_page
    assert 'id="rdPage"' in html_page
    assert "导出资源排班.xlsx" in html_page
    assert '<option value="team"' in html_page

    resp_data = client.get(f"/scheduler/resource-dispatch/data?{query}")
    _assert_status(resp_data, "GET /scheduler/resource-dispatch/data")
    payload = json.loads(resp_data.data.decode("utf-8", errors="ignore") or "{}")
    assert payload.get("success") is True
    data = payload.get("data") or {}
    assert (data.get("filters") or {}).get("scope_id") == "OP001"
    assert len(data.get("detail_rows") or []) == 1
    assert len(data.get("tasks") or []) == 1
    summary = data.get("summary") or {}
    assert summary.get("degraded") is False
    assert summary.get("degradation_events") == []
    assert summary.get("degradation_counters") == {}
    assert summary.get("empty_reason") is None
    assert len(data.get("calendar_rows") or []) == 1
    first_row = (data.get("detail_rows") or [])[0]
    assert first_row.get("team_relation_label") == "跨班组借调"
    assert first_row.get("is_overdue") is True

    resp_export = client.get(f"/scheduler/resource-dispatch/export?{query}")
    _assert_status(resp_export, "GET /scheduler/resource-dispatch/export")
    wb = openpyxl.load_workbook(io.BytesIO(resp_export.data))
    assert wb.sheetnames == ["查询摘要", "任务明细", "日历排班"]
    ws_detail = cast(Worksheet, wb["任务明细"])
    max_column = int(ws_detail.max_column)
    headers = [str(ws_detail.cell(1, idx).value or "").strip() for idx in range(1, max_column + 1)]
    assert headers[:8] == ["排程ID", "工序ID", "工序编码", "批次号", "图号", "工序", "工序名称", "开始时间"]
    assert str(ws_detail.cell(2, 4).value or "").strip() == "B001"

    query_team = "scope_type=team&team_id=TEAM-OP&team_axis=operator&period_preset=week&query_date=2026-03-02&version=1"

    resp_team_page = client.get(f"/scheduler/resource-dispatch?{query_team}")
    _assert_status(resp_team_page, "GET /scheduler/resource-dispatch (team)")
    html_team_page = resp_team_page.data.decode("utf-8", errors="ignore")
    assert "班组轴" in html_team_page
    assert '<option value="team" selected' in html_team_page

    resp_team_data = client.get(f"/scheduler/resource-dispatch/data?{query_team}")
    _assert_status(resp_team_data, "GET /scheduler/resource-dispatch/data (team)")
    payload_team = json.loads(resp_team_data.data.decode("utf-8", errors="ignore") or "{}")
    assert payload_team.get("success") is True
    data_team = payload_team.get("data") or {}
    assert (data_team.get("filters") or {}).get("scope_id") == "TEAM-OP"
    assert (data_team.get("filters") or {}).get("team_axis") == "operator"
    assert len(data_team.get("operator_rows") or []) == 1
    assert len(data_team.get("machine_rows") or []) == 0
    assert len(data_team.get("cross_team_rows") or []) == 1
    assert len(data_team.get("tasks") or []) == 1

    resp_team_export = client.get(f"/scheduler/resource-dispatch/export?{query_team}")
    _assert_status(resp_team_export, "GET /scheduler/resource-dispatch/export (team)")
    wb_team = openpyxl.load_workbook(io.BytesIO(resp_team_export.data))
    assert wb_team.sheetnames == [
        "查询摘要",
        "班组人员任务明细",
        "班组设备任务明细",
        "班组人员日历",
        "班组设备日历",
        "跨班组借调",
    ]

    resp_gantt = client.get("/scheduler/gantt/data?view=machine&week_start=2026-03-02&version=1")
    _assert_status(resp_gantt, "GET /scheduler/gantt/data")
    gantt_payload = json.loads(resp_gantt.data.decode("utf-8", errors="ignore") or "{}")
    assert gantt_payload.get("success") is True
    gantt_data = gantt_payload.get("data") or {}
    assert int(gantt_data.get("task_count") or 0) >= 1
    assert isinstance(gantt_data.get("critical_chain"), dict)

    resp_week_plan_export = client.get("/scheduler/week-plan/export?week_start=2026-03-02&version=1")
    _assert_status(resp_week_plan_export, "GET /scheduler/week-plan/export")
    wb_week_plan = openpyxl.load_workbook(io.BytesIO(resp_week_plan_export.data))
    assert wb_week_plan.sheetnames
