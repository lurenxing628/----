import os
import re
import sys
import tempfile

import openpyxl


def find_repo_root() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.abspath(os.path.join(here, ".."))
    if os.path.exists(os.path.join(repo_root, "app.py")) and os.path.exists(os.path.join(repo_root, "schema.sql")):
        return repo_root
    raise RuntimeError("未找到项目根目录：要求存在 app.py 与 schema.sql")


def _build_source_xlsx(path: str) -> None:
    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        assert ws is not None

        ws.title = "单元产品信息统计"
        ws.append(
            [
                "图号",
                "名称",
                "是否关重件",
                "关键特性",
                "工艺路线",
                "材料牌号",
                "材料规格",
                "进单元可装夹直径",
                "3140124 胡凡 罗辉",
                "换型时间(min)",
                "单件加工时间(min)",
                "批次加工时间(min)",
                "钳工\n程国华",
                "换型时间(min)",
                "单件加工时间(min)",
                "批次加工时间(min)",
            ]
        )
        # P001：seq=5/15 为“有工步+有人员” => internal；seq=10/20 => external
        ws.append(
            [
                "P001",
                "壳体A",
                "是",
                "",
                "5数车10热处理15数铣20总检",
                "",
                "",
                "",
                "5-1粗车",
                30,
                20,
                50,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "15-1精铣",
                20,
                15,
                35,
                "20去毛刺",  # 无 XX-X，不算工步 => 本工序按外协
                10,
                5,
                15,
            ]
        )
    # P002：seq=5 外协，但同名“数车”在 P001 自制工序出现，需自动改名为“数车（外协）”
        ws.append(
            [
                "P002",
                "壳体B",
                "否",
                "",
                "5数车10数铣",
                "",
                "",
                "",
                "10-1铣端面",
                10,
                12,
                22,
                None,
                None,
                None,
                None,
            ]
        )
        wb.save(path)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_headers(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        assert ws is not None

        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return list(first_row or [])
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main() -> None:
    repo_root = find_repo_root()
    if repo_root not in sys.path:
        sys.path.insert(0, repo_root)

    from core.infrastructure.database import ensure_schema, get_connection
    from core.services.common.excel_service import ImportMode, RowStatus
    from core.services.personnel import OperatorMachineService
    from core.services.process import UnitExcelConverter

    tmpdir = tempfile.mkdtemp(prefix="aps_regression_unit_converter_")
    src_xlsx = os.path.join(tmpdir, "source.xlsx")
    out_dir = os.path.join(tmpdir, "out")
    db_path = os.path.join(tmpdir, "test.db")

    _build_source_xlsx(src_xlsx)

    converter = UnitExcelConverter()
    converted = converter.convert(src_xlsx)
    output_paths = converter.write_templates(converted=converted, output_dir=out_dir)

    # 1) XX-X 工步并工序 + 内外协判定
    routes = {r["图号"]: r["工艺路线字符串"] for r in converted.routes_rows}
    assert "P001" in routes and "P002" in routes
    assert routes["P001"] == "5数车10热处理15数铣20总检", f"P001 路线异常：{routes['P001']!r}"
    assert routes["P002"] == "5数车（外协）10数铣", f"P002 路线异常：{routes['P002']!r}"

    # 2) 解析 “3140124 胡凡 罗辉”
    operators = {r["姓名"]: r["工号"] for r in converted.operators_rows}
    assert "胡凡" in operators and "罗辉" in operators, "人员解析失败（胡凡/罗辉）"
    for op_id in operators.values():
        assert re.fullmatch(r"\d{6}", str(op_id)), f"工号应为6位数字，实际={op_id!r}"

    # 2.1) 设备名称不应出现“设备XXXX”兜底前缀
    for m in converted.machines_rows:
        name = str(m.get("设备名称") or "")
        assert not name.startswith("设备"), f"设备名称不应带“设备”前缀，实际={name!r}"

    # 3) 人员设备关联输出对齐主模板 4 列（补默认技能等级/主操设备）
    assert converted.operator_machine_rows, "人员设备关联输出为空"
    for row in converted.operator_machine_rows:
        assert set(row.keys()) == {"工号", "设备编号", "技能等级", "主操设备"}, f"人员设备关联列不符合预期：{row.keys()}"
        assert row["技能等级"] == "normal", f"人员设备关联默认技能等级异常：{row!r}"
        assert row["主操设备"] == "no", f"人员设备关联默认主操设备异常：{row!r}"

    links = {(r["工号"], r["设备编号"]) for r in converted.operator_machine_rows}
    assert (operators["胡凡"], "3140124") in links
    assert (operators["罗辉"], "3140124") in links

    # 3.1) 供应商配置输出对齐主模板 6 列（补默认状态/备注）
    assert converted.suppliers_rows, "供应商配置输出为空"
    for row in converted.suppliers_rows:
        assert set(row.keys()) == {"供应商ID", "名称", "对应工种", "默认周期", "状态", "备注"}, f"供应商配置列不符合预期：{row.keys()}"
        assert row["状态"] == "启用", f"供应商默认状态异常：{row!r}"
        assert row["备注"] is None, f"供应商默认备注异常：{row!r}"

    # 4) 工种冲突自动改名：同名自制/外协并存时，外协侧加“（外协）”
    op_types = {r["工种名称"]: r["归属"] for r in converted.op_types_rows}
    assert op_types.get("数车") == "自制", f"数车归属异常：{op_types.get('数车')!r}"
    assert op_types.get("数车（外协）") == "外协", f"数车（外协）归属异常：{op_types.get('数车（外协）')!r}"

    # 5) 工步时间累计成工序工时（单位：小时）
    hours_map = {(r["图号"], int(r["工序"])): (float(r["换型时间(h)"]), float(r["单件工时(h)"])) for r in converted.part_operation_hours_rows}
    assert ("P001", 5) in hours_map and ("P001", 15) in hours_map and ("P002", 10) in hours_map
    assert abs(hours_map[("P001", 5)][0] - 0.5) < 1e-6 and abs(hours_map[("P001", 5)][1] - (20.0 / 60.0)) < 1e-6
    assert abs(hours_map[("P001", 15)][0] - (20.0 / 60.0)) < 1e-6 and abs(hours_map[("P001", 15)][1] - 0.25) < 1e-6

    # 6) 与现有导入链路兼容：人员设备关联缺少“技能等级/主操设备”仍可预览通过
    ensure_schema(db_path, logger=None, schema_path=os.path.join(repo_root, "schema.sql"))
    conn = get_connection(db_path)
    try:
        for r in converted.operators_rows:
            conn.execute(
                "INSERT INTO Operators (operator_id, name, status) VALUES (?, ?, ?)",
                (r["工号"], r["姓名"], "active"),
            )
        for r in converted.machines_rows:
            conn.execute(
                "INSERT INTO Machines (machine_id, name, status) VALUES (?, ?, ?)",
                (r["设备编号"], r["设备名称"], "active"),
            )
        conn.commit()

        link_svc = OperatorMachineService(conn)
        preview = link_svc.preview_import_links(rows=converted.operator_machine_rows, mode=ImportMode.OVERWRITE)
        assert preview, "preview_import_links 结果为空"
        bad = [p for p in preview if p.status == RowStatus.ERROR]
        assert not bad, f"关联预览出现错误：{[(p.row_num, p.message) for p in bad]}"
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # 7) 文件输出完整性
    expected_files = {
        "零件工艺路线.xlsx",
        "零件工序工时.xlsx",
        "人员基本信息.xlsx",
        "设备信息.xlsx",
        "人员设备关联.xlsx",
        "工种配置.xlsx",
        "供应商配置.xlsx",
    }
    assert expected_files.issubset(set(output_paths.keys()))
    for fn in expected_files:
        assert os.path.exists(output_paths[fn]), f"输出文件缺失：{fn}"
    assert _read_headers(output_paths["人员设备关联.xlsx"]) == ["工号", "设备编号", "技能等级", "主操设备"]
    assert _read_headers(output_paths["供应商配置.xlsx"]) == ["供应商ID", "名称", "对应工种", "默认周期", "状态", "备注"]

    print("OK")


if __name__ == "__main__":
    main()
