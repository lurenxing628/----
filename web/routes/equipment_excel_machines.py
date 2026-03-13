from __future__ import annotations

import io
import json
import os
import time
from typing import Any, Dict, List, Optional

from flask import current_app, flash, g, redirect, request, send_file, url_for

from core.infrastructure.errors import ValidationError
from core.models.enums import MACHINE_STATUS_VALUES
from core.services.common.enum_normalizers import normalize_machine_status
from core.services.common.excel_audit import log_excel_export, log_excel_import
from core.services.common.excel_backend_factory import get_excel_backend
from core.services.common.excel_service import ExcelService, ImportMode, RowStatus
from core.services.equipment import MachineService
from core.services.equipment.machine_excel_import_service import MachineExcelImportService
from core.services.personnel import ResourceTeamService
from core.services.process import OpTypeService
from web.ui_mode import render_ui_template as render_template

from .equipment_bp import _ensure_unique_ids, _parse_mode, _read_uploaded_xlsx, bp

# ============================================================
# Excel：设备信息（Machines）
# （Phase4-03 会在此基础上完善；先占位，避免后续拆文件）
# ============================================================


def _validate_machine_excel_row(row: Dict[str, Any]) -> Optional[str]:
    if not row.get("设备编号") or str(row.get("设备编号")).strip() == "":
        return "设备编号不能为空"
    if not row.get("设备名称") or str(row.get("设备名称")).strip() == "":
        return "设备名称不能为空"

    status = row.get("状态")
    if status is None or str(status).strip() == "":
        return "状态不能为空，请填写：可用 / 停用 / 维修（也兼容 active / inactive / maintain）。"
    st = _normalize_machine_status_for_excel(status)
    if st not in MACHINE_STATUS_VALUES:
        return "状态不合法，可填写：可用 / 停用 / 维修（也兼容 active / inactive / maintain）。"
    row["状态"] = st

    return None


def _normalize_machine_status_for_excel(value: Any) -> str:
    """
    Excel 中的状态允许：
    - 英文：active / inactive / maintain
    - 中文：可用 / 停用 / 维修/维护
    返回统一的英文枚举值。
    """
    return normalize_machine_status(value)


def _resolve_op_type(value: Any, op_type_svc: OpTypeService) -> Dict[str, Optional[str]]:
    """
    解析 Excel 的工种字段：
    - 支持填写 op_type_id 或 工种名称
    - 返回 dict：{op_type_id, op_type_name}
    """
    v = None if value is None else str(value).strip()
    if not v:
        return {"op_type_id": None, "op_type_name": None}
    ot = op_type_svc.get_optional(v)
    if not ot:
        ot = op_type_svc.get_by_name_optional(v)
    if not ot:
        raise ValidationError(f"工种{v}不存在，请先在工艺管理-工种配置中维护。", field="工种")
    return {"op_type_id": ot.op_type_id, "op_type_name": ot.name}


def _build_machine_template_output(template_path: str) -> io.BytesIO:
    import openpyxl

    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    desired_headers = ["设备编号", "设备名称", "工种", "班组", "状态"]
    current_headers = [str(ws.cell(1, idx + 1).value or "").strip() for idx in range(len(desired_headers))]
    if current_headers != desired_headers:
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(desired_headers)
        ws.append(["CNC-01", "数控车床1", "数车", "车工一组", "active"])
    elif ws.max_row < 2:
        ws.append(["CNC-01", "数控车床1", "数车", "车工一组", "active"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _parse_preview_rows_json(raw_rows_json: str) -> List[Dict[str, Any]]:
    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
        return rows
    except Exception as e:
        raise ValidationError("预览数据解析失败，请重新上传并预览。") from e


def _extract_error_rows(preview_rows: List[Any]) -> List[Any]:
    return [pr for pr in (preview_rows or []) if getattr(pr, "status", None) == RowStatus.ERROR]


def _format_error_sample(error_rows: List[Any]) -> str:
    items = [f"第{pr.row_num}行：{pr.message}" for pr in (error_rows or [])[:5] if pr and getattr(pr, "message", None)]
    return "；".join(items)


def _render_excel_machine_page(
    *,
    existing: Dict[str, Dict[str, Any]],
    preview_rows: Any,
    raw_rows_json: Optional[str],
    mode_value: str,
    filename: Optional[str],
):
    return render_template(
        "equipment/excel_import_machine.html",
        title="设备信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=raw_rows_json,
        mode=mode_value,
        filename=filename,
        preview_url=url_for("equipment.excel_machine_preview"),
        confirm_url=url_for("equipment.excel_machine_confirm"),
        template_download_url=url_for("equipment.excel_machine_template"),
        export_url=url_for("equipment.excel_machine_export"),
    )


@bp.get("/excel/machines")
def excel_machine_page():
    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel()
    return render_template(
        "equipment/excel_import_machine.html",
        title="设备信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("equipment.excel_machine_preview"),
        confirm_url=url_for("equipment.excel_machine_confirm"),
        template_download_url=url_for("equipment.excel_machine_template"),
        export_url=url_for("equipment.excel_machine_export"),
    )


@bp.post("/excel/machines/preview")
def excel_machine_preview():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="设备编号")

    op_type_svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    team_svc = ResourceTeamService(g.db, op_logger=getattr(g, "op_logger", None))

    normalized_rows: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        if "状态" in item:
            item["状态"] = _normalize_machine_status_for_excel(item.get("状态"))
        try:
            ot = _resolve_op_type(item.get("工种"), op_type_svc=op_type_svc)
            if ot.get("op_type_name") is not None:
                item["工种"] = ot.get("op_type_name")
        except ValidationError:
            pass
        if "班组" in item:
            try:
                item["班组"] = team_svc.resolve_team_name_optional(item.get("班组"))
            except ValidationError:
                pass
        normalized_rows.append(item)

    m_svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = m_svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        err = _validate_machine_excel_row(row)
        if err:
            return err

        v = row.get("工种")
        if v is not None and str(v).strip() != "":
            try:
                ot = _resolve_op_type(v, op_type_svc=op_type_svc)
                row["工种"] = ot.get("op_type_name")
            except ValidationError as e:
                return e.message
        if "班组" not in row:
            return None
        try:
            row["班组"] = team_svc.resolve_team_name_optional(row.get("班组"))
        except ValidationError as e:
            return e.message
        return None

    svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = svc.preview_import(
        rows=normalized_rows,
        id_column="设备编号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return _render_excel_machine_page(
        existing=existing,
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(normalized_rows, ensure_ascii=False),
        mode_value=mode.value,
        filename=file.filename,
    )


@bp.post("/excel/machines/confirm")
def excel_machine_confirm():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")
    rows = _parse_preview_rows_json(raw_rows_json)

    _ensure_unique_ids(rows, id_column="设备编号")

    op_type_svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    team_svc = ResourceTeamService(g.db, op_logger=getattr(g, "op_logger", None))
    m_svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = m_svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        err = _validate_machine_excel_row(row)
        if err:
            return err
        v = row.get("工种")
        if v is not None and str(v).strip() != "":
            try:
                ot = _resolve_op_type(v, op_type_svc=op_type_svc)
                row["工种"] = ot.get("op_type_name")
            except ValidationError as e:
                return e.message
        if "班组" not in row:
            return None
        try:
            row["班组"] = team_svc.resolve_team_name_optional(row.get("班组"))
        except ValidationError as e:
            return e.message
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="设备编号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    error_rows = _extract_error_rows(preview_rows)
    if error_rows:
        sample = _format_error_sample(error_rows)
        message = f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。"
        if sample:
            message += f"错误示例：{sample}"
        flash(
            message,
            "error",
        )
        return _render_excel_machine_page(
            existing=existing,
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode_value=mode.value,
            filename=filename,
        )

    import_svc = MachineExcelImportService(
        g.db,
        logger=getattr(g, "app_logger", None),
        op_logger=getattr(g, "op_logger", None),
    )
    import_stats = import_svc.apply_preview_rows(preview_rows, mode=mode, existing_ids=set(existing.keys()))
    new_count = int(import_stats.get("new_count", 0))
    update_count = int(import_stats.get("update_count", 0))
    skip_count = int(import_stats.get("skip_count", 0))
    error_count = int(import_stats.get("error_count", 0))

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        filename=filename,
        mode=mode,
        preview_or_result=import_stats,
        time_cost_ms=time_cost_ms,
    )

    flash(
        f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。",
        "success",
    )
    return redirect(url_for("equipment.excel_machine_page"))


@bp.get("/excel/machines/template")
def excel_machine_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "设备信息.xlsx")
    output = _build_machine_template_output(template_path)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        template_or_export_type="设备信息模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )
    return send_file(
        output,
        as_attachment=True,
        download_name="设备信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/machines/export")
def excel_machine_export():
    start = time.time()
    m_svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    rows = m_svc.list_for_export()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["设备编号", "设备名称", "工种", "班组", "状态"])
    for r in rows:
        ws.append([r["machine_id"], r["name"], r.get("op_type_name"), r.get("team_name"), r["status"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        template_or_export_type="设备信息导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="设备信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
