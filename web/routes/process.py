from __future__ import annotations

import io
import json
import os
import tempfile
import time
from typing import Any, Dict, List, Optional

from flask import Blueprint, current_app, flash, g, redirect, request, send_file, url_for

from web.ui_mode import render_ui_template as render_template

from core.infrastructure.errors import AppError, ErrorCode, ValidationError
from core.infrastructure.transaction import TransactionManager
from core.services.common.excel_audit import log_excel_export, log_excel_import
from core.services.common.excel_backend_factory import get_excel_backend
from core.services.common.excel_service import ExcelService, ImportMode, RowStatus
from core.services.process import ExternalGroupService, OpTypeService, PartService, SupplierService
from data.repositories import OpTypeRepository, SupplierRepository


bp = Blueprint("process", __name__)


def _merge_mode_zh(value: str) -> str:
    if value == "merged":
        return "合并设置"
    return "分别设置"


def _source_zh(value: str) -> str:
    if value == "external":
        return "外部"
    return "内部"


def _safe_float(value: Any, field: str) -> Optional[float]:
    if value is None:
        return None
    v = str(value).strip()
    if v == "":
        return None
    try:
        return float(v)
    except Exception:
        raise ValidationError(f"“{field}”必须是数字", field=field)


def _parse_mode(value: str) -> ImportMode:
    try:
        return ImportMode(value)
    except Exception:
        raise ValidationError("导入模式不合法", field="mode")


def _ensure_unique_ids(rows: List[Dict[str, Any]], id_column: str) -> None:
    seen = set()
    dup = set()
    for r in rows:
        v = r.get(id_column)
        if v is None:
            continue
        key = str(v).strip()
        if not key:
            continue
        if key in seen:
            dup.add(key)
        seen.add(key)
    if dup:
        sample = ", ".join(list(sorted(dup))[:10])
        raise ValidationError(f"Excel 中存在重复的“{id_column}”：{sample}。请去重后再导入。", field=id_column)


def _read_uploaded_xlsx(file_storage) -> List[Dict[str, Any]]:
    """
    把上传的 Excel（.xlsx）解析为 List[Dict]（key 为表头字符串）。
    - 跳过空行
    - 单元格字符串自动 strip；空串视为 None
    """
    data = file_storage.read()
    if not data:
        raise AppError(ErrorCode.EXCEL_FORMAT_ERROR, "上传文件为空，请重新选择。")

    fd, tmp_path = tempfile.mkstemp(prefix="aps_upload_", suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        backend = get_excel_backend()
        return backend.read(tmp_path)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


@bp.get("/")
def list_parts():
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    parts = svc.list()

    view_rows: List[Dict[str, Any]] = []
    for p in parts:
        route_short = (p.route_raw or "").strip()
        if len(route_short) > 48:
            route_short = route_short[:48] + "..."
        view_rows.append(
            {
                "part_no": p.part_no,
                "part_name": p.part_name,
                "route_raw": p.route_raw,
                "route_short": route_short if route_short else None,
                "route_parsed": p.route_parsed,
                "route_parsed_zh": "已解析" if p.route_parsed == "yes" else "未解析",
            }
        )

    return render_template("process/list.html", title="工艺管理", parts=view_rows)


@bp.post("/parts/create")
def create_part():
    part_no = request.form.get("part_no")
    part_name = request.form.get("part_name")
    route_raw = request.form.get("route_raw")
    remark = request.form.get("remark")

    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    p = svc.create(part_no=part_no, part_name=part_name, route_raw=route_raw, remark=remark)
    flash(f"已创建零件：{p.part_no} {p.part_name}", "success")
    return redirect(url_for("process.part_detail", part_no=p.part_no))


@bp.get("/parts/<part_no>")
def part_detail(part_no: str):
    p_svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    g_svc = ExternalGroupService(g.db, op_logger=getattr(g, "op_logger", None))

    detail = p_svc.get_template_detail(part_no)
    part = detail["part"].to_dict()
    ops = [o.to_dict() for o in detail["operations"]]
    groups = [gr.to_dict() for gr in detail["groups"]]

    # 统计
    active_ops = [o for o in ops if o.get("status") == "active"]
    internal_count = sum(1 for o in active_ops if o.get("source") == "internal")
    external_count = sum(1 for o in active_ops if o.get("source") == "external")
    total_count = len(active_ops)

    # 供应商名称映射（用于展示）
    suppliers = {s.supplier_id: s for s in SupplierService(g.db).list()}  # type: ignore[arg-type]

    # deletable group（按规则控制按钮）
    deletable_group_ids = set(p_svc.calc_deletable_external_group_ids(part_no))

    # 组内工序列表
    ops_by_group: Dict[str, List[Dict[str, Any]]] = {}
    for o in active_ops:
        gid = o.get("ext_group_id")
        if not gid:
            continue
        ops_by_group.setdefault(gid, []).append(o)
    for gid in ops_by_group:
        ops_by_group[gid].sort(key=lambda x: int(x.get("seq") or 0))

    # 组对象映射
    group_map = {g["group_id"]: g for g in groups}

    return render_template(
        "process/detail.html",
        title=f"工序模板 - {part.get('part_no')} {part.get('part_name')}",
        part=part,
        operations=ops,
        active_operations=active_ops,
        total_count=total_count,
        internal_count=internal_count,
        external_count=external_count,
        groups=groups,
        group_map=group_map,
        ops_by_group=ops_by_group,
        suppliers_map={k: v.to_dict() for k, v in suppliers.items()},
        deletable_group_ids=deletable_group_ids,
        source_zh=_source_zh,
        merge_mode_zh=_merge_mode_zh,
    )


@bp.post("/parts/<part_no>/update")
def update_part(part_no: str):
    part_name = request.form.get("part_name")
    route_raw = request.form.get("route_raw")
    remark = request.form.get("remark")
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update(part_no=part_no, part_name=part_name, route_raw=route_raw, remark=remark)
    flash("零件信息已保存。", "success")
    return redirect(url_for("process.part_detail", part_no=part_no))


@bp.post("/parts/<part_no>/delete")
def delete_part(part_no: str):
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    try:
        svc.delete(part_no)
        flash(f"已删除零件：{part_no}", "success")
    except AppError as e:
        flash(e.message, "error")
    return redirect(url_for("process.list_parts"))


@bp.post("/parts/bulk/delete")
def bulk_delete_parts():
    part_nos = request.form.getlist("part_nos")
    if not part_nos:
        flash("请至少选择 1 个零件。", "error")
        return redirect(url_for("process.list_parts"))

    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    ok = 0
    failed: List[str] = []
    for pn in part_nos:
        try:
            svc.delete(pn)
            ok += 1
        except Exception:
            failed.append(str(pn))
            continue

    flash(f"批量删除完成：成功 {ok}，失败 {len(failed)}。", "success" if ok else "warning")
    if failed:
        sample = "，".join(failed[:10])
        flash(f"删除失败（最多展示 10 个）：{sample}。常见原因：已被批次引用，请先删除/调整批次或停止引用。", "warning")
    return redirect(url_for("process.list_parts"))


@bp.post("/parts/<part_no>/reparse")
def reparse_part(part_no: str):
    route_raw = request.form.get("route_raw")
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))

    start = time.time()
    result = svc.reparse_and_save(part_no=part_no, route_raw=route_raw)
    ms = int((time.time() - start) * 1000)

    warn_text = f"，警告 {len(result.warnings)} 条" if result.warnings else ""
    flash(f"工艺路线解析完成：共 {result.stats.get('total', 0)} 道工序（内部 {result.stats.get('internal', 0)}，外部 {result.stats.get('external', 0)}）{warn_text}。耗时 {ms} ms。", "success")
    return redirect(url_for("process.part_detail", part_no=part_no))


@bp.post("/parts/<part_no>/ops/<int:seq>/hours")
def update_internal_hours(part_no: str, seq: int):
    setup_hours = request.form.get("setup_hours")
    unit_hours = request.form.get("unit_hours")
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update_internal_hours(part_no=part_no, seq=seq, setup_hours=setup_hours, unit_hours=unit_hours)
    flash("工序工时已保存。", "success")
    return redirect(url_for("process.part_detail", part_no=part_no))


@bp.post("/parts/<part_no>/groups/<group_id>/mode")
def set_group_mode(part_no: str, group_id: str):
    merge_mode = request.form.get("merge_mode") or "separate"
    total_days = request.form.get("total_days")

    per_op_days: Dict[int, Any] = {}
    for k, v in request.form.items():
        if not k.startswith("ext_days_"):
            continue
        try:
            seq = int(k.replace("ext_days_", ""))
        except Exception:
            continue
        per_op_days[seq] = v

    svc = ExternalGroupService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.set_merge_mode(
        group_id=group_id,
        merge_mode=merge_mode,
        total_days=total_days,
        per_op_days=per_op_days,
    )
    flash(f"外部工序组周期模式已更新：{_merge_mode_zh(merge_mode)}。", "success")
    return redirect(url_for("process.part_detail", part_no=part_no))


@bp.post("/parts/<part_no>/groups/<group_id>/delete")
def delete_group(part_no: str, group_id: str):
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    result = svc.delete_external_group(part_no=part_no, group_id=group_id)
    flash(f"已删除外部工序组：{group_id}（{result.get('message')}）", "success")
    return redirect(url_for("process.part_detail", part_no=part_no))


# ============================================================
# 工种配置（页面）
# ============================================================


@bp.get("/op-types")
def op_types_page():
    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    rows = [x.to_dict() for x in svc.list()]
    return render_template("process/op_types_list.html", title="工种配置", op_types=rows)


@bp.post("/op-types/create")
def create_op_type():
    op_type_id = request.form.get("op_type_id")
    name = request.form.get("name")
    category = request.form.get("category") or "internal"
    remark = request.form.get("remark")
    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    ot = svc.create(op_type_id=op_type_id, name=name, category=category, remark=remark)
    flash(f"已创建工种：{ot.op_type_id} {ot.name}", "success")
    return redirect(url_for("process.op_type_detail", op_type_id=ot.op_type_id))


@bp.get("/op-types/<op_type_id>")
def op_type_detail(op_type_id: str):
    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    ot = svc.get(op_type_id)
    return render_template("process/op_type_detail.html", title=f"工种详情 - {ot.op_type_id}", op_type=ot.to_dict())


@bp.post("/op-types/<op_type_id>/update")
def update_op_type(op_type_id: str):
    name = request.form.get("name")
    category = request.form.get("category")
    remark = request.form.get("remark")
    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update(op_type_id=op_type_id, name=name, category=category, remark=remark)
    flash("工种信息已保存。", "success")
    return redirect(url_for("process.op_type_detail", op_type_id=op_type_id))


@bp.post("/op-types/<op_type_id>/delete")
def delete_op_type(op_type_id: str):
    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.delete(op_type_id)
    flash("已删除工种。", "success")
    return redirect(url_for("process.op_types_page"))


# ============================================================
# 供应商配置（页面）
# ============================================================


def _supplier_status_zh(status: str) -> str:
    if status == "inactive":
        return "停用"
    return "启用"


@bp.get("/suppliers")
def suppliers_page():
    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    rows = [x.to_dict() for x in svc.list()]

    op_types = {ot.op_type_id: ot for ot in OpTypeService(g.db).list()}  # type: ignore[arg-type]
    view_rows = []
    for r in rows:
        ot = op_types.get(r.get("op_type_id") or "")
        view_rows.append(
            {
                **r,
                "status_zh": _supplier_status_zh(r.get("status") or "active"),
                "op_type_name": (ot.name if ot else None),
            }
        )

    op_type_options = [(ot.op_type_id, ot.name) for ot in sorted(op_types.values(), key=lambda x: x.name)]
    return render_template(
        "process/suppliers_list.html",
        title="供应商配置",
        suppliers=view_rows,
        op_type_options=op_type_options,
        status_options=[("active", "启用"), ("inactive", "停用")],
    )


@bp.post("/suppliers/create")
def create_supplier():
    supplier_id = request.form.get("supplier_id")
    name = request.form.get("name")
    op_type_id = request.form.get("op_type_id") or None
    default_days = request.form.get("default_days") or "1"
    status = request.form.get("status") or "active"
    remark = request.form.get("remark")

    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    s = svc.create(
        supplier_id=supplier_id,
        name=name,
        op_type_value=op_type_id,
        default_days=default_days,
        status=status,
        remark=remark,
    )
    flash(f"已创建供应商：{s.supplier_id} {s.name}", "success")
    return redirect(url_for("process.supplier_detail", supplier_id=s.supplier_id))


@bp.get("/suppliers/<supplier_id>")
def supplier_detail(supplier_id: str):
    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    s = svc.get(supplier_id)
    op_types = {ot.op_type_id: ot for ot in OpTypeService(g.db).list()}  # type: ignore[arg-type]
    op_type_options = [(ot.op_type_id, ot.name) for ot in sorted(op_types.values(), key=lambda x: x.name)]

    return render_template(
        "process/supplier_detail.html",
        title=f"供应商详情 - {s.supplier_id}",
        supplier=s.to_dict(),
        op_type_name=(op_types.get(s.op_type_id or "")).name if s.op_type_id and op_types.get(s.op_type_id) else None,
        op_type_options=op_type_options,
        status_options=[("active", "启用"), ("inactive", "停用")],
        supplier_status_zh=_supplier_status_zh(s.status),
    )


@bp.post("/suppliers/<supplier_id>/update")
def update_supplier(supplier_id: str):
    name = request.form.get("name")
    op_type_id = request.form.get("op_type_id")
    default_days = request.form.get("default_days")
    status = request.form.get("status")
    remark = request.form.get("remark")

    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update(
        supplier_id=supplier_id,
        name=name,
        op_type_value=op_type_id,
        default_days=default_days,
        status=status,
        remark=remark,
    )
    flash("供应商信息已保存。", "success")
    return redirect(url_for("process.supplier_detail", supplier_id=supplier_id))


@bp.post("/suppliers/<supplier_id>/delete")
def delete_supplier(supplier_id: str):
    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.delete(supplier_id)
    flash("已删除供应商。", "success")
    return redirect(url_for("process.suppliers_page"))


# ============================================================
# Excel：工种配置（OpTypes）
# ============================================================


def _normalize_op_type_category(value: Any) -> str:
    v = "" if value is None else str(value).strip()
    if v in ("内部", "内", "internal"):
        return "internal"
    if v in ("外部", "外", "external"):
        return "external"
    return v


@bp.get("/excel/op-types")
def excel_op_type_page():
    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel()
    return render_template(
        "process/excel_import_op_types.html",
        title="工种配置 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("process.excel_op_type_preview"),
        confirm_url=url_for("process.excel_op_type_confirm"),
        template_download_url=url_for("process.excel_op_type_template"),
        export_url=url_for("process.excel_op_type_export"),
    )


@bp.post("/excel/op-types/preview")
def excel_op_type_preview():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="工种ID")

    svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        if not row.get("工种ID") or str(row.get("工种ID")).strip() == "":
            return "“工种ID”不能为空"
        if not row.get("工种名称") or str(row.get("工种名称")).strip() == "":
            return "“工种名称”不能为空"
        cat = _normalize_op_type_category(row.get("归属") or "internal")
        if not cat:
            cat = "internal"
        if cat not in ("internal", "external"):
            return "“归属”不合法（允许：internal / external；或中文：内部/外部）"
        row["归属"] = cat
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="工种ID",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="op_type",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return render_template(
        "process/excel_import_op_types.html",
        title="工种配置 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("process.excel_op_type_preview"),
        confirm_url=url_for("process.excel_op_type_confirm"),
        template_download_url=url_for("process.excel_op_type_template"),
        export_url=url_for("process.excel_op_type_export"),
    )


@bp.post("/excel/op-types/confirm")
def excel_op_type_confirm():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    _ensure_unique_ids(rows, id_column="工种ID")

    op_type_svc = OpTypeService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_type_svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        if not row.get("工种ID") or str(row.get("工种ID")).strip() == "":
            return "“工种ID”不能为空"
        if not row.get("工种名称") or str(row.get("工种名称")).strip() == "":
            return "“工种名称”不能为空"
        cat = _normalize_op_type_category(row.get("归属") or "internal")
        if not cat:
            cat = "internal"
        if cat not in ("internal", "external"):
            return "“归属”不合法（允许：internal / external；或中文：内部/外部）"
        row["归属"] = cat
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="工种ID",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )
        return render_template(
            "process/excel_import_op_types.html",
            title="工种配置 - Excel 导入/导出",
            existing_list=list(existing.values()),
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("process.excel_op_type_preview"),
            confirm_url=url_for("process.excel_op_type_confirm"),
            template_download_url=url_for("process.excel_op_type_template"),
            export_url=url_for("process.excel_op_type_export"),
        )

    tx = TransactionManager(g.db)
    op_repo = OpTypeRepository(g.db)

    new_count = update_count = skip_count = error_count = 0
    errors_sample: List[Dict[str, Any]] = []

    with tx.transaction():
        if mode == ImportMode.REPLACE:
            op_type_svc.ensure_replace_allowed()
            g.db.execute("DELETE FROM OpTypes")

        for pr in preview_rows:
            if pr.status == RowStatus.ERROR:
                error_count += 1
                if pr.message and len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": pr.message})
                continue

            ot_id = str(pr.data.get("工种ID")).strip()
            name = str(pr.data.get("工种名称")).strip()
            cat = _normalize_op_type_category(pr.data.get("归属") or "internal") or "internal"

            if mode == ImportMode.APPEND and ot_id in existing:
                skip_count += 1
                continue

            try:
                if op_repo.get(ot_id):
                    op_repo.update(ot_id, {"name": name, "category": cat})
                    update_count += 1
                else:
                    op_repo.create({"op_type_id": ot_id, "name": name, "category": cat})
                    new_count += 1
            except AppError as e:
                error_count += 1
                if len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": e.message})
                continue

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="op_type",
        filename=filename,
        mode=mode,
        preview_or_result={
            "total_rows": len(preview_rows),
            "new_count": new_count,
            "update_count": update_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors_sample": errors_sample,
        },
        time_cost_ms=time_cost_ms,
    )

    flash(f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。", "success")
    return redirect(url_for("process.excel_op_type_page"))


@bp.get("/excel/op-types/template")
def excel_op_type_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "工种配置.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="process",
            target_type="op_type",
            template_or_export_type="工种配置模板.xlsx",
            filters={},
            row_count=2,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="工种配置.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工种ID", "工种名称", "归属"])
    ws.append(["OT001", "数车", "internal"])
    ws.append(["OT002", "标印", "external"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="op_type",
        template_or_export_type="工种配置模板.xlsx",
        filters={},
        row_count=2,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="工种配置.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/op-types/export")
def excel_op_type_export():
    start = time.time()
    rows = g.db.execute("SELECT op_type_id, name, category FROM OpTypes ORDER BY name").fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工种ID", "工种名称", "归属"])
    for r in rows:
        ws.append([r["op_type_id"], r["name"], r["category"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="op_type",
        template_or_export_type="工种配置导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="工种配置.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# Excel：供应商配置（Suppliers）
# ============================================================


def _normalize_supplier_status(value: Any) -> str:
    v = "" if value is None else str(value).strip()
    if v in ("启用", "在用", "正常", "active"):
        return "active"
    if v in ("停用", "禁用", "inactive"):
        return "inactive"
    return v or "active"


def _resolve_op_type_name(value: Any, op_type_repo: OpTypeRepository) -> Optional[str]:
    v = None if value is None else str(value).strip()
    if not v:
        return None
    ot = op_type_repo.get(v)
    if not ot:
        ot = op_type_repo.get_by_name(v)
    if not ot:
        raise ValidationError(f"工种“{v}”不存在，请先维护工种配置。", field="对应工种")
    return ot.name


def _resolve_op_type_id(value: Any, op_type_repo: OpTypeRepository) -> Optional[str]:
    v = None if value is None else str(value).strip()
    if not v:
        return None
    ot = op_type_repo.get(v)
    if not ot:
        ot = op_type_repo.get_by_name(v)
    if not ot:
        raise ValidationError(f"工种“{v}”不存在，请先维护工种配置。", field="对应工种")
    return ot.op_type_id


@bp.get("/excel/suppliers")
def excel_supplier_page():
    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel()
    return render_template(
        "process/excel_import_suppliers.html",
        title="供应商配置 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("process.excel_supplier_preview"),
        confirm_url=url_for("process.excel_supplier_confirm"),
        template_download_url=url_for("process.excel_supplier_template"),
        export_url=url_for("process.excel_supplier_export"),
    )


@bp.post("/excel/suppliers/preview")
def excel_supplier_preview():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="供应商ID")

    svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel()
    op_type_repo = OpTypeRepository(g.db)

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        if not row.get("供应商ID") or str(row.get("供应商ID")).strip() == "":
            return "“供应商ID”不能为空"
        if not row.get("名称") or str(row.get("名称")).strip() == "":
            return "“名称”不能为空"

        # 默认周期
        if row.get("默认周期") is None or str(row.get("默认周期")).strip() == "":
            row["默认周期"] = 1.0
        try:
            d = float(row.get("默认周期"))
            if d <= 0:
                return "“默认周期”必须大于 0"
            row["默认周期"] = d
        except Exception:
            return "“默认周期”必须是数字"

        # 状态可选
        if "状态" in row:
            row["状态"] = _normalize_supplier_status(row.get("状态"))
            if row["状态"] not in ("active", "inactive"):
                return "“状态”不合法（允许：active / inactive；或中文：启用/停用）"
        else:
            row["状态"] = "active"

        # 工种可选（允许 id 或 名称），预览阶段标准化为“名称”
        try:
            name = _resolve_op_type_name(row.get("对应工种"), op_type_repo=op_type_repo)
            row["对应工种"] = name
        except ValidationError as e:
            return e.message

        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="供应商ID",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="supplier",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return render_template(
        "process/excel_import_suppliers.html",
        title="供应商配置 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("process.excel_supplier_preview"),
        confirm_url=url_for("process.excel_supplier_confirm"),
        template_download_url=url_for("process.excel_supplier_template"),
        export_url=url_for("process.excel_supplier_export"),
    )


@bp.post("/excel/suppliers/confirm")
def excel_supplier_confirm():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    _ensure_unique_ids(rows, id_column="供应商ID")

    supplier_svc = SupplierService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = supplier_svc.build_existing_for_excel()
    op_type_repo = OpTypeRepository(g.db)

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        if not row.get("供应商ID") or str(row.get("供应商ID")).strip() == "":
            return "“供应商ID”不能为空"
        if not row.get("名称") or str(row.get("名称")).strip() == "":
            return "“名称”不能为空"
        if row.get("默认周期") is None or str(row.get("默认周期")).strip() == "":
            row["默认周期"] = 1.0
        try:
            d = float(row.get("默认周期"))
            if d <= 0:
                return "“默认周期”必须大于 0"
            row["默认周期"] = d
        except Exception:
            return "“默认周期”必须是数字"
        row["状态"] = _normalize_supplier_status(row.get("状态"))
        if row["状态"] not in ("active", "inactive"):
            return "“状态”不合法（允许：active / inactive；或中文：启用/停用）"
        try:
            name = _resolve_op_type_name(row.get("对应工种"), op_type_repo=op_type_repo)
            row["对应工种"] = name
        except ValidationError as e:
            return e.message
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="供应商ID",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )
        return render_template(
            "process/excel_import_suppliers.html",
            title="供应商配置 - Excel 导入/导出",
            existing_list=list(existing.values()),
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("process.excel_supplier_preview"),
            confirm_url=url_for("process.excel_supplier_confirm"),
            template_download_url=url_for("process.excel_supplier_template"),
            export_url=url_for("process.excel_supplier_export"),
        )

    tx = TransactionManager(g.db)
    s_repo = SupplierRepository(g.db)

    new_count = update_count = skip_count = error_count = 0
    errors_sample: List[Dict[str, Any]] = []

    with tx.transaction():
        if mode == ImportMode.REPLACE:
            supplier_svc.ensure_replace_allowed()
            g.db.execute("DELETE FROM Suppliers")

        for pr in preview_rows:
            if pr.status == RowStatus.ERROR:
                error_count += 1
                if pr.message and len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": pr.message})
                continue

            sid = str(pr.data.get("供应商ID")).strip()
            name = str(pr.data.get("名称")).strip()
            op_type_id = None
            try:
                op_type_id = _resolve_op_type_id(pr.data.get("对应工种"), op_type_repo=op_type_repo)
            except ValidationError:
                op_type_id = None
            default_days = float(pr.data.get("默认周期") or 1.0)
            status = _normalize_supplier_status(pr.data.get("状态"))
            remark = pr.data.get("备注")

            if mode == ImportMode.APPEND and sid in existing:
                skip_count += 1
                continue

            try:
                if s_repo.get(sid):
                    s_repo.update(
                        sid,
                        {
                            "name": name,
                            "op_type_id": op_type_id,
                            "default_days": default_days,
                            "status": status,
                            "remark": remark,
                        },
                    )
                    update_count += 1
                else:
                    s_repo.create(
                        {
                            "supplier_id": sid,
                            "name": name,
                            "op_type_id": op_type_id,
                            "default_days": default_days,
                            "status": status,
                            "remark": remark,
                        }
                    )
                    new_count += 1
            except AppError as e:
                error_count += 1
                if len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": e.message})
                continue

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="supplier",
        filename=filename,
        mode=mode,
        preview_or_result={
            "total_rows": len(preview_rows),
            "new_count": new_count,
            "update_count": update_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors_sample": errors_sample,
        },
        time_cost_ms=time_cost_ms,
    )

    flash(f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。", "success")
    return redirect(url_for("process.excel_supplier_page"))


@bp.get("/excel/suppliers/template")
def excel_supplier_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "供应商配置.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="process",
            target_type="supplier",
            template_or_export_type="供应商配置模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="供应商配置.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["供应商ID", "名称", "对应工种", "默认周期"])
    ws.append(["S001", "外协-标印厂", "标印", 1])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="supplier",
        template_or_export_type="供应商配置模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="供应商配置.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/suppliers/export")
def excel_supplier_export():
    start = time.time()
    rows = g.db.execute(
        """
        SELECT s.supplier_id, s.name, s.default_days, s.status, s.remark, ot.name AS op_type_name
        FROM Suppliers s
        LEFT JOIN OpTypes ot ON ot.op_type_id = s.op_type_id
        ORDER BY s.supplier_id
        """
    ).fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["供应商ID", "名称", "对应工种", "默认周期", "状态", "备注"])
    for r in rows:
        ws.append([r["supplier_id"], r["name"], r["op_type_name"], r["default_days"], r["status"], r["remark"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="supplier",
        template_or_export_type="供应商配置导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="供应商配置.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# Excel：零件工艺路线（Parts.route_raw + 解析生成模板）
# ============================================================


@bp.get("/excel/routes")
def excel_routes_page():
    svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel_routes()
    return render_template(
        "process/excel_import_routes.html",
        title="零件工艺路线 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("process.excel_routes_preview"),
        confirm_url=url_for("process.excel_routes_confirm"),
        template_download_url=url_for("process.excel_routes_template"),
        export_url=url_for("process.excel_routes_export"),
    )


@bp.post("/excel/routes/preview")
def excel_routes_preview():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="图号")

    part_svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = part_svc.build_existing_for_excel_routes()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        if not row.get("图号") or str(row.get("图号")).strip() == "":
            return "“图号”不能为空"
        if not row.get("名称") or str(row.get("名称")).strip() == "":
            return "“名称”不能为空"
        route_raw = row.get("工艺路线字符串")
        if route_raw is None or str(route_raw).strip() == "":
            return "“工艺路线字符串”不能为空"
        ok, msg = part_svc.validate_route_format(route_raw)
        if not ok:
            return f"工艺路线格式不合法：{msg}"
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="图号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="part_route",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return render_template(
        "process/excel_import_routes.html",
        title="零件工艺路线 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("process.excel_routes_preview"),
        confirm_url=url_for("process.excel_routes_confirm"),
        template_download_url=url_for("process.excel_routes_template"),
        export_url=url_for("process.excel_routes_export"),
    )


@bp.post("/excel/routes/confirm")
def excel_routes_confirm():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    _ensure_unique_ids(rows, id_column="图号")

    part_svc = PartService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = part_svc.build_existing_for_excel_routes()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        if not row.get("图号") or str(row.get("图号")).strip() == "":
            return "“图号”不能为空"
        if not row.get("名称") or str(row.get("名称")).strip() == "":
            return "“名称”不能为空"
        route_raw = row.get("工艺路线字符串")
        if route_raw is None or str(route_raw).strip() == "":
            return "“工艺路线字符串”不能为空"
        ok, msg = part_svc.validate_route_format(route_raw)
        if not ok:
            return f"工艺路线格式不合法：{msg}"
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="图号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )
        return render_template(
            "process/excel_import_routes.html",
            title="零件工艺路线 - Excel 导入/导出",
            existing_list=list(existing.values()),
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("process.excel_routes_preview"),
            confirm_url=url_for("process.excel_routes_confirm"),
            template_download_url=url_for("process.excel_routes_template"),
            export_url=url_for("process.excel_routes_export"),
        )

    tx = TransactionManager(g.db)
    new_count = update_count = skip_count = error_count = 0
    errors_sample: List[Dict[str, Any]] = []

    with tx.transaction():
        if mode == ImportMode.REPLACE:
            # 若存在批次引用，删除 Parts 会触发外键错误，因此这里做保护
            row = g.db.execute("SELECT 1 FROM Batches LIMIT 1").fetchone()
            if row is not None:
                raise ValidationError("已存在批次数据，不能执行“替换（清空后导入）”。请改用“覆盖/追加”。")
            g.db.execute("DELETE FROM Parts")

        for pr in preview_rows:
            if pr.status == RowStatus.ERROR:
                error_count += 1
                if pr.message and len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": pr.message})
                continue

            pn = str(pr.data.get("图号")).strip()
            name = str(pr.data.get("名称")).strip()
            route_raw = str(pr.data.get("工艺路线字符串")).strip()

            if mode == ImportMode.APPEND and pn in existing:
                skip_count += 1
                continue

            try:
                existed = pn in existing
                part_svc.upsert_and_parse_no_tx(part_no=pn, part_name=name, route_raw=route_raw)
                if existed:
                    update_count += 1
                else:
                    new_count += 1
            except AppError as e:
                error_count += 1
                if len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": e.message})
                continue

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="part_route",
        filename=filename,
        mode=mode,
        preview_or_result={
            "total_rows": len(preview_rows),
            "new_count": new_count,
            "update_count": update_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors_sample": errors_sample,
        },
        time_cost_ms=time_cost_ms,
    )

    flash(f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。", "success")
    return redirect(url_for("process.excel_routes_page"))


@bp.get("/excel/routes/template")
def excel_routes_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "零件工艺路线.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="process",
            target_type="part_route",
            template_or_export_type="零件工艺路线模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="零件工艺路线.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["图号", "名称", "工艺路线字符串"])
    ws.append(["A1234", "壳体-大", "5数铣10钳20数车35标印40总检45表处理"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="part_route",
        template_or_export_type="零件工艺路线模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="零件工艺路线.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/routes/export")
def excel_routes_export():
    start = time.time()
    rows = g.db.execute("SELECT part_no, part_name, route_raw FROM Parts ORDER BY part_no").fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["图号", "名称", "工艺路线字符串"])
    for r in rows:
        ws.append([r["part_no"], r["part_name"], r["route_raw"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="part_route",
        template_or_export_type="零件工艺路线导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="零件工艺路线.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# Excel：零件工序模板导出（PartOperations）
# ============================================================


@bp.get("/excel/part-operations")
def excel_part_ops_page():
    return render_template(
        "process/excel_part_ops_export.html",
        title="零件工序模板 - Excel 导出",
        export_url=url_for("process.excel_part_ops_export"),
    )


@bp.get("/excel/part-operations/export")
def excel_part_ops_export():
    start = time.time()
    rows = g.db.execute(
        """
        SELECT
          p.part_no,
          po.seq,
          po.op_type_name,
          po.source,
          po.supplier_id,
          s.name AS supplier_name,
          po.ext_days,
          po.ext_group_id,
          eg.merge_mode,
          eg.total_days
        FROM PartOperations po
        JOIN Parts p ON p.part_no = po.part_no
        LEFT JOIN Suppliers s ON s.supplier_id = po.supplier_id
        LEFT JOIN ExternalGroups eg ON eg.group_id = po.ext_group_id
        WHERE po.status = 'active'
        ORDER BY p.part_no, po.seq
        """
    ).fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["图号", "工序", "工种", "归属", "供应商", "周期"])
    for r in rows:
        supplier = r["supplier_name"] or ""
        days = None
        if r["source"] == "external":
            if r["merge_mode"] == "merged" and r["total_days"] is not None:
                days = r["total_days"]
            else:
                days = r["ext_days"]
        ws.append(
            [
                r["part_no"],
                r["seq"],
                r["op_type_name"],
                r["source"],
                supplier,
                days,
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="process",
        target_type="part_operation",
        template_or_export_type="零件工序模板导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="零件工序模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

