from __future__ import annotations

import io
import json
import os
import tempfile
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import Blueprint, current_app, flash, g, redirect, render_template, request, send_file, url_for

from core.infrastructure.errors import AppError, ErrorCode, ValidationError
from core.infrastructure.transaction import TransactionManager
from core.services.common.excel_audit import log_excel_export, log_excel_import
from core.services.common.excel_backend_factory import get_excel_backend
from core.services.common.excel_service import ExcelService, ImportMode, RowStatus
from core.services.equipment import MachineDowntimeService, MachineService
from core.services.personnel import OperatorMachineService
from data.repositories import MachineRepository, OpTypeRepository, OperatorRepository


bp = Blueprint("equipment", __name__)


def _machine_status_zh(status: str) -> str:
    if status == "active":
        return "可用"
    if status == "maintain":
        return "维修"
    if status == "inactive":
        return "停用"
    return status or "-"


def _operator_status_zh(status: str) -> str:
    if status == "active":
        return "在岗"
    if status == "inactive":
        return "停用/休假"
    return status or "-"


def _parse_mode(value: str) -> ImportMode:
    try:
        return ImportMode(value)
    except Exception:
        raise ValidationError("导入模式不合法", field="mode")


def _read_uploaded_xlsx(file_storage) -> List[Dict[str, Any]]:
    """
    把上传的 Excel（.xlsx）解析为 List[Dict]（key 为表头字符串）。
    - 跳过空行
    - 单元格字符串自动 strip；空串视为 None
    """
    data = file_storage.read()
    if not data:
        raise AppError(ErrorCode.EXCEL_FORMAT_ERROR, "上传文件为空，请重新选择。")

    fd, tmp_path = tempfile.mkstemp(prefix="aps_upload_", suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        backend = get_excel_backend()
        return backend.read(tmp_path)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def _ensure_unique_ids(rows: List[Dict[str, Any]], id_column: str) -> None:
    seen = set()
    dup = set()
    for r in rows:
        v = r.get(id_column)
        if v is None:
            continue
        key = str(v).strip()
        if not key:
            continue
        if key in seen:
            dup.add(key)
        seen.add(key)
    if dup:
        sample = ", ".join(list(sorted(dup))[:10])
        raise ValidationError(f"Excel 中存在重复的“{id_column}”：{sample}。请去重后再导入。", field=id_column)


@bp.get("/")
def list_page():
    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    op_type_repo = OpTypeRepository(g.db)

    machines = svc.list()
    op_types = {ot.op_type_id: ot for ot in op_type_repo.list()}

    # 自动绑定：若当前时间落在停机计划内，则页面显示“停机（计划）”
    downtime_now_set = set()
    try:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = g.db.execute(
            """
            SELECT DISTINCT machine_id
            FROM MachineDowntimes
            WHERE status = 'active'
              AND start_time <= ?
              AND end_time > ?
            """,
            (now_str, now_str),
        ).fetchall()
        downtime_now_set = {r["machine_id"] for r in rows if r and r["machine_id"]}
    except Exception:
        downtime_now_set = set()

    # 聚合关联人员
    link_rows = g.db.execute(
        """
        SELECT om.machine_id, om.operator_id, o.name AS operator_name, o.status AS operator_status
        FROM OperatorMachine om
        LEFT JOIN Operators o ON o.operator_id = om.operator_id
        ORDER BY om.machine_id, om.operator_id
        """
    ).fetchall()

    links_by_machine: Dict[str, List[Dict[str, Any]]] = {}
    for r in link_rows:
        mc_id = r["machine_id"]
        links_by_machine.setdefault(mc_id, []).append(
            {
                "operator_id": r["operator_id"],
                "operator_name": r["operator_name"],
                "operator_status": r["operator_status"],
            }
        )

    view_rows: List[Dict[str, Any]] = []
    for m in machines:
        ot = op_types.get(m.op_type_id or "")
        links = links_by_machine.get(m.machine_id, [])
        operator_text = ", ".join(
            [
                f"{x['operator_id']}{(' ' + x['operator_name']) if x.get('operator_name') else ''}".strip()
                for x in links
            ]
        )
        view_rows.append(
            {
                "machine_id": m.machine_id,
                "name": m.name,
                "op_type_id": m.op_type_id,
                "op_type_name": (ot.name if ot else None),
                "status": m.status,
                "status_zh": ("停机（计划）" if m.machine_id in downtime_now_set else _machine_status_zh(m.status)),
                "remark": m.remark,
                "operator_text": operator_text,
                "operator_count": len(links),
            }
        )

    # 工种下拉：按 name 排序
    op_type_options = [(ot.op_type_id, ot.name) for ot in sorted(op_types.values(), key=lambda x: x.name)]

    return render_template(
        "equipment/list.html",
        title="设备管理",
        machines=view_rows,
        op_type_options=op_type_options,
        status_options=[("active", "可用"), ("maintain", "维修"), ("inactive", "停用")],
    )


@bp.post("/create")
def create_machine():
    machine_id = request.form.get("machine_id")
    name = request.form.get("name")
    op_type_id = request.form.get("op_type_id") or None
    category = request.form.get("category") or None
    status = request.form.get("status") or "active"
    remark = request.form.get("remark")

    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    m = svc.create(machine_id=machine_id, name=name, op_type_id=op_type_id, category=category, status=status, remark=remark)
    flash(f"已创建设备：{m.machine_id} {m.name}", "success")
    return redirect(url_for("equipment.detail_page", machine_id=m.machine_id))


@bp.get("/<machine_id>")
def detail_page(machine_id: str):
    m_svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    link_svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    dt_svc = MachineDowntimeService(g.db, op_logger=getattr(g, "op_logger", None))

    op_type_repo = OpTypeRepository(g.db)
    operator_repo = OperatorRepository(g.db)

    m = m_svc.get(machine_id)
    links = link_svc.list_by_machine(machine_id)
    downtimes = dt_svc.list_by_machine(machine_id, include_cancelled=False)

    op_types = {ot.op_type_id: ot for ot in op_type_repo.list()}
    operators = {o.operator_id: o for o in operator_repo.list()}

    linked_operator_ids = {l.operator_id for l in links}

    linked_operators: List[Dict[str, Any]] = []
    for l in links:
        op = operators.get(l.operator_id)
        linked_operators.append(
            {
                "operator_id": l.operator_id,
                "operator_name": (op.name if op else None),
                "status": (op.status if op else None),
                "status_zh": _operator_status_zh(op.status) if op else "-",
                "skill_level": getattr(l, "skill_level", None),
                "is_primary": getattr(l, "is_primary", None),
            }
        )

    available_operators: List[Dict[str, Any]] = []
    for op in operators.values():
        if op.operator_id in linked_operator_ids:
            continue
        available_operators.append(
            {
                "operator_id": op.operator_id,
                "operator_name": op.name,
                "status": op.status,
                "status_zh": _operator_status_zh(op.status),
            }
        )
    available_operators.sort(key=lambda x: x["operator_id"])

    op_type_options = [(ot.op_type_id, ot.name) for ot in sorted(op_types.values(), key=lambda x: x.name)]

    return render_template(
        "equipment/detail.html",
        title=f"设备详情 - {m.machine_id} {m.name}",
        machine=m.to_dict(),
        machine_status_zh=_machine_status_zh(m.status),
        op_type_name=(op_types.get(m.op_type_id or "")).name if m.op_type_id and op_types.get(m.op_type_id) else None,
        op_type_options=op_type_options,
        status_options=[("active", "可用"), ("maintain", "维修"), ("inactive", "停用")],
        linked_operators=linked_operators,
        available_operators=available_operators,
        downtime_rows=[d.to_dict() for d in downtimes],
        downtime_reason_options=list(MachineDowntimeService.REASON_OPTIONS),
        skill_level_options=[("beginner", "初级"), ("normal", "普通"), ("expert", "熟练")],
    )


@bp.post("/<machine_id>/update")
def update_machine(machine_id: str):
    name = request.form.get("name")
    op_type_id = request.form.get("op_type_id")
    category = request.form.get("category")
    status = request.form.get("status")
    remark = request.form.get("remark")

    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update(machine_id=machine_id, name=name, op_type_id=op_type_id, category=category, status=status, remark=remark)
    flash("设备信息已保存。", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


@bp.post("/<machine_id>/status")
def set_status(machine_id: str):
    status = request.form.get("status")
    if not status:
        raise ValidationError("缺少状态参数", field="status")
    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    m = svc.set_status(machine_id=machine_id, status=status)
    flash(f"已更新状态：{m.machine_id} → {_machine_status_zh(m.status)}", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


@bp.post("/<machine_id>/delete")
def delete_machine(machine_id: str):
    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    try:
        svc.delete(machine_id)
        flash(f"已删除设备：{machine_id}", "success")
    except AppError as e:
        flash(e.message, "error")
    return redirect(url_for("equipment.list_page"))


@bp.post("/bulk/status")
def bulk_set_status():
    """
    批量设置设备状态（active/maintain/inactive）。
    """
    status = (request.form.get("status") or "").strip()
    machine_ids = request.form.getlist("machine_ids")
    if not machine_ids:
        flash("请至少选择 1 台设备。", "error")
        return redirect(url_for("equipment.list_page"))
    if status not in ("active", "maintain", "inactive"):
        raise ValidationError("状态不合法（允许：active / maintain / inactive）", field="status")

    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    ok = 0
    failed: List[str] = []
    for mid in machine_ids:
        try:
            svc.set_status(mid, status=status)
            ok += 1
        except Exception:
            failed.append(str(mid))
            continue

    flash(f"批量状态更新完成：成功 {ok}，失败 {len(failed)}。", "success" if ok else "warning")
    if failed:
        sample = "，".join(failed[:10])
        flash(f"失败设备（最多展示 10 个）：{sample}", "warning")
    return redirect(url_for("equipment.list_page"))


@bp.post("/bulk/delete")
def bulk_delete():
    """
    批量删除设备（受引用保护；建议优先批量“停用”）。
    """
    machine_ids = request.form.getlist("machine_ids")
    if not machine_ids:
        flash("请至少选择 1 台设备。", "error")
        return redirect(url_for("equipment.list_page"))

    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    ok = 0
    failed: List[str] = []
    for mid in machine_ids:
        try:
            svc.delete(mid)
            ok += 1
        except Exception:
            failed.append(str(mid))
            continue

    flash(f"批量删除完成：成功 {ok}，失败 {len(failed)}。", "success" if ok else "warning")
    if failed:
        sample = "，".join(failed[:10])
        flash(f"删除失败（最多展示 10 个）：{sample}。常见原因：被批次工序/排程引用，请改为“停用”。", "warning")
    return redirect(url_for("equipment.list_page"))


@bp.post("/<machine_id>/link/add")
def add_link(machine_id: str):
    operator_id = request.form.get("operator_id")
    svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.add_link(operator_id=operator_id, machine_id=machine_id)
    flash("已添加人员关联。", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


@bp.post("/<machine_id>/link/update")
def update_link(machine_id: str):
    operator_id = request.form.get("operator_id")
    skill_level = request.form.get("skill_level")
    is_primary = "yes" if request.form.get("is_primary") else "no"
    svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update_link_fields(operator_id=operator_id, machine_id=machine_id, skill_level=skill_level, is_primary=is_primary)
    flash("已更新关联字段（技能等级/主操设备）。", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


@bp.post("/<machine_id>/link/remove")
def remove_link(machine_id: str):
    operator_id = request.form.get("operator_id")
    svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.remove_link(operator_id=operator_id, machine_id=machine_id)
    flash("已解除人员关联。", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


# ============================================================
# 设备停机（MachineDowntimes）
# ============================================================


@bp.get("/downtimes/batch")
def downtime_batch_page():
    """
    批量停机计划：按设备/类别/全部创建停机区间。
    """
    repo = MachineRepository(g.db)
    machines = repo.list()
    categories = sorted({(m.category or "").strip() for m in machines if (m.category or "").strip()})
    machine_options = [(m.machine_id, f"{m.machine_id} {m.name}".strip()) for m in machines]
    machine_options.sort(key=lambda x: x[0])

    return render_template(
        "equipment/downtime_batch.html",
        title="批量停机计划",
        machine_options=machine_options,
        categories=categories,
        reason_options=list(MachineDowntimeService.REASON_OPTIONS),
    )


@bp.post("/downtimes/batch/create")
def downtime_batch_create():
    scope_type = request.form.get("scope_type")
    scope_value = request.form.get("scope_value") or None
    start_time = request.form.get("start_time")
    end_time = request.form.get("end_time")
    reason_code = request.form.get("reason_code")
    reason_detail = request.form.get("reason_detail")

    svc = MachineDowntimeService(g.db, op_logger=getattr(g, "op_logger", None))
    try:
        res = svc.create_by_scope(
            scope_type=scope_type,
            scope_value=scope_value,
            start_time=start_time,
            end_time=end_time,
            reason_code=reason_code,
            reason_detail=reason_detail,
        )
        flash(f"已创建停机计划：影响设备 {res.get('created_count')} 台。", "success")
        skipped = res.get("skipped_overlap") or []
        if skipped:
            sample = "，".join(list(skipped)[:10])
            flash(f"以下设备因时间段重叠已跳过（最多 10 台）：{sample}", "warning")
    except AppError as e:
        flash(e.message, "error")
    except Exception as e:
        flash(f"批量停机创建失败：{e}", "error")
    return redirect(url_for("equipment.downtime_batch_page"))


@bp.post("/<machine_id>/downtimes/create")
def create_downtime(machine_id: str):
    start_time = request.form.get("start_time")
    end_time = request.form.get("end_time")
    reason_code = request.form.get("reason_code")
    reason_detail = request.form.get("reason_detail")

    svc = MachineDowntimeService(g.db, op_logger=getattr(g, "op_logger", None))
    d = svc.create(
        machine_id=machine_id,
        start_time=start_time,
        end_time=end_time,
        reason_code=reason_code,
        reason_detail=reason_detail,
    )
    flash(f"已新增停机计划（ID={d.id}）。", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


@bp.post("/<machine_id>/downtimes/<int:downtime_id>/cancel")
def cancel_downtime(machine_id: str, downtime_id: int):
    svc = MachineDowntimeService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.cancel(downtime_id=downtime_id, machine_id=machine_id)
    flash("已取消停机计划。", "success")
    return redirect(url_for("equipment.detail_page", machine_id=machine_id))


# ============================================================
# Excel：设备信息（Machines）
# （Phase4-03 会在此基础上完善；先占位，避免后续拆文件）
# ============================================================


def _validate_machine_excel_row(row: Dict[str, Any]) -> Optional[str]:
    if not row.get("设备编号") or str(row.get("设备编号")).strip() == "":
        return "“设备编号”不能为空"
    if not row.get("设备名称") or str(row.get("设备名称")).strip() == "":
        return "“设备名称”不能为空"

    status = row.get("状态")
    if status is None or str(status).strip() == "":
        return "“状态”不能为空（允许：active / inactive / maintain）"
    status = str(status).strip()
    if status not in ("active", "inactive", "maintain", "可用", "停用", "维修", "维护", "维护中", "维修中"):
        return "“状态”不合法（允许：active / inactive / maintain；或中文：可用/停用/维修）"

    # 工种可为空；不为空时在 confirm/preview 中做存在性校验并给出中文提示
    return None


def _normalize_machine_status_for_excel(value: Any) -> str:
    """
    Excel 中的状态允许：
    - 英文：active / inactive / maintain
    - 中文：可用 / 停用 / 维修/维护
    返回统一的英文枚举值。
    """
    if value is None:
        return ""
    v = str(value).strip()
    if v in ("可用", "启用", "正常"):
        return "active"
    if v in ("维修", "维护", "维护中", "维修中", "保养"):
        return "maintain"
    if v in ("停用", "禁用", "不可用"):
        return "inactive"
    return v


def _resolve_op_type(value: Any, op_type_repo: OpTypeRepository) -> Dict[str, Optional[str]]:
    """
    解析 Excel 的“工种”字段：
    - 支持填写 op_type_id 或 工种名称
    - 返回 dict：{op_type_id, op_type_name}
    """
    v = None if value is None else str(value).strip()
    if not v:
        return {"op_type_id": None, "op_type_name": None}
    ot = op_type_repo.get(v)
    if not ot:
        ot = op_type_repo.get_by_name(v)
    if not ot:
        raise ValidationError(f"工种“{v}”不存在，请先在工艺管理-工种配置中维护。", field="工种")
    return {"op_type_id": ot.op_type_id, "op_type_name": ot.name}


@bp.get("/excel/machines")
def excel_machine_page():
    svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = svc.build_existing_for_excel()
    return render_template(
        "equipment/excel_import_machine.html",
        title="设备信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("equipment.excel_machine_preview"),
        confirm_url=url_for("equipment.excel_machine_confirm"),
        template_download_url=url_for("equipment.excel_machine_template"),
        export_url=url_for("equipment.excel_machine_export"),
    )


@bp.post("/excel/machines/preview")
def excel_machine_preview():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="设备编号")

    op_type_repo = OpTypeRepository(g.db)

    # 预先做字段标准化，便于差异对比与落库
    normalized_rows: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        if "状态" in item:
            item["状态"] = _normalize_machine_status_for_excel(item.get("状态"))
        # 工种：预览时也尽量标准化为“工种名称”
        try:
            ot = _resolve_op_type(item.get("工种"), op_type_repo=op_type_repo)
            if ot.get("op_type_name") is not None:
                item["工种"] = ot.get("op_type_name")
        except ValidationError:
            # 让 validator 输出更友好的错误，不在这里中断
            pass
        normalized_rows.append(item)

    m_svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = m_svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        err = _validate_machine_excel_row(row)
        if err:
            return err

        # 状态标准化
        row["状态"] = _normalize_machine_status_for_excel(row.get("状态"))
        if row["状态"] not in ("active", "inactive", "maintain"):
            return "“状态”不合法（允许：active / inactive / maintain；或中文：可用/停用/维修）"

        # 工种存在性校验 + 标准化（允许为空）
        v = row.get("工种")
        if v is None or str(v).strip() == "":
            return None
        try:
            ot = _resolve_op_type(v, op_type_repo=op_type_repo)
            row["工种"] = ot.get("op_type_name")
        except ValidationError as e:
            return e.message
        return None

    svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = svc.preview_import(
        rows=normalized_rows,
        id_column="设备编号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return render_template(
        "equipment/excel_import_machine.html",
        title="设备信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(normalized_rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("equipment.excel_machine_preview"),
        confirm_url=url_for("equipment.excel_machine_confirm"),
        template_download_url=url_for("equipment.excel_machine_template"),
        export_url=url_for("equipment.excel_machine_export"),
    )


@bp.post("/excel/machines/confirm")
def excel_machine_confirm():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    _ensure_unique_ids(rows, id_column="设备编号")

    op_type_repo = OpTypeRepository(g.db)
    m_svc = MachineService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = m_svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        err = _validate_machine_excel_row(row)
        if err:
            return err
        row["状态"] = _normalize_machine_status_for_excel(row.get("状态"))
        if row["状态"] not in ("active", "inactive", "maintain"):
            return "“状态”不合法（允许：active / inactive / maintain；或中文：可用/停用/维修）"
        v = row.get("工种")
        if v is None or str(v).strip() == "":
            return None
        try:
            ot = _resolve_op_type(v, op_type_repo=op_type_repo)
            row["工种"] = ot.get("op_type_name")
        except ValidationError as e:
            return e.message
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="设备编号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )
        return render_template(
            "equipment/excel_import_machine.html",
            title="设备信息 - Excel 导入/导出",
            existing_list=list(existing.values()),
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("equipment.excel_machine_preview"),
            confirm_url=url_for("equipment.excel_machine_confirm"),
            template_download_url=url_for("equipment.excel_machine_template"),
            export_url=url_for("equipment.excel_machine_export"),
        )

    # 落库：忽略 ERROR 行
    tx = TransactionManager(g.db)
    m_repo = MachineRepository(g.db)

    new_count = update_count = skip_count = error_count = 0
    errors_sample: List[Dict[str, Any]] = []

    with tx.transaction():
        if mode == ImportMode.REPLACE:
            m_svc.ensure_replace_allowed()
            g.db.execute("DELETE FROM Machines")

        for pr in preview_rows:
            if pr.status == RowStatus.ERROR:
                error_count += 1
                if pr.message and len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": pr.message})
                continue

            machine_id = str(pr.data.get("设备编号")).strip()
            name = pr.data.get("设备名称")
            status = _normalize_machine_status_for_excel(pr.data.get("状态"))
            if status not in ("active", "inactive", "maintain"):
                # 防御：理论上不会发生
                error_count += 1
                if len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": "状态不合法，无法写入。"})
                continue

            # 工种：预览阶段已标准化为名称（或空）；这里统一再解析一次拿到 op_type_id
            op_type_id = None
            try:
                ot = _resolve_op_type(pr.data.get("工种"), op_type_repo=op_type_repo)
                op_type_id = ot.get("op_type_id")
            except ValidationError:
                op_type_id = None

            if mode == ImportMode.APPEND and machine_id in existing:
                skip_count += 1
                continue

            if m_repo.exists(machine_id):
                m_repo.update(
                    machine_id,
                    {"name": name, "op_type_id": op_type_id, "status": status},
                )
                update_count += 1
            else:
                m_repo.create({"machine_id": machine_id, "name": name, "op_type_id": op_type_id, "status": status})
                new_count += 1

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        filename=filename,
        mode=mode,
        preview_or_result={
            "total_rows": len(preview_rows),
            "new_count": new_count,
            "update_count": update_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors_sample": errors_sample,
        },
        time_cost_ms=time_cost_ms,
    )

    flash(
        f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。",
        "success",
    )
    return redirect(url_for("equipment.excel_machine_page"))


@bp.get("/excel/machines/template")
def excel_machine_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "设备信息.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="equipment",
            target_type="machine",
            template_or_export_type="设备信息模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="设备信息.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["设备编号", "设备名称", "工种", "状态"])
    ws.append(["CNC-01", "数控车床1", "数车", "active"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        template_or_export_type="设备信息模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="设备信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/machines/export")
def excel_machine_export():
    start = time.time()
    rows = g.db.execute(
        """
        SELECT m.machine_id, m.name, m.status, ot.name AS op_type_name
        FROM Machines m
        LEFT JOIN OpTypes ot ON ot.op_type_id = m.op_type_id
        ORDER BY m.machine_id
        """
    ).fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["设备编号", "设备名称", "工种", "状态"])
    for r in rows:
        ws.append([r["machine_id"], r["name"], r["op_type_name"], r["status"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="machine",
        template_or_export_type="设备信息导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="设备信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# Excel：设备人员关联（OperatorMachine）
# ============================================================


@bp.get("/excel/links")
def excel_link_page():
    rows = g.db.execute(
        """
        SELECT om.machine_id, m.name AS machine_name, om.operator_id, o.name AS operator_name,
               om.skill_level, om.is_primary
        FROM OperatorMachine om
        LEFT JOIN Machines m ON m.machine_id = om.machine_id
        LEFT JOIN Operators o ON o.operator_id = om.operator_id
        ORDER BY om.machine_id, om.operator_id
        """
    ).fetchall()
    existing_list = [
        {
            "设备编号": r["machine_id"],
            "设备名称": r["machine_name"],
            "工号": r["operator_id"],
            "姓名": r["operator_name"],
            "技能等级": r["skill_level"],
            "主操设备": r["is_primary"],
        }
        for r in rows
    ]

    return render_template(
        "equipment/excel_import_machine_operator.html",
        title="设备人员关联 - Excel 导入/导出",
        existing_list=existing_list,
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("equipment.excel_link_preview"),
        confirm_url=url_for("equipment.excel_link_confirm"),
        template_download_url=url_for("equipment.excel_link_template"),
        export_url=url_for("equipment.excel_link_export"),
    )


@bp.post("/excel/links/preview")
def excel_link_preview():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)

    # 设备侧导入字段：设备编号、工号（底层服务使用：工号、设备编号）
    # 这里做一次兼容：若用户写成“工号/设备编号”，也能识别
    normalized_rows: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        if "工号" not in item and "操作工号" in item:
            item["工号"] = item.get("操作工号")
        if "设备编号" not in item and "机器编号" in item:
            item["设备编号"] = item.get("机器编号")
        # 若表头为“设备编号/工号”，这里已经满足
        if "工号" not in item and "工号" in r:
            item["工号"] = r.get("工号")
        normalized_rows.append(item)

    link_svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    preview_rows = link_svc.preview_import_links(rows=normalized_rows, mode=mode)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="operator_machine",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    # 刷新 existing list
    existing_rows = g.db.execute(
        """
        SELECT om.machine_id, m.name AS machine_name, om.operator_id, o.name AS operator_name,
               om.skill_level, om.is_primary
        FROM OperatorMachine om
        LEFT JOIN Machines m ON m.machine_id = om.machine_id
        LEFT JOIN Operators o ON o.operator_id = om.operator_id
        ORDER BY om.machine_id, om.operator_id
        """
    ).fetchall()
    existing_list = [
        {
            "设备编号": r["machine_id"],
            "设备名称": r["machine_name"],
            "工号": r["operator_id"],
            "姓名": r["operator_name"],
            "技能等级": r["skill_level"],
            "主操设备": r["is_primary"],
        }
        for r in existing_rows
    ]

    return render_template(
        "equipment/excel_import_machine_operator.html",
        title="设备人员关联 - Excel 导入/导出",
        existing_list=existing_list,
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(normalized_rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("equipment.excel_link_preview"),
        confirm_url=url_for("equipment.excel_link_confirm"),
        template_download_url=url_for("equipment.excel_link_template"),
        export_url=url_for("equipment.excel_link_export"),
    )


@bp.post("/excel/links/confirm")
def excel_link_confirm():
    start = time.time()
    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    link_svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    preview_rows = link_svc.preview_import_links(rows=rows, mode=mode)

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )

        # 刷新 existing list（与 preview 保持一致）
        existing_rows = g.db.execute(
            """
            SELECT om.machine_id, m.name AS machine_name, om.operator_id, o.name AS operator_name,
                   om.skill_level, om.is_primary
            FROM OperatorMachine om
            LEFT JOIN Machines m ON m.machine_id = om.machine_id
            LEFT JOIN Operators o ON o.operator_id = om.operator_id
            ORDER BY om.machine_id, om.operator_id
            """
        ).fetchall()
        existing_list = [
            {
                "设备编号": r["machine_id"],
                "设备名称": r["machine_name"],
                "工号": r["operator_id"],
                "姓名": r["operator_name"],
                "技能等级": r["skill_level"],
                "主操设备": r["is_primary"],
            }
            for r in existing_rows
        ]

        return render_template(
            "equipment/excel_import_machine_operator.html",
            title="设备人员关联 - Excel 导入/导出",
            existing_list=existing_list,
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("equipment.excel_link_preview"),
            confirm_url=url_for("equipment.excel_link_confirm"),
            template_download_url=url_for("equipment.excel_link_template"),
            export_url=url_for("equipment.excel_link_export"),
        )

    stats = link_svc.apply_import_links(preview_rows=preview_rows, mode=mode)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="operator_machine",
        filename=filename,
        mode=mode,
        preview_or_result=stats,
        time_cost_ms=time_cost_ms,
    )

    flash(
        f"导入完成：新增 {stats.get('new_count', 0)}，更新 {stats.get('update_count', 0)}，跳过 {stats.get('skip_count', 0)}，错误 {stats.get('error_count', 0)}。",
        "success",
    )
    return redirect(url_for("equipment.excel_link_page"))


@bp.get("/excel/links/template")
def excel_link_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "设备人员关联.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="equipment",
            target_type="operator_machine",
            template_or_export_type="设备人员关联模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="设备人员关联.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["设备编号", "工号", "技能等级", "主操设备"])
    ws.append(["CNC-01", "OP001", "normal", "yes"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="operator_machine",
        template_or_export_type="设备人员关联模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="设备人员关联.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/links/export")
def excel_link_export():
    start = time.time()
    rows = g.db.execute(
        "SELECT machine_id, operator_id, skill_level, is_primary FROM OperatorMachine ORDER BY machine_id, operator_id"
    ).fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["设备编号", "工号", "技能等级", "主操设备"])
    for r in rows:
        ws.append([r["machine_id"], r["operator_id"], r["skill_level"], r["is_primary"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="equipment",
        target_type="operator_machine",
        template_or_export_type="设备人员关联导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="设备人员关联.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

