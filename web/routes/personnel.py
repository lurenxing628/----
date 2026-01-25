from __future__ import annotations

import io
import json
import os
import tempfile
import time
from typing import Any, Dict, List, Optional

from flask import Blueprint, current_app, flash, g, redirect, render_template, request, send_file, url_for

from core.infrastructure.errors import AppError, ErrorCode, ValidationError
from core.infrastructure.transaction import TransactionManager
from core.services.common.excel_audit import log_excel_export, log_excel_import
from core.services.common.excel_backend_factory import get_excel_backend
from core.services.common.excel_service import ExcelService, ImportMode, RowStatus
from core.services.personnel import OperatorMachineService, OperatorService
from data.repositories import MachineRepository, OperatorRepository


bp = Blueprint("personnel", __name__)


def _operator_status_zh(status: str) -> str:
    if status == "active":
        return "在岗"
    if status == "inactive":
        return "停用/休假"
    return status or "-"


def _machine_status_zh(status: str) -> str:
    if status == "active":
        return "可用"
    if status == "maintain":
        return "维修"
    if status == "inactive":
        return "停用"
    return status or "-"


def _parse_mode(value: str) -> ImportMode:
    try:
        return ImportMode(value)
    except Exception:
        raise ValidationError("导入模式不合法", field="mode")


def _read_uploaded_xlsx(file_storage) -> List[Dict[str, Any]]:
    """
    把上传的 Excel（.xlsx）解析为 List[Dict]（key 为表头字符串）。
    - 跳过空行
    - 单元格字符串自动 strip；空串视为 None
    """
    data = file_storage.read()
    if not data:
        raise AppError(ErrorCode.EXCEL_FORMAT_ERROR, "上传文件为空，请重新选择。")

    fd, tmp_path = tempfile.mkstemp(prefix="aps_upload_", suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        backend = get_excel_backend()
        return backend.read(tmp_path)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def _validate_operator_excel_row(row: Dict[str, Any]) -> Optional[str]:
    if not row.get("工号") or str(row.get("工号")).strip() == "":
        return "“工号”不能为空"
    if not row.get("姓名") or str(row.get("姓名")).strip() == "":
        return "“姓名”不能为空"

    status = row.get("状态")
    if status is None or str(status).strip() == "":
        return "“状态”不能为空（允许：active / inactive）"
    status = str(status).strip()
    if status not in ("active", "inactive"):
        return "“状态”不合法（允许：active / inactive）"
    return None


def _ensure_unique_ids(rows: List[Dict[str, Any]], id_column: str) -> None:
    seen = set()
    dup = set()
    for r in rows:
        v = r.get(id_column)
        if v is None:
            continue
        key = str(v).strip()
        if not key:
            continue
        if key in seen:
            dup.add(key)
        seen.add(key)
    if dup:
        sample = ", ".join(list(sorted(dup))[:10])
        raise ValidationError(f"Excel 中存在重复的“{id_column}”：{sample}。请去重后再导入。", field=id_column)


@bp.get("/")
def list_page():
    op_svc = OperatorService(g.db, logger=getattr(g, "app_logger", None), op_logger=getattr(g, "op_logger", None))
    m_repo = MachineRepository(g.db)

    operators = op_svc.list()
    # 预加载所有设备（用于展示名称）
    machines = {m.machine_id: m for m in m_repo.list()}

    # 预加载关联并聚合到人员
    link_rows = g.db.execute(
        "SELECT operator_id, machine_id FROM OperatorMachine ORDER BY operator_id, machine_id"
    ).fetchall()
    links_by_operator: Dict[str, List[Dict[str, Any]]] = {}
    for r in link_rows:
        op_id = r["operator_id"]
        mc_id = r["machine_id"]
        m = machines.get(mc_id)
        links_by_operator.setdefault(op_id, []).append(
            {
                "machine_id": mc_id,
                "machine_name": (m.name if m else None),
                "machine_status": (m.status if m else None),
            }
        )

    view_rows: List[Dict[str, Any]] = []
    for op in operators:
        links = links_by_operator.get(op.operator_id, [])
        machine_text = ", ".join(
            [f"{x['machine_id']}{(' ' + x['machine_name']) if x.get('machine_name') else ''}".strip() for x in links]
        )
        view_rows.append(
            {
                "operator_id": op.operator_id,
                "name": op.name,
                "status": op.status,
                "status_zh": _operator_status_zh(op.status),
                "remark": op.remark,
                "machine_text": machine_text,
                "machine_count": len(links),
            }
        )

    return render_template(
        "personnel/list.html",
        title="人员管理",
        operators=view_rows,
        status_options=[("active", "在岗"), ("inactive", "停用/休假")],
    )


@bp.post("/create")
def create_operator():
    op_id = request.form.get("operator_id")
    name = request.form.get("name")
    status = request.form.get("status") or "active"
    remark = request.form.get("remark")

    svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    op = svc.create(operator_id=op_id, name=name, status=status, remark=remark)
    flash(f"已创建人员：{op.operator_id} {op.name}", "success")
    return redirect(url_for("personnel.detail_page", operator_id=op.operator_id))


@bp.get("/<operator_id>")
def detail_page(operator_id: str):
    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    link_svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    m_repo = MachineRepository(g.db)

    op = op_svc.get(operator_id)
    links = link_svc.list_by_operator(operator_id)

    machines = {m.machine_id: m for m in m_repo.list()}
    linked_machine_ids = {l.machine_id for l in links}

    linked_machines: List[Dict[str, Any]] = []
    for l in links:
        m = machines.get(l.machine_id)
        linked_machines.append(
            {
                "machine_id": l.machine_id,
                "machine_name": (m.name if m else None),
                "status": (m.status if m else None),
                "status_zh": _machine_status_zh(m.status) if m else "-",
                "skill_level": getattr(l, "skill_level", None),
                "is_primary": getattr(l, "is_primary", None),
            }
        )

    available_machines: List[Dict[str, Any]] = []
    for m in machines.values():
        if m.machine_id in linked_machine_ids:
            continue
        available_machines.append(
            {
                "machine_id": m.machine_id,
                "machine_name": m.name,
                "status": m.status,
                "status_zh": _machine_status_zh(m.status),
            }
        )
    available_machines.sort(key=lambda x: x["machine_id"])

    return render_template(
        "personnel/detail.html",
        title=f"人员详情 - {op.operator_id} {op.name}",
        operator=op.to_dict(),
        operator_status_zh=_operator_status_zh(op.status),
        status_options=[("active", "在岗"), ("inactive", "停用/休假")],
        linked_machines=linked_machines,
        available_machines=available_machines,
        skill_level_options=[("beginner", "初级"), ("normal", "普通"), ("expert", "熟练")],
    )


@bp.post("/<operator_id>/update")
def update_operator(operator_id: str):
    name = request.form.get("name")
    status = request.form.get("status")
    remark = request.form.get("remark")

    svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    op = svc.update(operator_id=operator_id, name=name, status=status, remark=remark)
    flash("人员信息已保存。", "success")
    return redirect(url_for("personnel.detail_page", operator_id=op.operator_id))


@bp.post("/<operator_id>/status")
def set_status(operator_id: str):
    status = request.form.get("status")
    if not status:
        raise ValidationError("缺少状态参数", field="status")
    svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    op = svc.set_status(operator_id=operator_id, status=status)
    flash(f"已更新状态：{op.operator_id} → {_operator_status_zh(op.status)}", "success")
    return redirect(url_for("personnel.detail_page", operator_id=op.operator_id))


@bp.post("/<operator_id>/delete")
def delete_operator(operator_id: str):
    svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    try:
        svc.delete(operator_id)
        flash(f"已删除人员：{operator_id}", "success")
    except AppError as e:
        flash(e.message, "error")
    return redirect(url_for("personnel.list_page"))


@bp.post("/bulk/status")
def bulk_set_status():
    """
    批量设置人员状态（active/inactive）。
    """
    status = (request.form.get("status") or "").strip()
    operator_ids = request.form.getlist("operator_ids")
    if not operator_ids:
        flash("请至少选择 1 个人员。", "error")
        return redirect(url_for("personnel.list_page"))
    if status not in ("active", "inactive"):
        raise ValidationError("状态不合法（允许：active / inactive）", field="status")

    svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    ok = 0
    failed: List[str] = []
    for oid in operator_ids:
        try:
            svc.set_status(oid, status=status)
            ok += 1
        except Exception:
            failed.append(str(oid))
            continue

    flash(f"批量状态更新完成：成功 {ok}，失败 {len(failed)}。", "success" if ok else "warning")
    if failed:
        sample = "，".join(failed[:10])
        flash(f"失败人员（最多展示 10 个）：{sample}", "warning")
    return redirect(url_for("personnel.list_page"))


@bp.post("/bulk/delete")
def bulk_delete():
    """
    批量删除人员（受引用保护；建议优先批量“停用”）。
    """
    operator_ids = request.form.getlist("operator_ids")
    if not operator_ids:
        flash("请至少选择 1 个人员。", "error")
        return redirect(url_for("personnel.list_page"))

    svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    ok = 0
    failed: List[str] = []
    for oid in operator_ids:
        try:
            svc.delete(oid)
            ok += 1
        except Exception:
            failed.append(str(oid))
            continue

    flash(f"批量删除完成：成功 {ok}，失败 {len(failed)}。", "success" if ok else "warning")
    if failed:
        sample = "，".join(failed[:10])
        flash(f"删除失败（最多展示 10 个）：{sample}。常见原因：被批次工序/排程引用，请改为“停用”。", "warning")
    return redirect(url_for("personnel.list_page"))


@bp.post("/<operator_id>/link/add")
def add_link(operator_id: str):
    machine_id = request.form.get("machine_id")
    svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.add_link(operator_id=operator_id, machine_id=machine_id)
    flash("已添加设备关联。", "success")
    return redirect(url_for("personnel.detail_page", operator_id=operator_id))


@bp.post("/<operator_id>/link/update")
def update_link(operator_id: str):
    machine_id = request.form.get("machine_id")
    skill_level = request.form.get("skill_level")
    # checkbox：未勾选时 form 中不存在该 key
    is_primary = "yes" if request.form.get("is_primary") else "no"
    svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.update_link_fields(operator_id=operator_id, machine_id=machine_id, skill_level=skill_level, is_primary=is_primary)
    flash("已更新关联字段（技能等级/主操设备）。", "success")
    return redirect(url_for("personnel.detail_page", operator_id=operator_id))


@bp.post("/<operator_id>/link/remove")
def remove_link(operator_id: str):
    machine_id = request.form.get("machine_id")
    svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    svc.remove_link(operator_id=operator_id, machine_id=machine_id)
    flash("已解除设备关联。", "success")
    return redirect(url_for("personnel.detail_page", operator_id=operator_id))


# ============================================================
# Excel：人员基本信息（Operators）
# ============================================================


@bp.get("/excel/operators")
def excel_operator_page():
    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_svc.build_existing_for_excel()

    return render_template(
        "personnel/excel_import_operator.html",
        title="人员基本信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("personnel.excel_operator_preview"),
        confirm_url=url_for("personnel.excel_operator_confirm"),
        template_download_url=url_for("personnel.excel_operator_template"),
        export_url=url_for("personnel.excel_operator_export"),
    )


@bp.post("/excel/operators/preview")
def excel_operator_preview():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="工号")

    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_svc.build_existing_for_excel()

    svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = svc.preview_import(
        rows=rows,
        id_column="工号",
        existing_data=existing,
        validators=[_validate_operator_excel_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return render_template(
        "personnel/excel_import_operator.html",
        title="人员基本信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("personnel.excel_operator_preview"),
        confirm_url=url_for("personnel.excel_operator_confirm"),
        template_download_url=url_for("personnel.excel_operator_template"),
        export_url=url_for("personnel.excel_operator_export"),
    )


@bp.post("/excel/operators/confirm")
def excel_operator_confirm():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    _ensure_unique_ids(rows, id_column="工号")

    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_svc.build_existing_for_excel()

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="工号",
        existing_data=existing,
        validators=[_validate_operator_excel_row],
        mode=mode,
    )

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )
        return render_template(
            "personnel/excel_import_operator.html",
            title="人员基本信息 - Excel 导入/导出",
            existing_list=list(existing.values()),
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("personnel.excel_operator_preview"),
            confirm_url=url_for("personnel.excel_operator_confirm"),
            template_download_url=url_for("personnel.excel_operator_template"),
            export_url=url_for("personnel.excel_operator_export"),
        )

    # 落库：忽略 ERROR 行
    tx = TransactionManager(g.db)
    op_repo = OperatorRepository(g.db)

    new_count = update_count = skip_count = error_count = 0
    errors_sample: List[Dict[str, Any]] = []

    with tx.transaction():
        if mode == ImportMode.REPLACE:
            op_svc.ensure_replace_allowed()
            g.db.execute("DELETE FROM Operators")

        for pr in preview_rows:
            if pr.status == RowStatus.ERROR:
                error_count += 1
                if pr.message and len(errors_sample) < 10:
                    errors_sample.append({"row": pr.row_num, "message": pr.message})
                continue

            op_id = str(pr.data.get("工号")).strip()
            name = pr.data.get("姓名")
            status = str(pr.data.get("状态")).strip()
            remark = pr.data.get("备注")

            if mode == ImportMode.APPEND and op_id in existing:
                skip_count += 1
                continue

            exists = op_repo.exists(op_id)
            if exists:
                op_repo.update(op_id, {"name": name, "status": status, "remark": remark})
                update_count += 1
            else:
                op_repo.create({"operator_id": op_id, "name": name, "status": status, "remark": remark})
                new_count += 1

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        filename=filename,
        mode=mode,
        preview_or_result={
            "total_rows": len(preview_rows),
            "new_count": new_count,
            "update_count": update_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors_sample": errors_sample,
        },
        time_cost_ms=time_cost_ms,
    )

    flash(
        f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。",
        "success",
    )
    return redirect(url_for("personnel.excel_operator_page"))


@bp.get("/excel/operators/template")
def excel_operator_template():
    start = time.time()

    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "人员基本信息.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="personnel",
            target_type="operator",
            template_or_export_type="人员基本信息模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="人员基本信息.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工号", "姓名", "状态", "备注"])
    ws.append(["OP001", "张三", "active", "示例备注"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        template_or_export_type="人员基本信息模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员基本信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/operators/export")
def excel_operator_export():
    start = time.time()

    # 导出全部人员
    rows = g.db.execute("SELECT operator_id, name, status, remark FROM Operators ORDER BY operator_id").fetchall()
    export_rows = [
        {"工号": r["operator_id"], "姓名": r["name"], "状态": r["status"], "备注": r["remark"]} for r in rows
    ]

    # 导出到浏览器：这里直接用 openpyxl 写入内存（无需落地文件）
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工号", "姓名", "状态", "备注"])
    for r in export_rows:
        ws.append([r.get("工号"), r.get("姓名"), r.get("状态"), r.get("备注")])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        template_or_export_type="人员基本信息导出.xlsx",
        filters={},
        row_count=len(export_rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员基本信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# Excel：人员设备关联（OperatorMachine）
# ============================================================


@bp.get("/excel/links")
def excel_link_page():
    # 当前关联展示（用于用户核对）
    rows = g.db.execute(
        """
        SELECT om.operator_id, o.name AS operator_name, om.machine_id, m.name AS machine_name,
               om.skill_level, om.is_primary
        FROM OperatorMachine om
        LEFT JOIN Operators o ON o.operator_id = om.operator_id
        LEFT JOIN Machines m ON m.machine_id = om.machine_id
        ORDER BY om.operator_id, om.machine_id
        """
    ).fetchall()
    existing_list = [
        {
            "工号": r["operator_id"],
            "姓名": r["operator_name"],
            "设备编号": r["machine_id"],
            "设备名称": r["machine_name"],
            "技能等级": r["skill_level"],
            "主操设备": r["is_primary"],
        }
        for r in rows
    ]

    return render_template(
        "personnel/excel_import_operator_machine.html",
        title="人员设备关联 - Excel 导入/导出",
        existing_list=existing_list,
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("personnel.excel_link_preview"),
        confirm_url=url_for("personnel.excel_link_confirm"),
        template_download_url=url_for("personnel.excel_link_template"),
        export_url=url_for("personnel.excel_link_export"),
    )


@bp.post("/excel/links/preview")
def excel_link_preview():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    link_svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    preview_rows = link_svc.preview_import_links(rows=rows, mode=mode)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator_machine",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    # 刷新 existing list
    existing_rows = g.db.execute(
        """
        SELECT om.operator_id, o.name AS operator_name, om.machine_id, m.name AS machine_name,
               om.skill_level, om.is_primary
        FROM OperatorMachine om
        LEFT JOIN Operators o ON o.operator_id = om.operator_id
        LEFT JOIN Machines m ON m.machine_id = om.machine_id
        ORDER BY om.operator_id, om.machine_id
        """
    ).fetchall()
    existing_list = [
        {
            "工号": r["operator_id"],
            "姓名": r["operator_name"],
            "设备编号": r["machine_id"],
            "设备名称": r["machine_name"],
            "技能等级": r["skill_level"],
            "主操设备": r["is_primary"],
        }
        for r in existing_rows
    ]

    return render_template(
        "personnel/excel_import_operator_machine.html",
        title="人员设备关联 - Excel 导入/导出",
        existing_list=existing_list,
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("personnel.excel_link_preview"),
        confirm_url=url_for("personnel.excel_link_confirm"),
        template_download_url=url_for("personnel.excel_link_template"),
        export_url=url_for("personnel.excel_link_export"),
    )


@bp.post("/excel/links/confirm")
def excel_link_confirm():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception:
        raise ValidationError("预览数据解析失败，请重新上传并预览。")

    link_svc = OperatorMachineService(g.db, op_logger=getattr(g, "op_logger", None))
    preview_rows = link_svc.preview_import_links(rows=rows, mode=mode)

    # 严格模式：只要存在错误行，就拒绝导入（规范用户行为）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )

        # 刷新 existing list（与 preview 保持一致）
        existing_rows = g.db.execute(
            """
            SELECT om.operator_id, o.name AS operator_name, om.machine_id, m.name AS machine_name,
                   om.skill_level, om.is_primary
            FROM OperatorMachine om
            LEFT JOIN Operators o ON o.operator_id = om.operator_id
            LEFT JOIN Machines m ON m.machine_id = om.machine_id
            ORDER BY om.operator_id, om.machine_id
            """
        ).fetchall()
        existing_list = [
            {
                "工号": r["operator_id"],
                "姓名": r["operator_name"],
                "设备编号": r["machine_id"],
                "设备名称": r["machine_name"],
                "技能等级": r["skill_level"],
                "主操设备": r["is_primary"],
            }
            for r in existing_rows
        ]

        return render_template(
            "personnel/excel_import_operator_machine.html",
            title="人员设备关联 - Excel 导入/导出",
            existing_list=existing_list,
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("personnel.excel_link_preview"),
            confirm_url=url_for("personnel.excel_link_confirm"),
            template_download_url=url_for("personnel.excel_link_template"),
            export_url=url_for("personnel.excel_link_export"),
        )

    stats = link_svc.apply_import_links(preview_rows=preview_rows, mode=mode)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator_machine",
        filename=filename,
        mode=mode,
        preview_or_result=stats,
        time_cost_ms=time_cost_ms,
    )

    flash(
        f"导入完成：新增 {stats.get('new_count', 0)}，更新 {stats.get('update_count', 0)}，跳过 {stats.get('skip_count', 0)}，错误 {stats.get('error_count', 0)}。",
        "success",
    )
    return redirect(url_for("personnel.excel_link_page"))


@bp.get("/excel/links/template")
def excel_link_template():
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "人员设备关联.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=getattr(g, "op_logger", None),
            module="personnel",
            target_type="operator_machine",
            template_or_export_type="人员设备关联模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_file(
            template_path,
            as_attachment=True,
            download_name="人员设备关联.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工号", "设备编号", "技能等级", "主操设备"])
    ws.append(["OP001", "CNC-01", "normal", "yes"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator_machine",
        template_or_export_type="人员设备关联模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员设备关联.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/links/export")
def excel_link_export():
    start = time.time()
    rows = g.db.execute(
        "SELECT operator_id, machine_id, skill_level, is_primary FROM OperatorMachine ORDER BY operator_id, machine_id"
    ).fetchall()

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工号", "设备编号", "技能等级", "主操设备"])
    for r in rows:
        ws.append([r["operator_id"], r["machine_id"], r["skill_level"], r["is_primary"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator_machine",
        template_or_export_type="人员设备关联导出.xlsx",
        filters={},
        row_count=len(rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员设备关联.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

