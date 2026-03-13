from __future__ import annotations

import io
import json
import os
import time
from typing import Any, Dict, List, Optional

from flask import current_app, flash, g, redirect, request, send_file, url_for

from core.infrastructure.errors import ValidationError
from core.models.enums import OperatorStatus
from core.services.common.enum_normalizers import normalize_operator_status
from core.services.common.excel_audit import log_excel_export, log_excel_import
from core.services.common.excel_backend_factory import get_excel_backend
from core.services.common.excel_service import ExcelService, ImportMode, RowStatus
from core.services.personnel import OperatorService, ResourceTeamService
from core.services.personnel.operator_excel_import_service import OperatorExcelImportService
from web.ui_mode import render_ui_template as render_template

from .personnel_bp import _ensure_unique_ids, _parse_mode, _read_uploaded_xlsx, bp


def _validate_operator_excel_row(row: Dict[str, Any]) -> Optional[str]:
    if not row.get("工号") or str(row.get("工号")).strip() == "":
        return "工号不能为空"
    if not row.get("姓名") or str(row.get("姓名")).strip() == "":
        return "姓名不能为空"

    st = normalize_operator_status(row.get("状态"))
    if not st:
        return "状态不能为空，请填写：在岗 或 停用（也兼容 active / inactive）。"
    if st not in (OperatorStatus.ACTIVE.value, OperatorStatus.INACTIVE.value):
        return "状态不合法，可填写：在岗 / 停用（也兼容 active / inactive）。"
    row["状态"] = st
    return None


def _build_operator_template_output(template_path: str) -> io.BytesIO:
    import openpyxl

    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    desired_headers = ["工号", "姓名", "状态", "班组", "备注"]
    current_headers = [str(ws.cell(1, idx + 1).value or "").strip() for idx in range(len(desired_headers))]
    if current_headers != desired_headers:
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(desired_headers)
        ws.append(["OP001", "张三", OperatorStatus.ACTIVE.value, "车工一组", "示例备注"])
    elif ws.max_row < 2:
        ws.append(["OP001", "张三", OperatorStatus.ACTIVE.value, "车工一组", "示例备注"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# Excel：人员基本信息（Operators）
# ============================================================


@bp.get("/excel/operators")
def excel_operator_page():
    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_svc.build_existing_for_excel()

    return render_template(
        "personnel/excel_import_operator.html",
        title="人员基本信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=None,
        raw_rows_json=None,
        mode=ImportMode.OVERWRITE.value,
        filename=None,
        preview_url=url_for("personnel.excel_operator_preview"),
        confirm_url=url_for("personnel.excel_operator_confirm"),
        template_download_url=url_for("personnel.excel_operator_template"),
        export_url=url_for("personnel.excel_operator_export"),
    )


@bp.post("/excel/operators/preview")
def excel_operator_preview():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    rows = _read_uploaded_xlsx(file)
    _ensure_unique_ids(rows, id_column="工号")

    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    team_svc = ResourceTeamService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_svc.build_existing_for_excel()

    normalized_rows: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        if "状态" in item:
            item["状态"] = normalize_operator_status(item.get("状态"))
        if "班组" in item:
            try:
                item["班组"] = team_svc.resolve_team_name_optional(item.get("班组"))
            except ValidationError:
                pass
        normalized_rows.append(item)

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        err = _validate_operator_excel_row(row)
        if err:
            return err
        if "班组" not in row:
            return None
        try:
            row["班组"] = team_svc.resolve_team_name_optional(row.get("班组"))
        except ValidationError as e:
            return e.message
        return None

    svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = svc.preview_import(
        rows=normalized_rows,
        id_column="工号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return render_template(
        "personnel/excel_import_operator.html",
        title="人员基本信息 - Excel 导入/导出",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(normalized_rows, ensure_ascii=False),
        mode=mode.value,
        filename=file.filename,
        preview_url=url_for("personnel.excel_operator_preview"),
        confirm_url=url_for("personnel.excel_operator_confirm"),
        template_download_url=url_for("personnel.excel_operator_template"),
        export_url=url_for("personnel.excel_operator_export"),
    )


@bp.post("/excel/operators/confirm")
def excel_operator_confirm():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception as e:
        raise ValidationError("预览数据解析失败，请重新上传并预览。") from e

    _ensure_unique_ids(rows, id_column="工号")

    op_svc = OperatorService(g.db, op_logger=getattr(g, "op_logger", None))
    team_svc = ResourceTeamService(g.db, op_logger=getattr(g, "op_logger", None))
    existing = op_svc.build_existing_for_excel()

    def validate_row(row: Dict[str, Any]) -> Optional[str]:
        err = _validate_operator_excel_row(row)
        if err:
            return err
        if "班组" not in row:
            return None
        try:
            row["班组"] = team_svc.resolve_team_name_optional(row.get("班组"))
        except ValidationError as e:
            return e.message
        return None

    excel_svc = ExcelService(backend=get_excel_backend(), logger=None, op_logger=getattr(g, "op_logger", None))
    preview_rows = excel_svc.preview_import(
        rows=rows,
        id_column="工号",
        existing_data=existing,
        validators=[validate_row],
        mode=mode,
    )

    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{pr.row_num}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        message = f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。"
        if sample:
            message += f"错误示例：{sample}"
        flash(
            message,
            "error",
        )
        return render_template(
            "personnel/excel_import_operator.html",
            title="人员基本信息 - Excel 导入/导出",
            existing_list=list(existing.values()),
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            mode=mode.value,
            filename=filename,
            preview_url=url_for("personnel.excel_operator_preview"),
            confirm_url=url_for("personnel.excel_operator_confirm"),
            template_download_url=url_for("personnel.excel_operator_template"),
            export_url=url_for("personnel.excel_operator_export"),
        )

    import_svc = OperatorExcelImportService(
        g.db,
        logger=getattr(g, "app_logger", None),
        op_logger=getattr(g, "op_logger", None),
    )
    import_stats = import_svc.apply_preview_rows(preview_rows, mode=mode, existing_ids=set(existing.keys()))
    new_count = int(import_stats.get("new_count", 0))
    update_count = int(import_stats.get("update_count", 0))
    skip_count = int(import_stats.get("skip_count", 0))
    error_count = int(import_stats.get("error_count", 0))

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        filename=filename,
        mode=mode,
        preview_or_result=import_stats,
        time_cost_ms=time_cost_ms,
    )

    flash(
        f"导入完成：新增 {new_count}，更新 {update_count}，跳过 {skip_count}，错误 {error_count}。",
        "success",
    )
    return redirect(url_for("personnel.excel_operator_page"))


@bp.get("/excel/operators/template")
def excel_operator_template():
    start = time.time()

    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "人员基本信息.xlsx")
    output = _build_operator_template_output(template_path)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        template_or_export_type="人员基本信息模板.xlsx",
        filters={},
        row_count=1,
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员基本信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/excel/operators/export")
def excel_operator_export():
    start = time.time()

    existing = OperatorService(g.db, op_logger=getattr(g, "op_logger", None)).build_existing_for_excel()
    export_rows = list(existing.values())

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["工号", "姓名", "状态", "班组", "备注"])
    for r in export_rows:
        ws.append([r.get("工号"), r.get("姓名"), r.get("状态"), r.get("班组"), r.get("备注")])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=getattr(g, "op_logger", None),
        module="personnel",
        target_type="operator",
        template_or_export_type="人员基本信息导出.xlsx",
        filters={},
        row_count=len(export_rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员基本信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
