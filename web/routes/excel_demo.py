from __future__ import annotations

import io
import json
import os
import time
from typing import Any, Dict, List, Optional

from flask import Blueprint, current_app, flash, g, redirect, request, send_file, url_for

from core.infrastructure.errors import AppError, ErrorCode, ValidationError
from core.models.enums import OperatorStatus
from core.services.common.enum_normalizers import normalize_operator_status
from core.services.common.excel_audit import log_excel_export, log_excel_import
from core.services.common.excel_service import ExcelService, ImportMode, RowStatus
from core.services.common.excel_templates import build_xlsx_bytes, get_template_definition
from core.services.common.normalize import is_blank_value
from core.services.common.openpyxl_backend import OpenpyxlBackend
from core.services.common.tabular_backend import SOURCE_ROW_NUM_KEY, SOURCE_SHEET_NAME_KEY
from core.services.personnel import OperatorService
from core.services.personnel.operator_excel_import_service import OperatorExcelImportService
from web.ui_mode import render_ui_template as render_template

from .excel_utils import (
    build_preview_baseline_token,
    flash_import_result,
    parse_import_mode,
    preview_baseline_matches,
    send_excel_template_file,
)

bp = Blueprint("excel_demo", __name__)


def _fetch_existing_operators(conn) -> Dict[str, Dict[str, Any]]:
    """以 Excel 列名（中文）构建 existing_data，便于 preview_import 做 diff。"""
    return OperatorService(conn, op_logger=None).build_existing_for_excel()


def _parse_mode(value: str) -> ImportMode:
    return parse_import_mode(value)


def _render_demo_page(
    *,
    existing: Dict[str, Dict[str, Any]],
    preview_rows: Any,
    raw_rows_json: Any,
    preview_baseline: str,
    mode_value: str,
    filename: Any,
):
    return render_template(
        "excel/demo.html",
        title="Excel 导入演示",
        existing_list=list(existing.values()),
        preview_rows=preview_rows,
        raw_rows_json=raw_rows_json,
        preview_baseline=preview_baseline,
        mode=mode_value,
        filename=filename,
        preview_url=url_for("excel_demo.preview"),
        confirm_url=url_for("excel_demo.confirm"),
        template_download_url=url_for("excel_demo.download_template"),
    )


def _validate_operator_row(row: Dict[str, Any]) -> Optional[str]:
    # 返回中文错误提示；返回 None 表示通过
    if is_blank_value(row.get("工号")):
        return "“工号”不能为空"
    if is_blank_value(row.get("姓名")):
        return "“姓名”不能为空"

    st = normalize_operator_status(row.get("状态"))
    if not st:
        return "“状态”不能为空，请填写：在岗 或 停用（也兼容 active / inactive）。"
    if st not in (OperatorStatus.ACTIVE.value, OperatorStatus.INACTIVE.value):
        return "“状态”不合法，可填写：在岗 / 停用（也兼容 active / inactive）。"
    row["状态"] = st
    return None


@bp.get("/")
def index():
    existing = _fetch_existing_operators(g.db)
    return _render_demo_page(
        existing=existing,
        preview_rows=None,
        raw_rows_json=None,
        preview_baseline="",
        mode_value=ImportMode.OVERWRITE.value,
        filename=None,
    )


@bp.post("/preview")
def preview():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValidationError("请先选择要上传的 Excel 文件", field="file")

    # 为了演示简单：直接读取上传流到内存（openpyxl 需要文件样式，因此保存到 BytesIO）
    data = file.read()
    if not data:
        raise AppError(ErrorCode.EXCEL_FORMAT_ERROR, "上传文件为空，请重新选择。")

    tmp = io.BytesIO(data)
    tmp.seek(0)

    backend = OpenpyxlBackend()
    # openpyxl.load_workbook 需要类文件对象时要用 filename=...；这里用临时文件流不稳定
    # 因此：写入一个临时内存文件并交给 openpyxl（openpyxl 支持 BytesIO）
    wb = None
    try:
        import openpyxl

        wb = openpyxl.load_workbook(tmp, data_only=True)
        ws = wb.active
        if ws is None:
            raise AppError(ErrorCode.EXCEL_FORMAT_ERROR, "Excel 缺少活动工作表，无法读取。")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise AppError(ErrorCode.EXCEL_FORMAT_ERROR, "Excel 内容为空，未读取到任何行。")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        parsed_rows: List[Dict[str, Any]] = []
        for excel_row_num, raw in enumerate(rows[1:], start=2):
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            item: Dict[str, Any] = {
                SOURCE_ROW_NUM_KEY: int(excel_row_num),
                SOURCE_SHEET_NAME_KEY: str(getattr(ws, "title", None) or "Sheet1"),
            }
            for idx, key in enumerate(headers):
                if not key:
                    continue
                val = raw[idx] if idx < len(raw) else None
                if isinstance(val, str):
                    val = val.strip()
                    if val == "":
                        val = None
                item[key] = val
            parsed_rows.append(item)
    except AppError:
        raise
    except Exception as e:
        raise AppError(ErrorCode.EXCEL_READ_ERROR, "读取 Excel 失败，请确认文件未损坏且未被占用。", cause=e) from e
    finally:
        if wb is not None:
            wb.close()

    existing = _fetch_existing_operators(g.db)
    svc = ExcelService(backend=backend, logger=None, op_logger=g.op_logger)
    preview_rows = svc.preview_import(
        rows=parsed_rows,
        id_column="工号",
        existing_data=existing,
        validators=[_validate_operator_row],
        mode=mode,
    )
    preview_baseline = build_preview_baseline_token(existing_data=existing, mode=mode, id_column="工号")

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=g.op_logger,
        module="excel_demo",
        target_type="operator",
        filename=file.filename,
        mode=mode,
        preview_or_result=preview_rows,
        time_cost_ms=time_cost_ms,
    )

    return _render_demo_page(
        existing=existing,
        preview_rows=preview_rows,
        raw_rows_json=json.dumps(parsed_rows, ensure_ascii=False),
        preview_baseline=preview_baseline,
        mode_value=mode.value,
        filename=file.filename,
    )


@bp.post("/confirm")
def confirm():
    start = time.time()

    mode = _parse_mode(request.form.get("mode", ImportMode.OVERWRITE.value))
    filename = request.form.get("filename") or "unknown.xlsx"
    raw_rows_json = request.form.get("raw_rows_json")
    preview_baseline = (request.form.get("preview_baseline") or "").strip()
    if not raw_rows_json:
        raise ValidationError("缺少预览数据，请重新上传并预览后再确认导入。")
    if not preview_baseline:
        raise ValidationError("缺少预览基线，请重新上传并预览后再确认导入。")

    try:
        rows = json.loads(raw_rows_json)
        if not isinstance(rows, list):
            raise ValueError("rows not list")
    except Exception as e:
        raise ValidationError("预览数据解析失败，请重新上传并预览。") from e

    existing = _fetch_existing_operators(g.db)
    if not preview_baseline_matches(preview_baseline, existing_data=existing, mode=mode, id_column="工号"):
        flash("导入被拒绝：数据已变化，需重新预览后再确认导入。", "error")
        return _render_demo_page(
            existing=existing,
            preview_rows=None,
            raw_rows_json=None,
            preview_baseline="",
            mode_value=mode.value,
            filename=filename,
        )
    backend = OpenpyxlBackend()
    svc = ExcelService(backend=backend, logger=None, op_logger=g.op_logger)
    preview_rows = svc.preview_import(
        rows=rows,
        id_column="工号",
        existing_data=existing,
        validators=[_validate_operator_row],
        mode=mode,
    )

    # 严格模式：只要存在错误行，就拒绝导入（演示页也保持一致）
    error_rows = [pr for pr in preview_rows if pr.status == RowStatus.ERROR]
    if error_rows:
        sample = "；".join([f"第{(getattr(pr, 'source_row_num', None) or pr.row_num)}行：{pr.message}" for pr in error_rows[:5] if pr and pr.message])
        flash(
            f"导入被拒绝：Excel 存在 {len(error_rows)} 行错误。请修正后重新预览并确认。{('错误示例：' + sample) if sample else ''}",
            "error",
        )
        return _render_demo_page(
            existing=existing,
            preview_rows=preview_rows,
            raw_rows_json=json.dumps(rows, ensure_ascii=False),
            preview_baseline=preview_baseline,
            mode_value=mode.value,
            filename=filename,
        )

    import_svc = OperatorExcelImportService(
        g.db,
        logger=getattr(g, "app_logger", None),
        op_logger=getattr(g, "op_logger", None),
    )
    import_stats = import_svc.apply_preview_rows(preview_rows, mode=mode, existing_ids=set(existing.keys()))
    new_count = int(import_stats.get("new_count", 0))
    update_count = int(import_stats.get("update_count", 0))
    skip_count = int(import_stats.get("skip_count", 0))
    error_count = int(import_stats.get("error_count", 0))

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_import(
        op_logger=g.op_logger,
        module="excel_demo",
        target_type="operator",
        filename=filename,
        mode=mode,
        preview_or_result=import_stats,
        time_cost_ms=time_cost_ms,
    )

    flash_import_result(
        new_count=new_count,
        update_count=update_count,
        skip_count=skip_count,
        error_count=error_count,
        errors_sample=list(import_stats.get("errors_sample") or []),
    )
    return redirect(url_for("excel_demo.index"))


@bp.get("/template")
def download_template():
    """
    下载“人员基本信息.xlsx”模板（演示用）：列名与文档一致（工号/姓名/状态/班组/备注）。
    """
    start = time.time()
    template_path = os.path.join(current_app.config["EXCEL_TEMPLATE_DIR"], "人员基本信息.xlsx")
    if os.path.exists(template_path):
        time_cost_ms = int((time.time() - start) * 1000)
        log_excel_export(
            op_logger=g.op_logger,
            module="excel_demo",
            target_type="operator",
            template_or_export_type="人员基本信息模板.xlsx",
            filters={},
            row_count=1,
            time_range={},
            time_cost_ms=time_cost_ms,
        )
        return send_excel_template_file(template_path, download_name="人员基本信息.xlsx")

    template_def = get_template_definition("人员基本信息.xlsx")
    sample_rows = template_def.get("sample_rows") or []
    output = build_xlsx_bytes(
        template_def["headers"],
        sample_rows,
        format_spec=template_def.get("format_spec"),
    )

    time_cost_ms = int((time.time() - start) * 1000)
    log_excel_export(
        op_logger=g.op_logger,
        module="excel_demo",
        target_type="operator",
        template_or_export_type="人员基本信息模板.xlsx",
        filters={},
        row_count=len(sample_rows),
        time_range={},
        time_cost_ms=time_cost_ms,
    )

    return send_file(
        output,
        as_attachment=True,
        download_name="人员基本信息.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

