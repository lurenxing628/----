from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl


@dataclass
class StationMeta:
    """源表中的设备/人员列块定义（每块 4 列）。"""

    col_start: int  # 0-based：工步描述列
    machine_id: str
    machine_label: str
    operators: List[str] = field(default_factory=list)


@dataclass
class StepRecord:
    """源表中一条工步/工序记录（按行+设备列块展开）。"""

    part_no: str
    seq: Optional[int]
    step_text: str
    has_step_code: bool  # 仅 XX-X 视为“工步”
    machine_id: str
    operators: List[str]
    setup_min: Optional[float]
    unit_min: Optional[float]
    batch_min: Optional[float]


@dataclass
class PartContext:
    part_no: str
    part_name: str
    route_raw: str
    route_map: Dict[int, str] = field(default_factory=dict)  # seq -> op_type_name
    step_records: List[StepRecord] = field(default_factory=list)


@dataclass
class ConvertedTemplates:
    routes_rows: List[Dict[str, Any]]
    part_operation_hours_rows: List[Dict[str, Any]]
    operators_rows: List[Dict[str, Any]]
    machines_rows: List[Dict[str, Any]]
    operator_machine_rows: List[Dict[str, Any]]
    op_types_rows: List[Dict[str, Any]]
    suppliers_rows: List[Dict[str, Any]]

    @staticmethod
    def output_specs() -> List[Tuple[str, List[str], str]]:
        return [
            ("零件工艺路线.xlsx", ["图号", "名称", "工艺路线字符串"], "routes_rows"),
            ("零件工序工时.xlsx", ["图号", "工序", "换型时间(h)", "单件工时(h)"], "part_operation_hours_rows"),
            ("人员基本信息.xlsx", ["工号", "姓名", "状态", "备注"], "operators_rows"),
            ("设备信息.xlsx", ["设备编号", "设备名称", "工种", "状态"], "machines_rows"),
            # 本阶段按用户要求：仅保留两列，不输出技能等级/主操设备
            ("人员设备关联.xlsx", ["工号", "设备编号"], "operator_machine_rows"),
            ("工种配置.xlsx", ["工种ID", "工种名称", "归属"], "op_types_rows"),
            ("供应商配置.xlsx", ["供应商ID", "名称", "对应工种", "默认周期"], "suppliers_rows"),
        ]

    def rows_by_filename(self) -> Dict[str, List[Dict[str, Any]]]:
        result: Dict[str, List[Dict[str, Any]]] = {}
        for filename, _headers, field_name in self.output_specs():
            result[filename] = list(getattr(self, field_name))
        return result


class UnitExcelConverter:
    """
    把“回转壳体单元产品数据.xlsx”转换为系统现有模板数据。

    关键规则（按用户要求）：
    - XX-X 视为工步，需并入工序（仅保留 XX 工序号）；
    - 只有“有工步 + 有人员操作”的工序视为单元内部工序；
    - 其他工序统一外协；
    - 人员设备关联仅输出 工号/设备编号 两列。
    """

    _SEPARATORS_RE = re.compile(r"[\s,，、\-—–→>＞]+")
    _CJK_RE = re.compile(r"[\u4e00-\u9fff]")

    def convert(self, input_path: str, sheet_name: Optional[str] = None) -> ConvertedTemplates:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            headers = [self._to_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            stations = self._build_station_columns(headers)

            parts: Dict[str, PartContext] = {}
            current_part_no: Optional[str] = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                row_values = list(row)

                part_no = self._to_text(self._pick_cell(row_values, 0))
                if part_no:
                    current_part_no = part_no
                    part_name = self._to_text(self._pick_cell(row_values, 1))
                    route_raw = self._to_text(self._pick_cell(row_values, 4))
                    route_map = self._parse_route_map(route_raw)
                    parts[part_no] = PartContext(
                        part_no=part_no,
                        part_name=part_name or part_no,
                        route_raw=route_raw,
                        route_map=route_map,
                    )

                if not current_part_no or current_part_no not in parts:
                    continue

                ctx = parts[current_part_no]
                for station in stations:
                    step_text = self._to_text(self._pick_cell(row_values, station.col_start))
                    if not step_text:
                        continue

                    seq, has_step_code = self._parse_step_seq(step_text)
                    rec = StepRecord(
                        part_no=ctx.part_no,
                        seq=seq,
                        step_text=step_text,
                        has_step_code=has_step_code,
                        machine_id=station.machine_id,
                        operators=list(station.operators),
                        setup_min=self._to_float(self._pick_cell(row_values, station.col_start + 1)),
                        unit_min=self._to_float(self._pick_cell(row_values, station.col_start + 2)),
                        batch_min=self._to_float(self._pick_cell(row_values, station.col_start + 3)),
                    )
                    ctx.step_records.append(rec)

            return self._build_templates(parts=parts, stations=stations)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def write_templates(self, converted: ConvertedTemplates, output_dir: str) -> Dict[str, str]:
        os.makedirs(output_dir, exist_ok=True)
        output_paths: Dict[str, str] = {}
        for filename, headers, field_name in ConvertedTemplates.output_specs():
            rows = list(getattr(converted, field_name))
            path = os.path.join(output_dir, filename)
            self._write_xlsx(path=path, headers=headers, rows=rows)
            output_paths[filename] = path
        return output_paths

    # -------------------------
    # 核心构建
    # -------------------------
    def _build_templates(self, parts: Dict[str, PartContext], stations: List[StationMeta]) -> ConvertedTemplates:
        machine_label_map = {s.machine_id: s.machine_label for s in stations}

        # 由“route 已映射工序”学习 machine -> op_type 主分布（用于 route 缺失序号时推断）
        machine_op_hint: Dict[str, Counter] = defaultdict(Counter)
        for ctx in parts.values():
            for rec in ctx.step_records:
                if rec.seq is None:
                    continue
                op_name = ctx.route_map.get(rec.seq)
                if op_name:
                    machine_op_hint[rec.machine_id][op_name] += 1

        # 操作记录（part_no, seq, base_name, source_internal, ext_days_hint, setup/unit hours）
        op_records: List[Dict[str, Any]] = []
        # 内部资源（仅“有工步+有人员”才纳入）
        machine_internal_counter: Dict[str, Counter] = defaultdict(Counter)
        operator_names: Set[str] = set()
        used_machine_ids: Set[str] = set()
        links: Set[Tuple[str, str]] = set()  # (operator_name, machine_id)

        for part_no in sorted(parts.keys()):
            ctx = parts[part_no]
            seq_to_records: Dict[int, List[StepRecord]] = defaultdict(list)
            internal_seq_set: Set[int] = set()

            for rec in ctx.step_records:
                if rec.seq is None:
                    continue
                seq_to_records[int(rec.seq)].append(rec)
                if rec.has_step_code and rec.operators:
                    internal_seq_set.add(int(rec.seq))

            inferred_missing_name: Dict[int, str] = {}
            for seq in sorted(internal_seq_set):
                if seq in ctx.route_map:
                    continue
                inferred_missing_name[seq] = self._infer_op_name_for_missing_seq(
                    seq=seq,
                    records=seq_to_records.get(seq, []),
                    machine_op_hint=machine_op_hint,
                )

            all_seqs = sorted(set(ctx.route_map.keys()) | internal_seq_set)
            for seq in all_seqs:
                base_name = (ctx.route_map.get(seq) or inferred_missing_name.get(seq) or f"工序{seq}").strip()
                source_internal = seq in internal_seq_set

                seq_records = seq_to_records.get(seq, [])
                setup_h, unit_h = self._aggregate_internal_hours(seq_records) if source_internal else (0.0, 0.0)
                ext_days_hint = None if source_internal else self._estimate_external_days(seq_records)

                op_records.append(
                    {
                        "part_no": ctx.part_no,
                        "part_name": ctx.part_name,
                        "seq": int(seq),
                        "base_name": base_name,
                        "source_internal": bool(source_internal),
                        "ext_days_hint": ext_days_hint,
                        "setup_hours": float(setup_h),
                        "unit_hours": float(unit_h),
                    }
                )

                if source_internal:
                    for rec in seq_records:
                        if not rec.has_step_code or not rec.operators:
                            continue
                        used_machine_ids.add(rec.machine_id)
                        machine_internal_counter[rec.machine_id][base_name] += 1
                        for name in rec.operators:
                            operator_names.add(name)
                            links.add((name, rec.machine_id))

        # 若同名工种既有 internal 又有 external，外协侧自动加后缀，避免导入后被统一判成 internal
        name_states: Dict[str, Set[str]] = defaultdict(set)
        for rec in op_records:
            state = "internal" if rec["source_internal"] else "external"
            name_states[str(rec["base_name"])].add(state)
        conflict_names = {name for name, states in name_states.items() if len(states) > 1}

        for rec in op_records:
            base_name = str(rec["base_name"])
            if rec["source_internal"]:
                final_name = base_name
            else:
                final_name = self._external_alias(base_name) if base_name in conflict_names else base_name
            rec["final_name"] = final_name

        # 1) 零件工艺路线
        route_rows: List[Dict[str, Any]] = []
        by_part: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for rec in op_records:
            by_part[str(rec["part_no"])].append(rec)
        for part_no in sorted(by_part.keys()):
            recs = sorted(by_part[part_no], key=lambda x: int(x["seq"]))
            part_name = str(recs[0]["part_name"] or part_no)
            route_string = "".join([f"{int(r['seq'])}{str(r['final_name'])}" for r in recs])
            route_rows.append({"图号": part_no, "名称": part_name, "工艺路线字符串": route_string})

        # 2) 零件工序工时（仅内部工序；单位小时）
        part_operation_hours_rows: List[Dict[str, Any]] = []
        for part_no in sorted(by_part.keys()):
            recs = sorted(by_part[part_no], key=lambda x: int(x["seq"]))
            for rec in recs:
                if not rec["source_internal"]:
                    continue
                part_operation_hours_rows.append(
                    {
                        "图号": part_no,
                        "工序": int(rec["seq"]),
                        "换型时间(h)": float(rec.get("setup_hours") or 0.0),
                        "单件工时(h)": float(rec.get("unit_hours") or 0.0),
                    }
                )

        # 3) 人员基本信息（工号使用 6 位数字稳定生成）
        operator_id_map: Dict[str, str] = {}
        used_operator_ids: Set[str] = set()
        operators_rows: List[Dict[str, Any]] = []
        next_operator_no = 900001
        for name in sorted(operator_names):
            while f"{next_operator_no:06d}" in used_operator_ids:
                next_operator_no += 1
            op_id = f"{next_operator_no:06d}"
            used_operator_ids.add(op_id)
            next_operator_no += 1
            operator_id_map[name] = op_id
            operators_rows.append({"工号": op_id, "姓名": name, "状态": "active", "备注": None})

        # 4) 设备信息（工种按内部工序主分布推断，可为空）
        machines_rows: List[Dict[str, Any]] = []
        for machine_id in sorted(used_machine_ids):
            display = machine_label_map.get(machine_id) or machine_id
            machine_name = self._build_machine_name(machine_id=machine_id, machine_label=display)
            op_name = self._most_common_key(machine_internal_counter.get(machine_id))
            machines_rows.append(
                {
                    "设备编号": machine_id,
                    "设备名称": machine_name,
                    "工种": op_name,
                    "状态": "active",
                }
            )

        # 5) 人员设备关联（仅工号/设备编号）
        operator_machine_rows: List[Dict[str, Any]] = []
        for operator_name, machine_id in sorted(links, key=lambda x: (operator_id_map.get(x[0], x[0]), x[1])):
            op_id = operator_id_map.get(operator_name)
            if not op_id:
                continue
            operator_machine_rows.append({"工号": op_id, "设备编号": machine_id})

        # 6) 工种配置
        op_type_states: Dict[str, Set[str]] = defaultdict(set)
        for rec in op_records:
            nm = str(rec["final_name"])
            op_type_states[nm].add("internal" if rec["source_internal"] else "external")

        op_types_rows: List[Dict[str, Any]] = []
        for idx, op_name in enumerate(sorted(op_type_states.keys()), start=1):
            # 理论上每个 final_name 只会归属单一状态，这里做保护
            cat = "internal" if "internal" in op_type_states[op_name] else "external"
            op_types_rows.append({"工种ID": f"OT{idx:03d}", "工种名称": op_name, "归属": cat})

        # 7) 供应商配置（外协工种生成默认供应商）
        external_days_by_name: Dict[str, List[float]] = defaultdict(list)
        for rec in op_records:
            if rec["source_internal"]:
                continue
            name = str(rec["final_name"])
            hint = rec.get("ext_days_hint")
            if isinstance(hint, (int, float)) and float(hint) > 0:
                external_days_by_name[name].append(float(hint))

        external_names = [r["工种名称"] for r in op_types_rows if str(r["归属"]).strip().lower() == "external"]
        suppliers_rows: List[Dict[str, Any]] = []
        for idx, op_name in enumerate(sorted(external_names), start=1):
            days_list = external_days_by_name.get(op_name) or []
            default_days = round(sum(days_list) / len(days_list), 4) if days_list else 1.0
            if default_days <= 0:
                default_days = 1.0
            suppliers_rows.append(
                {
                    "供应商ID": f"S{idx:03d}",
                    "名称": f"外协-{op_name}",
                    "对应工种": op_name,
                    "默认周期": default_days,
                }
            )

        return ConvertedTemplates(
            routes_rows=route_rows,
            part_operation_hours_rows=part_operation_hours_rows,
            operators_rows=operators_rows,
            machines_rows=machines_rows,
            operator_machine_rows=operator_machine_rows,
            op_types_rows=op_types_rows,
            suppliers_rows=suppliers_rows,
        )

    # -------------------------
    # 解析工具
    # -------------------------
    def _build_station_columns(self, headers: Sequence[str]) -> List[StationMeta]:
        stations: List[StationMeta] = []
        # A~H 为零件基础信息，后续按 4 列一个设备块（工步/换型/单件/批次）
        idx = 8
        while idx + 3 < len(headers):
            raw_header = self._to_text(headers[idx])
            if raw_header:
                machine_id, machine_label, operators = self._parse_station_header(raw_header)
                if machine_id:
                    stations.append(
                        StationMeta(
                            col_start=idx,
                            machine_id=machine_id,
                            machine_label=machine_label or machine_id,
                            operators=operators,
                        )
                    )
            idx += 4
        return stations

    def _parse_station_header(self, raw_header: str) -> Tuple[str, str, List[str]]:
        text = self._to_text(raw_header)
        lines = [x.strip() for x in re.split(r"[\r\n]+", text) if x and x.strip()]
        if not lines:
            return "", "", []

        machine_label = lines[0]
        machine_id = self._extract_machine_id(machine_label)

        operator_names: List[str] = []
        if len(lines) > 1:
            for ln in lines[1:]:
                operator_names.extend(self._extract_names(ln))
        else:
            # 兼容 “3140124 胡凡 罗辉” 这种单行写法
            tokens = [t for t in re.split(r"[、,，/\s]+", machine_label) if t]
            if tokens:
                first = tokens[0]
                machine_label = first
                machine_id = self._extract_machine_id(first) or first
                operator_names.extend(self._extract_names(" ".join(tokens[1:])))

        operator_names = self._dedupe_keep_order(operator_names)
        return machine_id, machine_label, operator_names

    @staticmethod
    def _extract_machine_id(text: str) -> str:
        s = (text or "").strip()
        if not s:
            return ""
        m = re.search(r"(\d{4,})", s)
        if m:
            return m.group(1)
        # 去掉括号说明，如 “钳工（备用）”
        s = re.sub(r"[（(].*?[）)]", "", s).strip()
        return s

    def _extract_names(self, text: str) -> List[str]:
        tokens = [x.strip() for x in re.split(r"[、,，/\s]+", self._to_text(text)) if x and x.strip()]
        result: List[str] = []
        for tk in tokens:
            # 过滤掉纯数字/机台标注
            if tk.isdigit():
                continue
            if not self._CJK_RE.search(tk):
                continue
            # 去掉括号注释
            cleaned = re.sub(r"[（(].*?[）)]", "", tk).strip()
            if cleaned:
                result.append(cleaned)
        return result

    @staticmethod
    def _parse_step_seq(step_text: str) -> Tuple[Optional[int], bool]:
        s = (step_text or "").strip()
        if not s:
            return None, False
        m = re.match(r"^(\d{1,3})-([0-9A-Za-z]+)", s)
        if m:
            try:
                return int(m.group(1)), True
            except Exception:
                return None, True
        m2 = re.match(r"^(\d{1,3})", s)
        if m2:
            try:
                return int(m2.group(1)), False
            except Exception:
                return None, False
        return None, False

    def _parse_route_map(self, route_raw: str) -> Dict[int, str]:
        normalized = self._normalize_route(route_raw)
        matches = re.findall(r"(\d+)([^\d]+)", normalized)
        route: Dict[int, str] = {}
        for seq_s, op_name in matches:
            try:
                seq = int(seq_s)
            except Exception:
                continue
            name = (op_name or "").strip()
            if not name:
                continue
            # 与现有 RouteParser 一致：重复序号保留首个
            if seq not in route:
                route[seq] = name
        return route

    def _infer_op_name_for_missing_seq(
        self,
        seq: int,
        records: Iterable[StepRecord],
        machine_op_hint: Dict[str, Counter],
    ) -> str:
        rec_list = list(records)
        # 1) 优先用 machine->op_type 历史分布推断
        for rec in rec_list:
            c = machine_op_hint.get(rec.machine_id)
            name = self._most_common_key(c)
            if name:
                return name

        # 2) 再用工步文本关键字推断
        merged_text = " ".join([r.step_text for r in rec_list])
        if "铣" in merged_text:
            return "数铣"
        if "车" in merged_text or "镗" in merged_text:
            return "数车"
        if any(k in merged_text for k in ("攻丝", "修锉", "制钝", "去毛刺", "钳")):
            return "钳工"
        if "检" in merged_text:
            return "检验"

        # 3) 最后兜底
        return f"工序{seq}"

    def _aggregate_internal_hours(self, records: Sequence[StepRecord]) -> Tuple[float, float]:
        """
        把工步时间累计为工序时间（小时）。
        - 仅统计 XX-X 工步；
        - 同一工步在多设备重复出现时按工步去重（取该工步在各设备上的最大值后再累加）。
        """
        step_time: Dict[str, Dict[str, float]] = {}
        for rec in records:
            if not rec.has_step_code:
                continue
            step_key = self._extract_step_key(rec.step_text)
            if not step_key:
                continue
            if step_key not in step_time:
                step_time[step_key] = {"setup_min": 0.0, "unit_min": 0.0}
            if isinstance(rec.setup_min, (int, float)) and float(rec.setup_min) > 0:
                step_time[step_key]["setup_min"] = max(step_time[step_key]["setup_min"], float(rec.setup_min))
            if isinstance(rec.unit_min, (int, float)) and float(rec.unit_min) > 0:
                step_time[step_key]["unit_min"] = max(step_time[step_key]["unit_min"], float(rec.unit_min))

        setup_min_sum = sum(v.get("setup_min", 0.0) for v in step_time.values())
        unit_min_sum = sum(v.get("unit_min", 0.0) for v in step_time.values())
        return round(setup_min_sum / 60.0, 6), round(unit_min_sum / 60.0, 6)

    @staticmethod
    def _extract_step_key(step_text: str) -> str:
        s = (step_text or "").strip()
        if not s:
            return ""
        m = re.match(r"^(\d{1,3}-[0-9A-Za-z]+)", s)
        return m.group(1) if m else ""

    @staticmethod
    def _estimate_external_days(records: Sequence[StepRecord]) -> Optional[float]:
        vals: List[float] = []
        for rec in records:
            candidate = None
            if isinstance(rec.batch_min, (int, float)) and rec.batch_min > 0:
                candidate = float(rec.batch_min)
            elif isinstance(rec.setup_min, (int, float)) and isinstance(rec.unit_min, (int, float)):
                total = float(rec.setup_min) + float(rec.unit_min)
                if total > 0:
                    candidate = total
            elif isinstance(rec.unit_min, (int, float)) and rec.unit_min > 0:
                candidate = float(rec.unit_min)
            if candidate and candidate > 0:
                vals.append(candidate / 480.0)  # 8h = 480min

        if not vals:
            return None
        avg = sum(vals) / len(vals)
        return round(max(avg, 0.0001), 6)

    @staticmethod
    def _external_alias(name: str) -> str:
        s = (name or "").strip()
        if not s:
            return "外协工序"
        return s if s.endswith("（外协）") else f"{s}（外协）"

    @staticmethod
    def _most_common_key(counter: Optional[Counter]) -> Optional[str]:
        if not counter:
            return None
        try:
            return counter.most_common(1)[0][0]
        except Exception:
            return None

    @staticmethod
    def _build_machine_name(machine_id: str, machine_label: str) -> str:
        label = (machine_label or "").strip()
        label = re.sub(r"[（(].*?[）)]", "", label).strip()
        mid = (machine_id or "").strip()
        if label and label != mid:
            return label
        return mid or label

    @staticmethod
    def _write_xlsx(path: str, headers: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        wb = openpyxl.Workbook()
        try:
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(list(headers))
            for row in rows:
                ws.append([row.get(h) for h in headers])
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @staticmethod
    def _to_text(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return str(v).strip()

    @staticmethod
    def _to_float(v: Any) -> Optional[float]:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            try:
                fv = float(v)
            except Exception:
                return None
            return fv if fv > 0 else None
        s = str(v).strip()
        if not s:
            return None
        s = s.replace(",", "")
        try:
            fv = float(s)
            return fv if fv > 0 else None
        except Exception:
            return None

    @staticmethod
    def _pick_cell(row_values: Sequence[Any], idx: int) -> Any:
        return row_values[idx] if 0 <= idx < len(row_values) else None

    @staticmethod
    def _dedupe_keep_order(items: Iterable[str]) -> List[str]:
        seen: Set[str] = set()
        out: List[str] = []
        for it in items:
            v = (it or "").strip()
            if not v or v in seen:
                continue
            seen.add(v)
            out.append(v)
        return out

    def _normalize_route(self, route_raw: str) -> str:
        s = self._to_text(route_raw)
        if not s:
            return ""
        s = self._SEPARATORS_RE.sub("", s)
        s = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        return s

