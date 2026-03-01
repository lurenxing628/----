from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl


@dataclass
class StationMeta:
    col_start: int
    machine_id: str
    machine_label: str
    operators: List[str] = field(default_factory=list)


@dataclass
class StepRecord:
    part_no: str
    seq: Optional[int]
    step_text: str
    has_step_code: bool
    machine_id: str
    operators: List[str]
    setup_min: Optional[float]
    unit_min: Optional[float]
    batch_min: Optional[float]


@dataclass
class PartContext:
    part_no: str
    part_name: str
    route_raw: str
    route_map: Dict[int, str] = field(default_factory=dict)
    step_records: List[StepRecord] = field(default_factory=list)


class UnitExcelParser:
    _SEPARATORS_RE = re.compile(r"[\s,，、\-—–→>＞]+")
    _CJK_RE = re.compile(r"[\u4e00-\u9fff]")

    def parse(self, input_path: str, sheet_name: Optional[str] = None) -> Tuple[Dict[str, PartContext], List[StationMeta]]:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            headers = [self._to_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            stations = self._build_station_columns(headers)

            parts: Dict[str, PartContext] = {}
            current_part_no: Optional[str] = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                row_values = list(row)

                current_part_no = self._maybe_update_part_context(parts, current_part_no=current_part_no, row_values=row_values)
                if not current_part_no:
                    continue

                ctx = parts[current_part_no]
                self._append_station_step_records(ctx, row_values=row_values, stations=stations)
            return parts, stations
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _maybe_update_part_context(
        self, parts: Dict[str, PartContext], *, current_part_no: Optional[str], row_values: Sequence[Any]
    ) -> Optional[str]:
        part_no = self._to_text(self._pick_cell(row_values, 0))
        if not part_no:
            return current_part_no if current_part_no in parts else None

        part_name = self._to_text(self._pick_cell(row_values, 1))
        route_raw = self._to_text(self._pick_cell(row_values, 4))
        route_map = self._parse_route_map(route_raw)

        existing_ctx = parts.get(part_no)
        if existing_ctx is None:
            parts[part_no] = PartContext(
                part_no=part_no,
                part_name=part_name or part_no,
                route_raw=route_raw,
                route_map=route_map,
            )
        else:
            if part_name:
                existing_ctx.part_name = part_name
            if route_raw:
                existing_ctx.route_raw = route_raw
            if route_map:
                existing_ctx.route_map = route_map

        return part_no

    def _append_station_step_records(self, ctx: PartContext, *, row_values: Sequence[Any], stations: Sequence[StationMeta]) -> None:
        for station in stations:
            step_text = self._to_text(self._pick_cell(row_values, station.col_start))
            if not step_text:
                continue

            seq, has_step_code = self._parse_step_seq(step_text)
            rec = StepRecord(
                part_no=ctx.part_no,
                seq=seq,
                step_text=step_text,
                has_step_code=has_step_code,
                machine_id=station.machine_id,
                operators=list(station.operators),
                setup_min=self._to_float(self._pick_cell(row_values, station.col_start + 1)),
                unit_min=self._to_float(self._pick_cell(row_values, station.col_start + 2)),
                batch_min=self._to_float(self._pick_cell(row_values, station.col_start + 3)),
            )
            ctx.step_records.append(rec)

    def _build_station_columns(self, headers: Sequence[str]) -> List[StationMeta]:
        stations: List[StationMeta] = []
        idx = 8
        while idx + 3 < len(headers):
            raw_header = self._to_text(headers[idx])
            if raw_header:
                machine_id, machine_label, operators = self._parse_station_header(raw_header)
                if machine_id:
                    stations.append(
                        StationMeta(
                            col_start=idx,
                            machine_id=machine_id,
                            machine_label=machine_label or machine_id,
                            operators=operators,
                        )
                    )
            idx += 4
        return stations

    def _parse_station_header(self, raw_header: str) -> Tuple[str, str, List[str]]:
        text = self._to_text(raw_header)
        lines = [x.strip() for x in re.split(r"[\r\n]+", text) if x and x.strip()]
        if not lines:
            return "", "", []

        machine_label = lines[0]
        machine_id = self._extract_machine_id(machine_label)

        operator_names: List[str] = []
        if len(lines) > 1:
            for ln in lines[1:]:
                operator_names.extend(self._extract_names(ln))
        else:
            tokens = [t for t in re.split(r"[、,，/\s]+", machine_label) if t]
            if tokens:
                first = tokens[0]
                machine_label = first
                machine_id = self._extract_machine_id(first) or first
                operator_names.extend(self._extract_names(" ".join(tokens[1:])))

        operator_names = self._dedupe_keep_order(operator_names)
        return machine_id, machine_label, operator_names

    @staticmethod
    def _extract_machine_id(text: str) -> str:
        s = (text or "").strip()
        if not s:
            return ""
        m = re.search(r"(\d{4,})", s)
        if m:
            return m.group(1)
        s = re.sub(r"[（(].*?[）)]", "", s).strip()
        return s

    def _extract_names(self, text: str) -> List[str]:
        tokens = [x.strip() for x in re.split(r"[、,，/\s]+", self._to_text(text)) if x and x.strip()]
        result: List[str] = []
        for tk in tokens:
            if tk.isdigit():
                continue
            if not self._CJK_RE.search(tk):
                continue
            cleaned = re.sub(r"[（(].*?[）)]", "", tk).strip()
            if cleaned:
                result.append(cleaned)
        return result

    @staticmethod
    def _parse_step_seq(step_text: str) -> Tuple[Optional[int], bool]:
        s = (step_text or "").strip()
        if not s:
            return None, False
        m = re.match(r"^(\d{1,3})-([0-9A-Za-z]+)", s)
        if m:
            try:
                return int(m.group(1)), True
            except Exception:
                return None, True
        m2 = re.match(r"^(\d{1,3})", s)
        if m2:
            try:
                return int(m2.group(1)), False
            except Exception:
                return None, False
        return None, False

    def _parse_route_map(self, route_raw: str) -> Dict[int, str]:
        normalized = self._normalize_route(route_raw)
        matches = re.findall(r"(\d+)([^\d]+)", normalized)
        route: Dict[int, str] = {}
        for seq_s, op_name in matches:
            try:
                seq = int(seq_s)
            except Exception:
                continue
            name = (op_name or "").strip()
            if not name:
                continue
            if seq not in route:
                route[seq] = name
        return route

    @staticmethod
    def _to_text(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return str(v).strip()

    @staticmethod
    def _to_float(v: Any) -> Optional[float]:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            try:
                fv = float(v)
            except Exception:
                return None
            return fv if fv > 0 else None
        s = str(v).strip()
        if not s:
            return None
        s = s.replace(",", "")
        try:
            fv = float(s)
            return fv if fv > 0 else None
        except Exception:
            return None

    @staticmethod
    def _pick_cell(row_values: Sequence[Any], idx: int) -> Any:
        return row_values[idx] if 0 <= idx < len(row_values) else None

    @staticmethod
    def _dedupe_keep_order(items: Iterable[str]) -> List[str]:
        seen: Set[str] = set()
        out: List[str] = []
        for it in items:
            v = (it or "").strip()
            if not v or v in seen:
                continue
            seen.add(v)
            out.append(v)
        return out

    def _normalize_route(self, route_raw: str) -> str:
        s = self._to_text(route_raw)
        if not s:
            return ""
        s = self._SEPARATORS_RE.sub("", s)
        s = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        return s

