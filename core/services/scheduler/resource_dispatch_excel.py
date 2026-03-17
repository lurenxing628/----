from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from core.services.common.excel_templates import _sanitize_export_cell


def _auto_width(ws) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if column_letter is None:
                column_letter = cell.column_letter
            text = "" if cell.value is None else str(cell.value)
            if len(text) > max_length:
                max_length = len(text)
        if column_letter is not None:
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)


def _append_row(ws, values: Sequence[Any]) -> None:
    ws.append([_sanitize_export_cell(v) for v in values])


def _write_table(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    _append_row(ws, list(headers))
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        _append_row(ws, list(row))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _auto_width(ws)


def _detail_headers() -> List[str]:
    return [
        "排程ID",
        "工序ID",
        "工序编码",
        "批次号",
        "图号",
        "工序",
        "工序名称",
        "开始时间",
        "结束时间",
        "时长(分钟)",
        "查询对象",
        "对应资源",
        "查询对象班组",
        "对应资源班组",
        "班组关系",
        "来源",
        "锁定状态",
        "是否跨天",
        "是否超期",
    ]


def _first_present(row: Dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return default


def _yes_no_label(value: Any) -> str:
    return "是" if value else "否"


def _build_detail_row(row: Dict[str, Any]) -> List[Any]:
    seq = row.get("seq")
    return [
        _first_present(row, "schedule_id"),
        _first_present(row, "op_id"),
        _first_present(row, "op_code"),
        _first_present(row, "batch_id"),
        _first_present(row, "part_no"),
        seq if seq is not None else "",
        _first_present(row, "op_type_name"),
        _first_present(row, "start_time"),
        _first_present(row, "end_time"),
        _first_present(row, "duration_minutes", default=0),
        _first_present(row, "current_resource_label", "scope_label"),
        _first_present(row, "counterpart_resource_label"),
        _first_present(row, "current_team_name", "current_team_id"),
        _first_present(row, "counterpart_team_name", "counterpart_team_id"),
        _first_present(row, "team_relation_label"),
        _first_present(row, "source"),
        _first_present(row, "lock_status"),
        _yes_no_label(row.get("is_cross_day")),
        _yes_no_label(row.get("is_overdue")),
    ]


def _detail_table_rows(detail_rows: Sequence[Dict[str, Any]]) -> List[List[Any]]:
    return [_build_detail_row(row) for row in detail_rows]


def _calendar_table_rows(calendar_rows: Sequence[Dict[str, Any]]) -> List[List[Any]]:
    table: List[List[Any]] = []
    for row in calendar_rows:
        line = [row.get("scope_label") or ""]
        cells = row.get("cells") or []
        for cell in cells:
            line.append(cell.get("text") or "")
        table.append(line)
    return table


def _summary_pairs(filters: Dict[str, Any], summary: Dict[str, Any]) -> List[List[Any]]:
    return [
        ["视角", filters.get("scope_type_label") or ""],
        ["查询对象", filters.get("scope_label") or ""],
        ["班组轴", filters.get("team_axis_label") or ""],
        ["区间类型", filters.get("period_preset_label") or ""],
        ["开始日期", filters.get("start_date") or ""],
        ["结束日期", filters.get("end_date") or ""],
        ["排产版本", filters.get("version") or ""],
        ["任务数量", summary.get("total_tasks") or 0],
        ["总工时（小时）", summary.get("total_hours") or 0],
        ["跨天任务", summary.get("cross_day_count") or 0],
        ["超期批次任务", summary.get("overdue_count") or 0],
        ["外协/未分配", summary.get("external_count") or 0],
        ["跨班组借调", summary.get("cross_team_count") or 0],
    ]


def _write_summary_sheet(wb: Workbook, filters: Dict[str, Any], summary: Dict[str, Any]) -> None:
    ws_summary = wb.create_sheet("查询摘要")
    for key, value in _summary_pairs(filters, summary):
        _append_row(ws_summary, [key, value])
    ws_summary.freeze_panes = "A2"
    for row in ws_summary.iter_rows():
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _auto_width(ws_summary)


def _write_calendar_sheet(wb: Workbook, title: str, headers: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    _write_table(ws, ["查询对象"] + list(headers), _calendar_table_rows(rows))


def _write_detail_sheet(wb: Workbook, title: str, rows: Sequence[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    _write_table(ws, _detail_headers(), _detail_table_rows(rows))


def _write_team_scope_sheets(wb: Workbook, payload: Dict[str, Any]) -> None:
    _write_detail_sheet(wb, "班组人员任务明细", list(payload.get("operator_rows") or []))
    _write_detail_sheet(wb, "班组设备任务明细", list(payload.get("machine_rows") or []))
    _write_calendar_sheet(
        wb,
        "班组人员日历",
        list(payload.get("operator_calendar_headers") or []),
        list(payload.get("operator_calendar_rows") or []),
    )
    _write_calendar_sheet(
        wb,
        "班组设备日历",
        list(payload.get("machine_calendar_headers") or []),
        list(payload.get("machine_calendar_rows") or []),
    )
    cross_team_rows = list(payload.get("cross_team_rows") or [])
    if cross_team_rows:
        _write_detail_sheet(wb, "跨班组借调", cross_team_rows)


def _write_resource_scope_sheets(wb: Workbook, payload: Dict[str, Any]) -> None:
    _write_detail_sheet(wb, "任务明细", list(payload.get("detail_rows") or []))
    _write_calendar_sheet(
        wb,
        "日历排班",
        list(payload.get("calendar_headers") or []),
        list(payload.get("calendar_rows") or []),
    )


def build_resource_dispatch_workbook(payload: Dict[str, Any]) -> BytesIO:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    filters = payload.get("filters") or {}
    summary = payload.get("summary") or {}
    _write_summary_sheet(wb, filters, summary)

    if str(filters.get("scope_type") or "") == "team":
        _write_team_scope_sheets(wb, payload)
    else:
        _write_resource_scope_sheets(wb, payload)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
