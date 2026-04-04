from __future__ import annotations

import importlib
import os
from typing import Any, Dict, Iterable, List, Optional, cast

from core.infrastructure.errors import AppError, ErrorCode

from .tabular_backend import SOURCE_ROW_NUM_KEY, SOURCE_SHEET_NAME_KEY, TabularBackend


def _normalize_header_key(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        s = value.strip()
        return None if s == "" else s
    return value


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _normalize_record(record: Any) -> Dict[str, Any]:
    item: Dict[str, Any] = {}
    if not record:
        return item
    for k, v in (record or {}).items():
        key = _normalize_header_key(k)
        if not key:
            continue
        item[key] = _normalize_cell_value(v)
    return item


def _is_blank_item(item: Dict[str, Any]) -> bool:
    if not item:
        return True
    return all(_is_blank_value(v) for v in item.values())


def _iter_sheet_rows(ws: Any) -> Iterable[Any]:
    iter_rows = getattr(ws, "iter_rows", None)
    if not callable(iter_rows):
        raise AppError(
            code=ErrorCode.EXCEL_READ_ERROR,
            message="读取 Excel 文件失败（pandas 辅助行号恢复）：工作表对象不支持逐行读取。",
        )
    rows_iter = cast(Iterable[Any], iter_rows(values_only=True))
    return rows_iter


def _openpyxl_source_rows(file_path: str, sheet: Optional[str]) -> List[int]:
    """
    pandas 读取时可能压缩空行；这里用 openpyxl 复算非空数据行的原始 Excel 行号。
    若辅助读取失败，调用方会退回到顺序行号。
    """
    import openpyxl

    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        source_rows: List[int] = []
        for excel_row_num, row in enumerate(_iter_sheet_rows(ws), start=1):
            if excel_row_num == 1:
                continue
            if row is None:
                continue
            if all(_is_blank_value(v) for v in row):
                continue
            source_rows.append(int(excel_row_num))
        return source_rows
    finally:
        close_fn = getattr(wb, "close", None)
        if callable(close_fn):
            close_fn()


class PandasBackend(TabularBackend):
    """
    可选 Excel 后端：pandas + numpy（读取/写入 xlsx）。

    约束：对外仍只收发 List[Dict]，不暴露 DataFrame。
    """

    def read(self, file_path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
        try:
            pd = importlib.import_module("pandas")

            df = pd.read_excel(file_path, sheet_name=sheet or 0, dtype=object)
            # 统一空值
            df = df.where(pd.notnull(df), None)
            raw_rows = df.to_dict(orient="records")
            try:
                source_row_nums = _openpyxl_source_rows(file_path, sheet)
            except Exception:
                source_row_nums = []

            result: List[Dict[str, Any]] = []
            for idx, r in enumerate(raw_rows, start=2):
                item = _normalize_record(r)
                # 跳过空行
                if _is_blank_item(item):
                    continue
                item[SOURCE_ROW_NUM_KEY] = int(source_row_nums[len(result)] if len(result) < len(source_row_nums) else idx)
                if sheet:
                    item[SOURCE_SHEET_NAME_KEY] = str(sheet)
                result.append(item)
            return result
        except AppError:
            raise
        except Exception as e:
            raise AppError(
                code=ErrorCode.EXCEL_READ_ERROR,
                message="读取 Excel 文件失败（pandas），请确认文件未损坏且未被其他程序占用。",
                details={"file_path": file_path, "sheet": sheet},
                cause=e,
            ) from e

    def write(self, rows: List[Dict[str, Any]], file_path: str, sheet: str = "Sheet1") -> None:
        try:
            pd = importlib.import_module("pandas")

            os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)

            if not rows:
                # 仍生成一个空文件（只有标题行/空表）
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    pd.DataFrame([]).to_excel(writer, index=False, sheet_name=sheet)
                return

            headers = list(rows[0].keys())
            df = pd.DataFrame(rows)
            # 尽量保持列顺序与首行一致
            df = df.reindex(columns=headers)

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet)
        except AppError:
            raise
        except Exception as e:
            raise AppError(
                code=ErrorCode.EXCEL_WRITE_ERROR,
                message="写入 Excel 文件失败（pandas），请确认目标路径可写。",
                details={"file_path": file_path, "sheet": sheet},
                cause=e,
            ) from e
