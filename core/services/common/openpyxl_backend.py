from __future__ import annotations

import os
from typing import Any, Dict, Iterable, List, Optional, cast

import openpyxl

from core.infrastructure.errors import AppError, ErrorCode

from .tabular_backend import SOURCE_ROW_NUM_KEY, SOURCE_SHEET_NAME_KEY, TabularBackend


def _normalize_header_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        s = value.strip()
        return None if s == "" else s
    return value


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    try:
        return str(value).strip() == ""
    except Exception:
        return False


def _is_blank_row(raw: Any) -> bool:
    if raw is None:
        return True
    return all(_is_blank_cell(v) for v in raw)


def _iter_sheet_rows(ws: Any) -> List[Any]:
    iter_rows = getattr(ws, "iter_rows", None)
    if not callable(iter_rows):
        raise AppError(
            code=ErrorCode.EXCEL_READ_ERROR,
            message="读取 Excel 文件失败：工作表对象不支持逐行读取。",
        )
    rows_iter = cast(Iterable[Any], iter_rows(values_only=True))
    return list(rows_iter)


def _append_sheet_row(ws: Any, row: List[Any]) -> None:
    append = getattr(ws, "append", None)
    if not callable(append):
        raise AppError(
            code=ErrorCode.EXCEL_WRITE_ERROR,
            message="写入 Excel 文件失败：工作表对象不支持追加行。",
        )
    append(row)


def _row_to_item(
    headers: List[str],
    raw: Any,
    *,
    source_row_num: int,
    source_sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    item: Dict[str, Any] = {SOURCE_ROW_NUM_KEY: int(source_row_num)}
    if source_sheet_name:
        item[SOURCE_SHEET_NAME_KEY] = str(source_sheet_name)
    for idx, key in enumerate(headers):
        if not key:
            continue
        val = raw[idx] if idx < len(raw) else None
        item[key] = _normalize_cell_value(val)
    return item


def _convert_worksheet_rows(rows: List[Any], *, source_sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    if not rows:
        return []

    headers = [_normalize_header_cell(h) for h in rows[0]]
    result: List[Dict[str, Any]] = []

    for excel_row_num, raw in enumerate(rows[1:], start=2):
        # 跳过空行，但保留原始 Excel 行号，避免后续错误定位漂移。
        if _is_blank_row(raw):
            continue
        result.append(
            _row_to_item(
                headers,
                raw,
                source_row_num=excel_row_num,
                source_sheet_name=source_sheet_name,
            )
        )

    return result


class OpenpyxlBackend(TabularBackend):
    """V1 Excel 后端：只依赖 openpyxl；对外仅收发 List[Dict]。"""

    def read(self, file_path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet] if sheet else wb.active
            rows = _iter_sheet_rows(ws)

            return _convert_worksheet_rows(rows, source_sheet_name=getattr(ws, "title", None))
        except AppError:
            raise
        except Exception as e:
            raise AppError(
                code=ErrorCode.EXCEL_READ_ERROR,
                message="读取 Excel 文件失败，请确认文件未损坏且未被其他程序占用。",
                details={"file_path": file_path, "sheet": sheet},
                cause=e,
            ) from e
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    def write(self, rows: List[Dict[str, Any]], file_path: str, sheet: str = "Sheet1") -> None:
        wb = None
        try:
            os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise AppError(code=ErrorCode.EXCEL_WRITE_ERROR, message="写入 Excel 文件失败：未能创建默认工作表。")
            ws.title = sheet

            if not rows:
                wb.save(file_path)
                return

            headers = list(rows[0].keys())
            _append_sheet_row(ws, headers)
            for r in rows:
                _append_sheet_row(ws, [r.get(h) for h in headers])

            wb.save(file_path)
        except AppError:
            raise
        except Exception as e:
            raise AppError(
                code=ErrorCode.EXCEL_WRITE_ERROR,
                message="写入 Excel 文件失败，请确认目标路径可写。",
                details={"file_path": file_path, "sheet": sheet},
                cause=e,
            ) from e
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
