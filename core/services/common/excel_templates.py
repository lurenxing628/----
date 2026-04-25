from __future__ import annotations

import io
import os
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Set

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def _close_workbook_quietly(wb: Any) -> None:
    try:
        wb.close()
    except Exception:
        return


def _active_sheet_or_none(wb: Any) -> Any:
    try:
        return wb.active
    except Exception:
        return None


def _require_active_sheet(wb: Any) -> Any:
    ws = _active_sheet_or_none(wb)
    if ws is None:
        raise ValueError("Workbook 缺少活动工作表")
    return ws


def _sanitize_export_cell(value: Any) -> Any:
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _iter_column_indices(spec: Mapping[str, Any], key: str) -> Iterable[int]:
    for col_idx in spec.get(key, []) or []:
        if isinstance(col_idx, int) and col_idx >= 0:
            yield col_idx


def _apply_number_format(ws, col_indices: Iterable[int], fmt: str, max_row: int) -> None:
    for col_idx in col_indices:
        col_letter = get_column_letter(col_idx + 1)
        for row_idx in range(2, max_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = fmt


def _apply_enum_validations(ws, enum_cols: Mapping[int, Sequence[str]], max_row: int) -> None:
    for col_idx, values in (enum_cols or {}).items():
        if not isinstance(col_idx, int) or col_idx < 0:
            continue
        items = [str(v).strip() for v in (values or []) if str(v).strip()]
        if not items:
            continue
        dv = DataValidation(type="list", formula1=f"\"{','.join(items)}\"", allow_blank=True)
        col_letter = get_column_letter(col_idx + 1)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")
        ws.add_data_validation(dv)


def _auto_width(ws, *, explicit_widths: Mapping[int, int] | None = None) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(text))
    for col_idx, content_width in widths.items():
        width = min(max(content_width + 2, 12), 36)
        if explicit_widths and (col_idx - 1) in explicit_widths:
            width = max(width, int(explicit_widths[col_idx - 1]))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if explicit_widths:
        for zero_based_idx, width in explicit_widths.items():
            ws.column_dimensions[get_column_letter(int(zero_based_idx) + 1)].width = max(
                ws.column_dimensions[get_column_letter(int(zero_based_idx) + 1)].width or 0,
                int(width),
            )


def _apply_sheet_layout(
    ws,
    *,
    format_spec: Mapping[str, Any] | None,
    data_row_count: int,
) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    spec = format_spec or {}
    max_row = max(int(data_row_count) + 200, 500)
    _apply_number_format(ws, _iter_column_indices(spec, "text_cols"), "@", max_row)
    _apply_number_format(ws, _iter_column_indices(spec, "int_cols"), "0", max_row)
    _apply_number_format(ws, _iter_column_indices(spec, "float_cols"), "0.00", max_row)
    _apply_number_format(ws, _iter_column_indices(spec, "date_cols"), "yyyy-mm-dd", max_row)
    _apply_number_format(ws, _iter_column_indices(spec, "time_cols"), "hh:mm", max_row)
    _apply_enum_validations(ws, spec.get("enum_cols", {}) or {}, max_row)
    _auto_width(ws, explicit_widths=spec.get("column_widths", {}) or {})


def build_xlsx_bytes(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]] = (),
    *,
    format_spec: Mapping[str, Any] | None = None,
    sheet_title: str = "Sheet1",
    sanitize_formula: bool = False,
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    try:
        ws = _require_active_sheet(wb)
        ws.title = sheet_title
        ws.append(list(headers))
        for r in rows:
            row_values = [_sanitize_export_cell(v) if sanitize_formula else v for v in r]
            ws.append(list(row_values))
        _apply_sheet_layout(ws, format_spec=format_spec, data_row_count=len(rows))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    finally:
        _close_workbook_quietly(wb)


def _write_xlsx(
    path: str,
    headers: Sequence[str],
    sample_rows: Sequence[Sequence[Any]] = (),
    *,
    format_spec: Mapping[str, Any] | None = None,
) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    output = build_xlsx_bytes(headers, sample_rows, format_spec=format_spec)
    with open(path, "wb") as f:
        f.write(output.getvalue())


def _read_xlsx_headers(path: str) -> List[str]:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        ws = _active_sheet_or_none(wb)
        if ws is None:
            return []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            return []
        return ["" if value is None else str(value).strip() for value in first_row]
    except Exception:
        return []
    finally:
        _close_workbook_quietly(wb)


_LEGACY_TEMPLATE_ENUM_VALUES: Set[str] = {
    "active",
    "inactive",
    "maintain",
    "internal",
    "external",
    "beginner",
    "normal",
    "expert",
    "urgent",
    "critical",
    "partial",
    "workday",
    "holiday",
    "yes",
    "no",
}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_has_value(row: Sequence[Any]) -> bool:
    return any(_cell_text(value) for value in row)


def _inline_validation_values(formula: Any) -> List[str]:
    text = _cell_text(formula)
    if len(text) < 2 or not (text.startswith('"') and text.endswith('"')):
        return []
    return [item.strip() for item in text[1:-1].split(",") if item.strip()]


def _non_empty_data_rows(ws: Any) -> List[Sequence[Any]]:
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if _row_has_value(row)]


def _enum_texts_from_row(row: Sequence[Any], enum_col_indices: Iterable[int]) -> Set[str]:
    texts: Set[str] = set()
    for col_idx in enum_col_indices:
        if col_idx >= len(row):
            continue
        text = _cell_text(row[col_idx]).lower()
        if text:
            texts.add(text)
    return texts


def _enum_texts_from_rows(rows: Iterable[Sequence[Any]], enum_col_indices: Iterable[int]) -> Set[str]:
    texts: Set[str] = set()
    for row in rows:
        texts.update(_enum_texts_from_row(row, enum_col_indices))
    return texts


def _inline_validation_texts(ws: Any) -> Set[str]:
    texts: Set[str] = set()
    for dv in getattr(ws.data_validations, "dataValidation", []) or []:
        for value in _inline_validation_values(getattr(dv, "formula1", "")):
            texts.add(value.lower())
    return texts


def _known_generated_template_needs_refresh(path: str, template_def: Mapping[str, Any]) -> bool:
    """Refresh old generated templates that still expose raw enum defaults.

    The guard is intentionally narrow: if a workbook contains user data beyond the
    built-in sample rows, keep it untouched and let download serve the disk file.
    """
    format_spec = template_def.get("format_spec", {}) or {}
    enum_col_indices = set(format_spec.get("enum_cols", {}) or {})
    if not enum_col_indices:
        return False

    sample_row_count = len(template_def.get("sample_rows") or [])
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return False
    try:
        ws = _active_sheet_or_none(wb)
        if ws is None:
            return False

        non_empty_data_rows = _non_empty_data_rows(ws)
        if len(non_empty_data_rows) > sample_row_count:
            return False

        enum_texts = _enum_texts_from_rows(non_empty_data_rows, enum_col_indices)
        enum_texts.update(_inline_validation_texts(ws))

        return bool(enum_texts & _LEGACY_TEMPLATE_ENUM_VALUES)
    except Exception:
        return False
    finally:
        _close_workbook_quietly(wb)


def get_template_definition(filename: str) -> Dict[str, Any]:
    for item in get_default_templates():
        if str(item.get("filename")) == str(filename):
            return item
    raise KeyError(f"unknown template: {filename}")


def get_default_templates() -> List[Dict[str, Any]]:
    """
    返回需要交付的 Excel 模板清单（文件名 + 表头 + 示例行）。

    说明：
    - 这些模板与各模块的“下载模板”接口保持一致（列名为中文）。
    - 即便某些模块（如批次/日历）尚未在 Phase0~5 提供页面/接口，也可先把模板交付到目录中，
      便于后续直接复用，减少再回读开发文档的成本。
    """
    return [
        # 人员
        {
            "filename": "人员基本信息.xlsx",
            "headers": ["工号", "姓名", "状态", "班组", "备注"],
            "sample_rows": [["OP001", "张三", "在岗", None, "示例备注"]],
            "format_spec": {
                "text_cols": [0, 1, 3, 4],
                "enum_cols": {2: ["在岗", "停用"]},
                "column_widths": {0: 14, 1: 12, 2: 12, 3: 14, 4: 18},
            },
        },
        {
            "filename": "人员设备关联.xlsx",
            "headers": ["工号", "设备编号", "技能等级", "主操设备"],
            "sample_rows": [["OP001", "CNC-01", "普通", "是"]],
            "format_spec": {
                "text_cols": [0, 1],
                "enum_cols": {2: ["初级", "普通", "熟练"], 3: ["是", "否"]},
                "column_widths": {0: 14, 1: 14, 2: 12, 3: 12},
            },
        },
        # 设备
        {
            "filename": "设备信息.xlsx",
            "headers": ["设备编号", "设备名称", "工种", "班组", "状态"],
            "sample_rows": [["CNC-01", "数控车床1", "数车", None, "可用"]],
            "format_spec": {
                "text_cols": [0, 1, 2, 3],
                "enum_cols": {4: ["可用", "停用", "维修"]},
                "column_widths": {0: 14, 1: 18, 2: 12, 3: 14, 4: 12},
            },
        },
        {
            "filename": "设备人员关联.xlsx",
            "headers": ["设备编号", "工号", "技能等级", "主操设备"],
            "sample_rows": [["CNC-01", "OP001", "普通", "是"]],
            "format_spec": {
                "text_cols": [0, 1],
                "enum_cols": {2: ["初级", "普通", "熟练"], 3: ["是", "否"]},
                "column_widths": {0: 14, 1: 14, 2: 12, 3: 12},
            },
        },
        # 工艺
        {
            "filename": "工种配置.xlsx",
            "headers": ["工种ID", "工种名称", "归属"],
            "sample_rows": [["OT001", "数车", "内部"], ["OT002", "标印", "外部"]],
            "format_spec": {
                "text_cols": [0, 1],
                "enum_cols": {2: ["内部", "外部"]},
                "column_widths": {0: 14, 1: 16, 2: 12},
            },
        },
        {
            "filename": "供应商配置.xlsx",
            "headers": ["供应商ID", "名称", "对应工种", "默认周期", "状态", "备注"],
            "sample_rows": [["S001", "外协-标印厂", "标印", 1, "启用", "示例备注"]],
            "format_spec": {
                "text_cols": [0, 1, 2, 4, 5],
                "float_cols": [3],
                "enum_cols": {4: ["启用", "停用"]},
                "column_widths": {0: 14, 1: 18, 2: 14, 3: 12, 4: 12, 5: 18},
            },
        },
        {
            "filename": "零件工艺路线.xlsx",
            "headers": ["图号", "名称", "工艺路线字符串"],
            "sample_rows": [["A1234", "壳体-大", "5数铣10钳20数车35标印40总检45表处理"]],
            "format_spec": {
                "text_cols": [0, 1, 2],
                "column_widths": {0: 14, 1: 18, 2: 32},
            },
        },
        {
            "filename": "零件工序工时.xlsx",
            "headers": ["图号", "工序", "换型时间(h)", "单件工时(h)"],
            "sample_rows": [["A1234", 5, 1.0, 0.25], ["A1234", 10, 0.5, 0.1]],
            "format_spec": {
                "text_cols": [0],
                "int_cols": [1],
                "float_cols": [2, 3],
                "column_widths": {0: 14, 1: 10, 2: 14, 3: 14},
            },
        },
        # 排产（后续阶段会接入接口/页面；模板先交付）
        {
            "filename": "批次信息.xlsx",
            # 对齐路由 `web/routes/scheduler_excel_batches.py` 的兜底模板与导入字段（含齐套日期）
            "headers": ["批次号", "图号", "数量", "交期", "优先级", "齐套", "齐套日期", "备注"],
            "sample_rows": [
                ["B001", "A1234", 50, "2026-01-25", "急件", "齐套", "2026-01-24", "示例"]
            ],
            "format_spec": {
                "text_cols": [0, 1, 7],
                "int_cols": [2],
                "date_cols": [3, 6],
                "enum_cols": {
                    4: ["普通", "急件", "特急"],
                    5: ["齐套", "未齐套", "部分齐套"],
                },
                "column_widths": {0: 14, 1: 14, 2: 10, 3: 12, 4: 12, 5: 12, 6: 12, 7: 18},
            },
        },
        {
            "filename": "工作日历.xlsx",
            "headers": ["日期", "类型", "可用工时", "效率", "允许普通件", "允许急件", "说明"],
            "sample_rows": [["2026-01-21", "工作日", 8, 1.0, "是", "是", "示例"]],
            "format_spec": {
                "date_cols": [0],
                "float_cols": [2, 3],
                "enum_cols": {1: ["工作日", "假期"], 4: ["是", "否"], 5: ["是", "否"]},
                "text_cols": [6],
                "column_widths": {0: 12, 1: 12, 2: 12, 3: 12, 4: 12, 5: 12, 6: 18},
            },
        },
        {
            "filename": "人员专属工作日历.xlsx",
            "headers": ["工号", "日期", "类型", "班次开始", "班次结束", "可用工时", "效率", "允许普通件", "允许急件", "说明"],
            "sample_rows": [
                ["OP001", "2026-01-25", "假期", "08:00", "", 0, 0.8, "否", "否", "示例：休假"],
                ["OP001", "2026-01-26", "假期", "08:00", "16:00", "", "", "是", "是", "示例：假期加班（用班次结束推导工时）"],
            ],
            "format_spec": {
                "text_cols": [0, 9],
                "date_cols": [1],
                "time_cols": [3, 4],
                "float_cols": [5, 6],
                "enum_cols": {2: ["工作日", "假期"], 7: ["是", "否"], 8: ["是", "否"]},
                "column_widths": {0: 14, 1: 12, 2: 12, 3: 10, 4: 10, 5: 12, 6: 12, 7: 12, 8: 12, 9: 18},
            },
        },
    ]


def ensure_excel_templates(template_dir: str) -> Dict[str, Any]:
    """
    确保 templates_excel 目录下存在所有交付模板文件。
    - 若文件存在则不覆盖（避免用户手工修改被覆盖）
    - 若文件存在但表头与当前真源不一致，则按当前真源重建
    - 若目录为空/缺失，则补齐
    """
    template_dir = os.path.abspath(template_dir or "templates_excel")
    os.makedirs(template_dir, exist_ok=True)

    created: List[str] = []
    skipped: List[str] = []

    for t in get_default_templates():
        filename = t["filename"]
        path = os.path.join(template_dir, filename)
        expected_headers = [str(header).strip() for header in (t.get("headers") or [])]
        if (
            os.path.exists(path)
            and _read_xlsx_headers(path) == expected_headers
            and not _known_generated_template_needs_refresh(path, t)
        ):
            skipped.append(filename)
            continue
        _write_xlsx(
            path,
            headers=t["headers"],
            sample_rows=t.get("sample_rows") or [],
            format_spec=t.get("format_spec"),
        )
        created.append(filename)

    return {"template_dir": template_dir, "created": created, "skipped": skipped}
