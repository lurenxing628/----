from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from core.infrastructure.errors import ValidationError
from core.services.scheduler.calendar_service import CalendarService
from data.repositories import ScheduleHistoryRepository, ScheduleRepository


def _parse_date(value: Any, field: str) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value or "").strip().replace("/", "-")
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        raise ValidationError("日期格式不合法（期望：YYYY-MM-DD）", field=field)


def _parse_dt(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip().replace("/", "-").replace("T", " ").replace("：", ":")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def _overlap_seconds(a_start: datetime, a_end: datetime, b_start: datetime, b_end: datetime) -> float:
    s = max(a_start, b_start)
    e = min(a_end, b_end)
    if e <= s:
        return 0.0
    return float((e - s).total_seconds())


@dataclass
class ReportExport:
    filename: str
    content_type: str
    data: io.BytesIO


class ReportEngine:
    """
    报表引擎（基于现有 DB 表直接汇总）。
    """

    def __init__(self, conn, logger=None):
        self.conn = conn
        self.logger = logger
        self.schedule_repo = ScheduleRepository(conn, logger=logger)
        self.history_repo = ScheduleHistoryRepository(conn, logger=logger)
        self.calendar = CalendarService(conn, logger=logger)

    # -------------------------
    # Version helpers
    # -------------------------
    def list_versions(self, limit: int = 30) -> List[Dict[str, Any]]:
        return list(self.history_repo.list_versions(limit=int(limit)))

    def latest_version(self) -> int:
        return int(self.history_repo.get_latest_version() or 0)

    # -------------------------
    # 1) 超期清单
    # -------------------------
    def overdue_batches(self, version: int) -> Dict[str, Any]:
        v = int(version or 0)
        rows = self.conn.execute(
            """
            SELECT
              b.batch_id AS batch_id,
              b.part_no AS part_no,
              b.part_name AS part_name,
              b.quantity AS quantity,
              b.due_date AS due_date,
              MAX(s.end_time) AS finish_time
            FROM Batches b
            LEFT JOIN BatchOperations bo ON bo.batch_id = b.batch_id
            LEFT JOIN Schedule s ON s.op_id = bo.id AND s.version = ?
            WHERE b.due_date IS NOT NULL AND TRIM(CAST(b.due_date AS TEXT)) <> ''
            GROUP BY b.batch_id
            ORDER BY b.due_date ASC, b.batch_id ASC
            """,
            (v,),
        ).fetchall()

        items: List[Dict[str, Any]] = []
        for r in rows:
            due_s = r["due_date"]
            finish_s = r["finish_time"]
            due_d = _parse_dt(due_s)
            finish_dt = _parse_dt(finish_s)
            if not due_d or not finish_dt:
                continue
            due_end = datetime(due_d.year, due_d.month, due_d.day, 23, 59, 59)
            if finish_dt <= due_end:
                continue
            delay_sec = (finish_dt - due_end).total_seconds()
            delay_hours = round(delay_sec / 3600.0, 2)
            delay_days = round(delay_sec / 86400.0, 2)
            items.append(
                {
                    "batch_id": r["batch_id"],
                    "part_no": r["part_no"],
                    "part_name": r["part_name"],
                    "quantity": r["quantity"],
                    "due_date": due_s,
                    "finish_time": finish_s,
                    "delay_hours": delay_hours,
                    "delay_days": delay_days,
                }
            )

        return {"version": v, "count": len(items), "items": items}

    def export_overdue_xlsx(self, version: int) -> ReportExport:
        rep = self.overdue_batches(version)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "overdue"
        ws.append(["批次号", "图号", "名称", "数量", "交期", "完工时间", "超期(小时)", "超期(天)"])
        for it in rep["items"]:
            ws.append(
                [
                    it.get("batch_id"),
                    it.get("part_no"),
                    it.get("part_name"),
                    it.get("quantity"),
                    it.get("due_date"),
                    it.get("finish_time"),
                    it.get("delay_hours"),
                    it.get("delay_days"),
                ]
            )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return ReportExport(
            filename=f"超期清单_v{int(rep['version'])}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=buf,
        )

    # -------------------------
    # 2) 资源负荷/利用率
    # -------------------------
    def _capacity_hours(self, start_d: date, end_d: date) -> float:
        # 以“日历的工作窗 * efficiency”作为单资源可用工时（简化：不区分设备/人员差异）
        total = 0.0
        cur = start_d
        while cur <= end_d:
            p = self.calendar._policy_for_datetime(datetime.combine(cur, datetime.min.time()))  # type: ignore[attr-defined]
            if float(p.shift_hours or 0.0) > 0:
                total += float(p.shift_hours or 0.0) * float(p.efficiency or 1.0)
            cur = cur + timedelta(days=1)
        return float(round(total, 6))

    def utilization(self, version: int, start_date: Any, end_date: Any) -> Dict[str, Any]:
        v = int(version or 0)
        sd = _parse_date(start_date, field="start_date")
        ed = _parse_date(end_date, field="end_date")
        if ed < sd:
            raise ValidationError("结束日期不能早于开始日期", field="end_date")

        start_dt = datetime(sd.year, sd.month, sd.day, 0, 0, 0)
        end_dt_excl = datetime(ed.year, ed.month, ed.day, 0, 0, 0) + timedelta(days=1)
        start_s = start_dt.strftime("%Y-%m-%d %H:%M:%S")
        end_s = end_dt_excl.strftime("%Y-%m-%d %H:%M:%S")

        rows = self.schedule_repo.list_overlapping_with_details(start_s, end_s, v)

        cap_hours = self._capacity_hours(sd, ed)
        if cap_hours <= 0:
            cap_hours = 0.0

        by_machine: Dict[str, Dict[str, Any]] = {}
        by_operator: Dict[str, Dict[str, Any]] = {}

        for r in rows:
            if (r.get("source") or "").strip() != "internal":
                continue
            s_dt = _parse_dt(r.get("start_time"))
            e_dt = _parse_dt(r.get("end_time"))
            if not s_dt or not e_dt:
                continue
            sec = _overlap_seconds(s_dt, e_dt, start_dt, end_dt_excl)
            if sec <= 0:
                continue
            hours = sec / 3600.0

            mc = (r.get("machine_id") or "").strip()
            if mc:
                it = by_machine.setdefault(
                    mc,
                    {"machine_id": mc, "machine_name": r.get("machine_name"), "hours": 0.0, "task_count": 0},
                )
                it["hours"] = float(it["hours"]) + float(hours)
                it["task_count"] = int(it["task_count"]) + 1

            op = (r.get("operator_id") or "").strip()
            if op:
                it = by_operator.setdefault(
                    op,
                    {"operator_id": op, "operator_name": r.get("operator_name"), "hours": 0.0, "task_count": 0},
                )
                it["hours"] = float(it["hours"]) + float(hours)
                it["task_count"] = int(it["task_count"]) + 1

        machine_rows = []
        for it in by_machine.values():
            h = float(it["hours"])
            util = (h / cap_hours) if cap_hours > 0 else None
            machine_rows.append({**it, "hours": round(h, 2), "capacity_hours": round(cap_hours, 2), "utilization": round(util, 4) if util is not None else None})
        machine_rows.sort(key=lambda x: (-(x.get("hours") or 0.0), x.get("machine_id") or ""))

        operator_rows = []
        for it in by_operator.values():
            h = float(it["hours"])
            util = (h / cap_hours) if cap_hours > 0 else None
            operator_rows.append({**it, "hours": round(h, 2), "capacity_hours": round(cap_hours, 2), "utilization": round(util, 4) if util is not None else None})
        operator_rows.sort(key=lambda x: (-(x.get("hours") or 0.0), x.get("operator_id") or ""))

        return {
            "version": v,
            "start_date": sd.isoformat(),
            "end_date": ed.isoformat(),
            "capacity_hours_per_resource": round(cap_hours, 2),
            "machines": machine_rows,
            "operators": operator_rows,
        }

    def export_utilization_xlsx(self, version: int, start_date: Any, end_date: Any) -> ReportExport:
        rep = self.utilization(version, start_date, end_date)
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "machines"
        ws1.append(["设备编号", "设备名称", "负荷(小时)", "任务数", "可用工时(小时)", "利用率"])
        for r in rep["machines"]:
            ws1.append(
                [
                    r.get("machine_id"),
                    r.get("machine_name"),
                    r.get("hours"),
                    r.get("task_count"),
                    r.get("capacity_hours"),
                    r.get("utilization"),
                ]
            )

        ws2 = wb.create_sheet("operators")
        ws2.append(["工号", "姓名", "负荷(小时)", "任务数", "可用工时(小时)", "利用率"])
        for r in rep["operators"]:
            ws2.append(
                [
                    r.get("operator_id"),
                    r.get("operator_name"),
                    r.get("hours"),
                    r.get("task_count"),
                    r.get("capacity_hours"),
                    r.get("utilization"),
                ]
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return ReportExport(
            filename=f"资源负荷与利用率_v{int(rep['version'])}_{rep['start_date']}_to_{rep['end_date']}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=buf,
        )

    # -------------------------
    # 3) 停机影响统计
    # -------------------------
    def downtime_impact(self, version: int, start_date: Any, end_date: Any) -> Dict[str, Any]:
        v = int(version or 0)
        sd = _parse_date(start_date, field="start_date")
        ed = _parse_date(end_date, field="end_date")
        if ed < sd:
            raise ValidationError("结束日期不能早于开始日期", field="end_date")

        start_dt = datetime(sd.year, sd.month, sd.day, 0, 0, 0)
        end_dt_excl = datetime(ed.year, ed.month, ed.day, 0, 0, 0) + timedelta(days=1)
        start_s = start_dt.strftime("%Y-%m-%d %H:%M:%S")
        end_s = end_dt_excl.strftime("%Y-%m-%d %H:%M:%S")

        # 停机区间
        dt_rows = self.conn.execute(
            """
            SELECT md.machine_id, m.name AS machine_name, md.start_time, md.end_time, md.reason_code, md.reason_detail
            FROM MachineDowntimes md
            LEFT JOIN Machines m ON m.machine_id = md.machine_id
            WHERE md.status = 'active'
              AND md.start_time < ?
              AND md.end_time > ?
            ORDER BY md.machine_id, md.start_time, md.id
            """,
            (end_s, start_s),
        ).fetchall()

        by_machine_dt: Dict[str, List[Tuple[datetime, datetime, str, str]]] = {}
        for r in dt_rows:
            mc = str(r["machine_id"] or "").strip()
            if not mc:
                continue
            s_dt = _parse_dt(r["start_time"])
            e_dt = _parse_dt(r["end_time"])
            if not s_dt or not e_dt:
                continue
            by_machine_dt.setdefault(mc, []).append((s_dt, e_dt, str(r["reason_code"] or ""), str(r["reason_detail"] or "")))

        # 排程区间（用于检查是否有“与停机重叠”的任务）
        sch_rows = self.schedule_repo.list_overlapping_with_details(start_s, end_s, v)
        by_machine_sch: Dict[str, List[Tuple[datetime, datetime]]] = {}
        for r in sch_rows:
            if (r.get("source") or "").strip() != "internal":
                continue
            mc = str(r.get("machine_id") or "").strip()
            if not mc:
                continue
            s_dt = _parse_dt(r.get("start_time"))
            e_dt = _parse_dt(r.get("end_time"))
            if not s_dt or not e_dt:
                continue
            by_machine_sch.setdefault(mc, []).append((s_dt, e_dt))

        items: List[Dict[str, Any]] = []
        for mc, dts in by_machine_dt.items():
            # 停机总时长（截断到范围内）
            downtime_sec = 0.0
            for ds, de, _, _ in dts:
                downtime_sec += _overlap_seconds(ds, de, start_dt, end_dt_excl)

            # 与排程重叠（理论上应为 0；用于发现“停机计划变更后未重排”的风险）
            overlap_sec = 0.0
            impact_count = 0
            segs = by_machine_sch.get(mc) or []
            for ds, de, _, _ in dts:
                for ss, se in segs:
                    sec = _overlap_seconds(ds, de, ss, se)
                    if sec > 0:
                        overlap_sec += sec
                        impact_count += 1

            # 设备名称（从第一条取）
            name = None
            for r in dt_rows:
                if str(r["machine_id"] or "").strip() == mc:
                    name = r["machine_name"]
                    break

            items.append(
                {
                    "machine_id": mc,
                    "machine_name": name,
                    "downtime_hours": round(downtime_sec / 3600.0, 2),
                    "downtime_count": len(dts),
                    "schedule_overlap_hours": round(overlap_sec / 3600.0, 2),
                    "schedule_overlap_count": int(impact_count),
                }
            )

        items.sort(key=lambda x: (-(x.get("downtime_hours") or 0.0), x.get("machine_id") or ""))

        return {
            "version": v,
            "start_date": sd.isoformat(),
            "end_date": ed.isoformat(),
            "machines": items,
        }

    def export_downtime_impact_xlsx(self, version: int, start_date: Any, end_date: Any) -> ReportExport:
        rep = self.downtime_impact(version, start_date, end_date)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "downtime"
        ws.append(["设备编号", "设备名称", "停机时长(小时)", "停机次数", "与排程重叠(小时)", "重叠次数"])
        for r in rep["machines"]:
            ws.append(
                [
                    r.get("machine_id"),
                    r.get("machine_name"),
                    r.get("downtime_hours"),
                    r.get("downtime_count"),
                    r.get("schedule_overlap_hours"),
                    r.get("schedule_overlap_count"),
                ]
            )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return ReportExport(
            filename=f"停机影响统计_v{int(rep['version'])}_{rep['start_date']}_to_{rep['end_date']}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=buf,
        )

