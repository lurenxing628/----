from __future__ import annotations

import io
import tempfile
from typing import Any, BinaryIO, Dict, List, cast

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.services.common.excel_templates import _sanitize_export_cell


def _auto_width(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(text))
    for col_idx, content_width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(content_width + 2, 12), 36)


def _append_row(ws, values: List[Any]) -> None:
    ws.append([_sanitize_export_cell(v) for v in values])


def _append_write_only_row(ws, values: List[Any], *, is_header: bool = False) -> None:
    row_cells: List[Any] = []
    for value in values:
        cell = WriteOnlyCell(ws, value=_sanitize_export_cell(value))
        if is_header:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_cells.append(cell)
    ws.append(row_cells)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _auto_width(ws)


def _make_output_buffer(*, write_only: bool) -> BinaryIO:
    if write_only:
        return cast(BinaryIO, tempfile.SpooledTemporaryFile(max_size=4 * 1024 * 1024, mode="w+b"))
    return io.BytesIO()


def _utilization_percent(value: Any) -> Any:
    if value is None:
        return None
    try:
        return round(float(value) * 100.0, 2)
    except Exception:
        return value


def export_overdue_xlsx(items: List[Dict[str, Any]], *, write_only: bool = False) -> BinaryIO:
    wb = openpyxl.Workbook(write_only=write_only)
    try:
        ws = wb.create_sheet("超期清单") if write_only else wb.active
        if ws is None:
            raise RuntimeError("无法创建超期清单工作表")
        if not write_only:
            ws.title = "超期清单"
            _append_row(ws, ["类别", "批次号", "图号", "名称", "数量", "交期", "完工/截至时间", "超期(天)", "超期(小时)"])
            for it in items:
                _append_row(
                    ws,
                    [
                        it.get("bucket_label"),
                        it.get("batch_id"),
                        it.get("part_no"),
                        it.get("part_name"),
                        it.get("quantity"),
                        it.get("due_date"),
                        it.get("finish_time") or it.get("as_of_time"),
                        it.get("delay_days"),
                        it.get("delay_hours"),
                    ],
                )
            _format_sheet(ws)
        else:
            _append_write_only_row(
                ws,
                ["类别", "批次号", "图号", "名称", "数量", "交期", "完工/截至时间", "超期(天)", "超期(小时)"],
                is_header=True,
            )
            for it in items:
                _append_write_only_row(
                    ws,
                    [
                        it.get("bucket_label"),
                        it.get("batch_id"),
                        it.get("part_no"),
                        it.get("part_name"),
                        it.get("quantity"),
                        it.get("due_date"),
                        it.get("finish_time") or it.get("as_of_time"),
                        it.get("delay_days"),
                        it.get("delay_hours"),
                    ],
                )

        buf = _make_output_buffer(write_only=write_only)
        wb.save(buf)
        buf.seek(0)
        return buf
    finally:
        try:
            wb.close()
        except Exception:
            pass


def export_utilization_xlsx(
    machines: List[Dict[str, Any]],
    operators: List[Dict[str, Any]],
    *,
    write_only: bool = False,
) -> BinaryIO:
    wb = openpyxl.Workbook(write_only=write_only)
    try:
        if write_only:
            ws1 = wb.create_sheet("设备负荷")
            _append_write_only_row(ws1, ["设备编号", "设备名称", "负荷(小时)", "任务数", "可用工时(小时)", "利用率(%)"], is_header=True)
            for r in machines:
                _append_write_only_row(
                    ws1,
                    [
                        r.get("machine_id"),
                        r.get("machine_name"),
                        r.get("hours"),
                        r.get("task_count"),
                        r.get("capacity_hours"),
                        _utilization_percent(r.get("utilization")),
                    ],
                )

            ws2 = wb.create_sheet("人员负荷")
            _append_write_only_row(ws2, ["工号", "姓名", "负荷(小时)", "任务数", "可用工时(小时)", "利用率(%)"], is_header=True)
            for r in operators:
                _append_write_only_row(
                    ws2,
                    [
                        r.get("operator_id"),
                        r.get("operator_name"),
                        r.get("hours"),
                        r.get("task_count"),
                        r.get("capacity_hours"),
                        _utilization_percent(r.get("utilization")),
                    ],
                )
        else:
            ws1 = wb.active
            if ws1 is None:
                raise RuntimeError("无法创建设备负荷工作表")
            ws1.title = "设备负荷"
            _append_row(ws1, ["设备编号", "设备名称", "负荷(小时)", "任务数", "可用工时(小时)", "利用率(%)"])
            for r in machines:
                _append_row(
                    ws1,
                    [
                        r.get("machine_id"),
                        r.get("machine_name"),
                        r.get("hours"),
                        r.get("task_count"),
                        r.get("capacity_hours"),
                        _utilization_percent(r.get("utilization")),
                    ],
                )
            _format_sheet(ws1)

            ws2 = wb.create_sheet("人员负荷")
            _append_row(ws2, ["工号", "姓名", "负荷(小时)", "任务数", "可用工时(小时)", "利用率(%)"])
            for r in operators:
                _append_row(
                    ws2,
                    [
                        r.get("operator_id"),
                        r.get("operator_name"),
                        r.get("hours"),
                        r.get("task_count"),
                        r.get("capacity_hours"),
                        _utilization_percent(r.get("utilization")),
                    ],
                )
            _format_sheet(ws2)

        buf = _make_output_buffer(write_only=write_only)
        wb.save(buf)
        buf.seek(0)
        return buf
    finally:
        try:
            wb.close()
        except Exception:
            pass


def export_downtime_impact_xlsx(machines: List[Dict[str, Any]], *, write_only: bool = False) -> BinaryIO:
    wb = openpyxl.Workbook(write_only=write_only)
    try:
        ws = wb.create_sheet("停机影响") if write_only else wb.active
        if ws is None:
            raise RuntimeError("无法创建停机影响工作表")
        if not write_only:
            ws.title = "停机影响"
            _append_row(ws, ["设备编号", "设备名称", "停机时长(小时)", "停机次数", "与排程重叠(小时)", "重叠次数"])
            for r in machines:
                _append_row(
                    ws,
                    [
                        r.get("machine_id"),
                        r.get("machine_name"),
                        r.get("downtime_hours"),
                        r.get("downtime_count"),
                        r.get("schedule_overlap_hours"),
                        r.get("schedule_overlap_count"),
                    ],
                )
            _format_sheet(ws)
        else:
            _append_write_only_row(
                ws,
                ["设备编号", "设备名称", "停机时长(小时)", "停机次数", "与排程重叠(小时)", "重叠次数"],
                is_header=True,
            )
            for r in machines:
                _append_write_only_row(
                    ws,
                    [
                        r.get("machine_id"),
                        r.get("machine_name"),
                        r.get("downtime_hours"),
                        r.get("downtime_count"),
                        r.get("schedule_overlap_hours"),
                        r.get("schedule_overlap_count"),
                    ],
                )

        buf = _make_output_buffer(write_only=write_only)
        wb.save(buf)
        buf.seek(0)
        return buf
    finally:
        try:
            wb.close()
        except Exception:
            pass
